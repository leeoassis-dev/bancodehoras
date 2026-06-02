from flask import Flask, render_template, request, redirect, url_for, flash, make_response, jsonify, session, g
from database import init_db, get_db, _consumir_fifo_raw, six_months_ago, five_months_ago, IS_POSTGRES
from datetime import datetime, date, timedelta
from urllib.parse import urlencode
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import os, json, io, csv, secrets, string, smtplib, time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from database import six_months_ago, five_months_ago

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "ibipora_banco_horas_2024_seguro")
app.url_map.strict_slashes = False
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("MAX_UPLOAD_MB", "8")) * 1024 * 1024

LIMITE_PAGAMENTO_MINUTOS = 45 * 60
MESES_PT = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
MESES_FULL = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
              "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

_CACHE = {}

def cache_get(chave):
    item = _CACHE.get(chave)
    if not item:
        return None
    expira_em, valor = item
    if expira_em < time.time():
        _CACHE.pop(chave, None)
        return None
    return valor

def cache_set(chave, valor, ttl=30):
    _CACHE[chave] = (time.time() + ttl, valor)
    return valor

def limpar_cache():
    _CACHE.clear()

# â”€â”€â”€ Auth helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

ROTAS_PUBLICAS = {'login','logout','recuperar_senha','recuperar_senha_token','setup','static'}

@app.errorhandler(404)
def pagina_nao_encontrada(e):
    """Evita 404 por pequenas variações de URL em produção."""
    caminho = (request.path or "").rstrip("/")
    rotas = {
        "/login": "login",
        "/recuperar-senha": "recuperar_senha",
        "/criar-conta": "criar_conta",
        "/portal": "portal",
        "/": "portal",
    }
    if caminho in rotas:
        return redirect(url_for(rotas[caminho]))
    return render_template("404.html"), 404

@app.errorhandler(413)
def arquivo_muito_grande(e):
    flash("Arquivo muito grande. Envie um CSV de até 8 MB.", "danger")
    return redirect(request.referrer or url_for("portal"))

def gerar_senha_temp(n=10):
    chars = string.ascii_letters + string.digits + "!@#$"
    return ''.join(secrets.choice(chars) for _ in range(n))

def obter_config_smtp():
    """Lê SMTP do banco e, se faltar, usa variáveis de ambiente do Render."""
    db = get_db()
    cfg_db = {r['chave']: r['valor'] for r in db.execute("SELECT * FROM config").fetchall()}
    cfg = {
        'smtp_host': cfg_db.get('smtp_host') or os.environ.get('SMTP_HOST', ''),
        'smtp_port': cfg_db.get('smtp_port') or os.environ.get('SMTP_PORT', '587'),
        'smtp_user': cfg_db.get('smtp_user') or os.environ.get('SMTP_USER', ''),
        'smtp_pass': cfg_db.get('smtp_pass') or os.environ.get('SMTP_PASS', ''),
        'smtp_from': cfg_db.get('smtp_from') or os.environ.get('SMTP_FROM', ''),
        'smtp_tls': cfg_db.get('smtp_tls') or os.environ.get('SMTP_TLS', 'true'),
    }
    cfg['smtp_from'] = cfg['smtp_from'] or cfg['smtp_user']
    cfg['_configurado'] = bool(cfg['smtp_host'] and cfg['smtp_user'] and cfg['smtp_pass'])
    cfg['_origem'] = 'Render/env' if os.environ.get('SMTP_HOST') and not cfg_db.get('smtp_host') else 'Banco de dados'
    return cfg

def enviar_email_smtp(para, assunto, corpo_html):
    """Envia e-mail via SMTP configurado no banco. Retorna (ok, msg)."""
    try:
        cfg = obter_config_smtp()
        host = cfg.get('smtp_host','')
        porta = int(cfg.get('smtp_port','587') or 587)
        user = cfg.get('smtp_user','')
        pwd  = cfg.get('smtp_pass','')
        de   = cfg.get('smtp_from', user)
        tls  = str(cfg.get('smtp_tls','true')).lower() not in ('0','false','nao','não','no')
        if not host or not user or not pwd:
            return False, "SMTP_NAO_CONFIGURADO"
        msg = MIMEMultipart('alternative')
        msg['Subject'] = assunto; msg['From'] = de; msg['To'] = para
        msg.attach(MIMEText(corpo_html, 'html', 'utf-8'))
        smtp_cls = smtplib.SMTP_SSL if porta == 465 else smtplib.SMTP
        with smtp_cls(host, porta, timeout=20) as srv:
            srv.ehlo()
            if tls and porta != 465:
                srv.starttls()
                srv.ehlo()
            srv.login(user, pwd)
            srv.sendmail(de, para, msg.as_string())
        return True, ""
    except Exception as e:
        return False, str(e)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'uid' not in session:
            flash("Faça login para continuar.", "warning")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def master_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'uid' not in session:
            flash("Faça login para continuar.", "warning")
            return redirect(url_for('login'))
        if session.get('nivel') != 'master':
            try:
                registrar_auditoria(
                    get_db(), "Acesso negado", "seguranca", request.endpoint or request.path,
                    session.get("matricula"), session.get("nome"),
                    f"Tentativa de acesso restrito ao RH: {request.path}"
                )
                get_db().commit()
            except Exception:
                pass
            flash("Acesso restrito ao RH.", "danger")
            return redirect(url_for('portal'))
        return f(*args, **kwargs)
    return decorated

@app.context_processor
def injetar_usuario():
    ctx = {
        'u_nivel': session.get('nivel'),
        'u_nome':  session.get('nome'),
        'u_cpf':   session.get('cpf'),
        'pendentes_tarefas': 0,
        'pendentes_aprovacao': 0,
        'u_mat_servidor': None,
        'u_sol_habilitado': False,
    }
    uid   = session.get('uid')
    nivel = session.get('nivel')
    if uid:
        try:
            db = get_db()
            # Matricula do servidor vinculado ao usuário (qualquer nível)
            u_row = db.execute("SELECT matricula FROM usuarios WHERE id=?", (uid,)).fetchone()
            if u_row and u_row['matricula']:
                srv_row = db.execute(
                    "SELECT matricula FROM servidores WHERE matricula=? AND arquivado=0",
                    (u_row['matricula'],)).fetchone()
                ctx['u_mat_servidor'] = srv_row['matricula'] if srv_row else None
            elif nivel == 'servidor':
                ctx['u_mat_servidor'] = session.get('matricula')
            if ctx['u_mat_servidor']:
                ctx['u_sol_habilitado'] = _solicitacoes_habilitado(db)
            if nivel == 'master':
                ctx['pendentes_tarefas'] = db.execute(
                    "SELECT COUNT(*) FROM solicitacoes WHERE status='autorizado'"
                ).fetchone()[0] or 0
            if nivel in ('chefia', 'secretario'):
                vinculos = session.get('vinculos') or []
                if vinculos:
                    ph = ','.join('?' * len(vinculos))
                    campo = 'secretaria' if nivel == 'secretario' else 'setor'
                    ctx['pendentes_aprovacao'] = db.execute(
                        f"""SELECT COUNT(*) FROM solicitacoes sol
                            JOIN servidores s ON s.matricula=sol.matricula
                            WHERE sol.status='solicitado' AND s.{campo} IN ({ph})
                            AND sol.criado_por_uid != ?""",
                        vinculos + [uid]
                    ).fetchone()[0] or 0
        except Exception:
            pass
    return ctx

@app.before_request
def verificar_acesso():
    if request.endpoint in ROTAS_PUBLICAS or request.endpoint is None:
        return None

    if 'uid' not in session:
        # Se não há usuários, redireciona para setup
        db = get_db()
        if db.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0] == 0:
            return redirect(url_for('setup'))
        flash("Faça login para continuar.", "warning")
        return redirect(url_for('login'))

    nivel = session.get('nivel')

    # Senha temporária → forçar troca
    if session.get('temp') and request.endpoint != 'trocar_senha':
        flash("Sua senha é temporária. Defina uma nova senha para continuar.", "warning")
        return redirect(url_for('trocar_senha'))

    # Servidor → meu_banco + seus próprios dados de eleição + solicitações
    if nivel == 'servidor' and request.endpoint not in (
            'meu_banco', 'meu_cadastro', 'trocar_senha', 'logout',
            'eleicao_servidor', 'eleicao_exportar',
            'solicitacoes_nova', 'solicitacoes_cancelar', 'api_saldo_solicitacao'):
        try:
            registrar_auditoria(get_db(), "Acesso negado", "seguranca", request.endpoint or request.path,
                                session.get("matricula"), session.get("nome"),
                                f"Servidor tentou acessar {request.path}")
            get_db().commit()
        except Exception:
            pass
        return redirect(url_for('meu_banco'))

    # Secretário / Chefia → consulta + api_historico + eleição + solicitações
    if nivel in ('secretario', 'chefia') and request.endpoint not in (
            'portal', 'consulta', 'api_historico', 'meu_cadastro', 'trocar_senha', 'logout',
            'eleicao_index', 'eleicao_servidor', 'eleicao_exportar',
            'solicitacoes_nova', 'solicitacoes_autorizar', 'solicitacoes_indeferir',
            'solicitacoes_cancelar', 'api_saldo_solicitacao'):
        try:
            registrar_auditoria(get_db(), "Acesso negado", "seguranca", request.endpoint or request.path,
                                session.get("matricula"), session.get("nome"),
                                f"{nivel} tentou acessar {request.path}")
            get_db().commit()
        except Exception:
            pass
        return redirect(url_for('consulta'))

    # Master: bloqueia apenas se tentar acessar rota de outro nível
    # (master tem acesso total, nada a bloquear)
    registrar_visualizacao()

@app.after_request
def invalidar_cache_em_gravacao(response):
    if request.method in ("POST", "PUT", "PATCH", "DELETE"):
        limpar_cache()
    return response

# â”€â”€â”€ Utilitários â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def horas_para_minutos(s):
    try:
        h, m = map(int, str(s).split(":")); return h*60+m
    except: return 0

def minutos_para_horas(m):
    m = int(m or 0)
    s = "-" if m < 0 else ""; m = abs(m)
    return f"{s}{m//60:02d}:{m%60:02d}"

def minutos_num(m):
    return int(m or 0)

@app.template_filter('fmt_data')
def fmt_data(valor):
    """YYYY-MM-DD → dd/mm/aaaa para exibição."""
    if not valor:
        return '–'
    s = str(valor).strip()[:10]
    try:
        return date.fromisoformat(s).strftime('%d/%m/%Y')
    except Exception:
        return s

@app.template_filter('fmt_datetime')
def fmt_datetime(valor):
    """YYYY-MM-DD HH:MM[:SS] → dd/mm/aaaa HH:MM para exibição."""
    if not valor:
        return '–'
    s = str(valor).strip()
    try:
        if len(s) >= 16:
            return datetime.strptime(s[:16], '%Y-%m-%d %H:%M').strftime('%d/%m/%Y %H:%M')
        return date.fromisoformat(s[:10]).strftime('%d/%m/%Y')
    except Exception:
        return s

def calcular_saldo_eleicao(db, matricula):
    creditos = db.execute(
        "SELECT COALESCE(SUM(quantidade_dias),0) FROM eleicao_creditos WHERE matricula=?",
        (matricula,)).fetchone()[0]
    baixas = db.execute(
        "SELECT COUNT(*) FROM eleicao_baixas WHERE matricula=?",
        (matricula,)).fetchone()[0]
    pendentes = db.execute(
        "SELECT COALESCE(SUM(quantidade),0) FROM solicitacoes WHERE matricula=? AND tipo='eleicao' AND status IN ('solicitado','autorizado')",
        (matricula,)).fetchone()[0]
    return int(creditos or 0) - int(baixas or 0) - int(pendentes or 0)

def registrar_auditoria(db, acao, entidade, entidade_id=None, matricula=None, servidor_nome=None, detalhe=""):
    return db.insert("""
        INSERT INTO auditoria
            (criado_em, usuario_id, usuario_nome, usuario_cpf, acao, entidade, entidade_id, matricula, servidor_nome, detalhe)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        session.get("uid"), session.get("nome"), session.get("cpf"),
        acao, entidade, str(entidade_id or ""), matricula, servidor_nome, detalhe
    ))

def _rows_dict(rows):
    return [dict(r) for r in rows]

def _snapshot_servidor_exclusao(db, matricula):
    lancs = _rows_dict(db.execute("SELECT * FROM lancamentos WHERE matricula=? ORDER BY id", (matricula,)).fetchall())
    comps = _rows_dict(db.execute("SELECT * FROM compensacoes WHERE matricula=? ORDER BY id", (matricula,)).fetchall())
    pags = _rows_dict(db.execute("SELECT * FROM pagamentos WHERE matricula=? ORDER BY id", (matricula,)).fetchall())
    lanc_ids = [l["id"] for l in lancs]
    comp_ids = [c["id"] for c in comps]
    pag_ids = [p["id"] for p in pags]
    consumos = []
    if lanc_ids:
        ph = ",".join("?" * len(lanc_ids))
        consumos += _rows_dict(db.execute(f"SELECT * FROM consumos WHERE lancamento_id IN ({ph})", lanc_ids).fetchall())
    if comp_ids:
        ph = ",".join("?" * len(comp_ids))
        consumos += _rows_dict(db.execute(f"SELECT * FROM consumos WHERE tipo='compensacao' AND referencia_id IN ({ph})", comp_ids).fetchall())
    if pag_ids:
        ph = ",".join("?" * len(pag_ids))
        consumos += _rows_dict(db.execute(f"SELECT * FROM consumos WHERE tipo='pagamento' AND referencia_id IN ({ph})", pag_ids).fetchall())
    return {"lancamentos": lancs, "compensacoes": comps, "pagamentos": pags, "consumos": consumos}

def _snapshot_servidor_cadastro(srv):
    if not srv:
        return None
    return {
        "matricula": srv["matricula"], "nome": srv["nome"], "cpf": srv["cpf"], "email": srv["email"],
        "cargo": srv["cargo"], "secretaria": srv["secretaria"], "setor": srv["setor"],
        "funcao_gratificada": srv["funcao_gratificada"], "arquivado": srv["arquivado"],
    }

def _valor_csv(row, *nomes):
    normalizado = {str(k).strip().lower(): (v or "").strip() for k, v in row.items()}
    for nome in nomes:
        if nome.lower() in normalizado:
            return normalizado[nome.lower()]
    return ""

def registrar_visualizacao():
    if request.method != "GET" or "uid" not in session:
        return
    if request.endpoint in ROTAS_PUBLICAS or request.endpoint == "static":
        return
    if (request.path or "").startswith("/api/"):
        return
    titulos = {
        "dashboard": "Dashboard",
        "servidores": "Servidores",
        "arquivados": "Arquivados",
        "pagamentos_index": "Pagamentos",
        "relatorios": "Relatórios",
        "admin_usuarios": "Usuários",
        "admin_acessos": "Permissões",
        "admin_auditoria": "Auditoria",
        "consulta": "Consulta",
        "meu_banco": "Meu Banco",
        "eleicao_index": "Dias de Eleição",
        "eleicao_servidor": "Dias de Eleição – Servidor",
        "admin_tarefas": "Tarefas",
    }
    try:
        db = get_db()
        db.execute("""
            INSERT INTO visualizacoes (usuario_id,usuario_nome,usuario_cpf,endpoint,caminho,titulo,criado_em)
            VALUES (?,?,?,?,?,?,?)
        """, (
            session.get("uid"), session.get("nome"), session.get("cpf"),
            request.endpoint or "", request.full_path.rstrip("?"),
            titulos.get(request.endpoint, request.path),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))
        db.commit()
    except Exception:
        pass

def _csv_response(filename, headers, rows):
    def safe_csv(v):
        txt = "" if v is None else str(v)
        return "'" + txt if txt[:1] in ("=", "+", "-", "@") else txt
    out = io.StringIO()
    w = csv.writer(out, delimiter=";")
    w.writerow([safe_csv(h) for h in headers])
    w.writerows([[safe_csv(v) for v in row] for row in rows])
    r = make_response("\ufeff" + out.getvalue())
    r.headers["Content-Type"] = "text/csv; charset=utf-8"
    r.headers["Content-Disposition"] = f"attachment; filename={filename}.csv"
    return r

def _validar_csv_upload(arquivo):
    nome = (arquivo.filename or "").lower()
    if not nome.endswith(".csv"):
        return False, "Arquivo inválido. Envie somente arquivo CSV."
    return True, ""

def _csv_reader_upload(arquivo):
    ok, msg = _validar_csv_upload(arquivo)
    if not ok:
        raise ValueError(msg)
    conteudo = arquivo.read().decode("utf-8-sig", errors="replace")
    amostra = conteudo[:2048]
    try:
        dialect = csv.Sniffer().sniff(amostra, delimiters=";,")
    except Exception:
        dialect = csv.excel
    return csv.DictReader(io.StringIO(conteudo), dialect=dialect)

def _safe_excel(v):
    if not isinstance(v, str):
        return v
    txt = "" if v is None else str(v)
    return "'" + txt if txt[:1] in ("=", "+", "-", "@") else v

def _registrar_relatorio_importacao(tipo, arquivo, total, sucessos, erros, payload=None, atualizados=0):
    """Persiste histórico da importação e prepara relatório visual pós-importação."""
    db = get_db()
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    payload = payload or []
    imp_id = db.insert("""INSERT INTO importacoes
        (tipo,arquivo,usuario_id,usuario_nome,usuario_cpf,total_linhas,criados,atualizados,erros,payload,criado_em)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (tipo, arquivo, session.get("uid"), session.get("nome"), session.get("cpf"),
         total, len(sucessos), atualizados, len(erros), json.dumps(payload, ensure_ascii=False), agora))
    registrar_auditoria(db, f"Importou {tipo}", "importacao", imp_id, None, None,
                        f"Arquivo {arquivo}; importados {len(sucessos)}; erros {len(erros)}.")
    db.commit()
    session["import_report"] = {
        "tipo": tipo.replace("_", " ").title(),
        "arquivo": arquivo,
        "total": total,
        "sucessos_qtd": len(sucessos),
        "erros_qtd": len(erros),
        "sucessos": sucessos[:30],
        "erros": erros[:30],
        "data_hora": agora,
        "usuario": session.get("nome") or "-",
    }
    return imp_id

def _erro_importacao(erros, linha, identificador, motivo):
    erros.append({"linha": linha, "identificador": identificador or "-", "motivo": motivo})

def _sucesso_importacao(sucessos, linha, identificador, detalhe):
    sucessos.append({"linha": linha, "identificador": identificador or "-", "detalhe": detalhe})

def _xlsx_response(filename, title, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    cell = ws.cell(1, 1, title)
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.fill = PatternFill("solid", fgColor="1A3A6B")
    cell.alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(headers)
    for c in ws[3]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A3A6B")
        c.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([_safe_excel(v) for v in row])
        label = str(row[0] if row else "")
        if label.startswith("TOTAL DO GRUPO") or label.startswith("TOTAL GERAL"):
            fill = "D9EAF7" if label.startswith("TOTAL DO GRUPO") else "1A3A6B"
            font_color = "000000" if label.startswith("TOTAL DO GRUPO") else "FFFFFF"
            for c in ws[ws.max_row]:
                c.font = Font(bold=True, color=font_color)
                c.fill = PatternFill("solid", fgColor=fill)
                c.alignment = Alignment(horizontal="center" if c.column >= 5 else "left")
    for col in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 45)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = make_response(buf.getvalue())
    r.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    r.headers["Content-Disposition"] = f"attachment; filename={filename}.xlsx"
    return r

def _pdf_response(filename, title, headers, rows):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from xml.sax.saxutils import escape

    buf = io.BytesIO()
    page_size = landscape(A3 if len(headers) >= 10 else A4)
    doc = SimpleDocTemplate(buf, pagesize=page_size, rightMargin=14, leftMargin=14, topMargin=18, bottomMargin=18)
    styles = getSampleStyleSheet()
    cell_font = 6 if len(headers) >= 10 else 7
    header_font = 6 if len(headers) >= 10 else 7
    cell_style = ParagraphStyle("CellWrap", parent=styles["BodyText"], fontSize=cell_font, leading=cell_font + 2, wordWrap="CJK")
    header_style = ParagraphStyle("HeaderWrap", parent=cell_style, fontName="Helvetica-Bold", textColor=colors.white, alignment=1)
    title_style = ParagraphStyle("TitleBlue", parent=styles["Title"], fontSize=16, leading=19, alignment=1)
    story = [Paragraph(f"<b>{escape(title)}</b>", title_style), Spacer(1, 10)]

    def pcell(value, style=cell_style):
        return Paragraph(escape(str(value or "")), style)

    table_data = [[pcell(h, header_style) for h in headers]] + [[pcell(v) for v in row] for row in rows]
    weights = []
    for idx, h in enumerate(headers):
        sample = max([len(str(h))] + [len(str(row[idx] if idx < len(row) else "")) for row in rows[:80]])
        weights.append(max(7, min(sample, 24)))
    total_weight = sum(weights) or 1
    col_widths = [doc.width * w / total_weight for w in weights]
    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    estilos = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3A6B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), header_font),
        ("FONTSIZE", (0, 1), (-1, -1), cell_font),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D7DE")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F9")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for idx, row in enumerate(rows, start=1):
        label = str(row[0] if row else "")
        if label.startswith("TOTAL DO GRUPO"):
            estilos.extend([
                ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#D9EAF7")),
                ("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold"),
            ])
        elif label.startswith("TOTAL GERAL"):
            estilos.extend([
                ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#1A3A6B")),
                ("TEXTCOLOR", (0, idx), (-1, idx), colors.white),
                ("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold"),
            ])
    table.setStyle(TableStyle(estilos))
    story.append(table)
    doc.build(story)
    buf.seek(0)
    r = make_response(buf.getvalue())
    r.headers["Content-Type"] = "application/pdf"
    r.headers["Content-Disposition"] = f"attachment; filename={filename}.pdf"
    return r

def _export_response(fmt_out, filename, title, headers, rows):
    if fmt_out == "csv":
        return _csv_response(filename, headers, rows)
    if fmt_out == "xlsx":
        return _xlsx_response(filename, title, headers, rows)
    if fmt_out == "pdf":
        return _pdf_response(filename, title, headers, rows)
    return None

def _grupo_relatorio(row, agr):
    if agr == "secretaria":
        return row.get("secretaria") or "Sem Secretaria"
    if agr == "departamento":
        return row.get("setor") or "Sem Departamento"
    if agr == "cargo":
        return row.get("cargo") or "Sem Cargo"
    nome = row.get("nome") or row.get("servidor_nome") or "Servidor"
    matricula = row.get("matricula") or row.get("mat") or ""
    return f"{nome} ({matricula})" if matricula else nome

def _agrupar_itens(itens, agr):
    grupos = {}
    for item in itens:
        d = dict(item)
        grupos.setdefault(_grupo_relatorio(d, agr), []).append(d)
    return grupos

def _somar_meses_iso(data_iso, meses):
    try:
        d = datetime.strptime(data_iso, "%Y-%m-%d").date()
        mes = d.month - 1 + meses
        ano = d.year + mes // 12
        mes = mes % 12 + 1
        ultimo = [31, 29 if ano % 4 == 0 and (ano % 100 != 0 or ano % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][mes - 1]
        return date(ano, mes, min(d.day, ultimo)).isoformat()
    except Exception:
        return data_iso

def calcular_saldo(db, matricula):
    c = db.execute("SELECT COALESCE(SUM(minutos_creditados),0) FROM lancamentos WHERE matricula=?", (matricula,)).fetchone()[0]
    u = db.execute("""SELECT COALESCE(SUM(c.minutos),0) FROM consumos c
                      JOIN lancamentos l ON l.id=c.lancamento_id WHERE l.matricula=?""", (matricula,)).fetchone()[0]
    p = db.execute(
        "SELECT COALESCE(SUM(quantidade),0) FROM solicitacoes WHERE matricula=? AND tipo='banco_horas' AND status IN ('solicitado','autorizado')",
        (matricula,)).fetchone()[0]
    return c - u - int(p or 0)

def lancamentos_com_saldo(db, matricula, apenas_vencidos=False):
    seis_meses = six_months_ago()
    f = "AND l.data <= ?" if apenas_vencidos else ""
    params = [matricula, seis_meses] if apenas_vencidos else [matricula]
    rows = db.execute(f"""
        SELECT l.*,
               COALESCE((SELECT SUM(c.minutos) FROM consumos c WHERE c.lancamento_id=l.id),0) AS consumido,
               CASE WHEN l.data<=? THEN 1 ELSE 0 END AS vencido
        FROM lancamentos l WHERE l.matricula=? {f} ORDER BY l.data ASC, l.id ASC
    """, [seis_meses] + params).fetchall()
    out = []
    for l in rows:
        sm = l["minutos_creditados"] - l["consumido"]
        if sm <= 0: continue
        mb = l["minutos_base"] if l["minutos_base"] else horas_para_minutos(l["horas_base"])
        sb = round(sm * mb / l["minutos_creditados"]) if l["minutos_creditados"] > 0 else 0
        out.append({**dict(l), "saldo_minutos": sm, "saldo_base_minutos": sb})
    return out

def _filtro_servidores(busca="", secretaria="", setor="", cargo="", arquivado=0):
    f = f"WHERE s.arquivado={arquivado}"
    p = []
    if busca:
        f += " AND (s.matricula LIKE ? OR s.nome LIKE ?)"; p += [f"%{busca}%", f"%{busca}%"]
    if secretaria:
        f += " AND s.secretaria LIKE ?"; p.append(f"%{secretaria}%")
    if setor:
        f += " AND s.setor LIKE ?"; p.append(f"%{setor}%")
    if cargo:
        f += " AND s.cargo LIKE ?"; p.append(f"%{cargo}%")
    return f, p

def _listas_filtro(db, arquivado=0):
    chave = f"listas_filtro:{arquivado}"
    cached = cache_get(chave)
    if cached is not None:
        return cached
    secs = _cadastros_nomes(db, "secretaria")
    sets = _cadastros_nomes(db, "departamento")
    if not secs:
        secs = [r[0] for r in db.execute(
            f"SELECT DISTINCT secretaria FROM servidores WHERE secretaria IS NOT NULL AND secretaria!='' AND arquivado={arquivado} ORDER BY secretaria").fetchall()]
    if not sets:
        sets = [r[0] for r in db.execute(
            f"SELECT DISTINCT setor FROM servidores WHERE setor IS NOT NULL AND setor!='' AND arquivado={arquivado} ORDER BY setor").fetchall()]
    return cache_set(chave, (secs, sets), ttl=60)

def _cadastros_nomes(db, tipo, somente_ativos=True):
    chave = f"cadastros:{tipo}:{1 if somente_ativos else 0}"
    cached = cache_get(chave)
    if cached is not None:
        return cached
    where = "WHERE tipo=?"
    params = [tipo]
    if somente_ativos:
        where += " AND ativo=1"
    rows = db.execute(
        f"SELECT nome FROM cadastros_auxiliares {where} ORDER BY nome",
        params
    ).fetchall()
    return cache_set(chave, [r["nome"] for r in rows], ttl=60)

def _opcoes_servidor_form(db):
    return {
        "secretarias": _cadastros_nomes(db, "secretaria"),
        "setores": _cadastros_nomes(db, "departamento"),
        "cargos": _cadastros_nomes(db, "cargo"),
    }

def _garantir_cadastro_auxiliar(db, tipo, nome):
    nome = (nome or "").strip()
    if not nome:
        return
    db.upsert(
        "INSERT OR IGNORE INTO cadastros_auxiliares (tipo,nome,ativo) VALUES (?,?,1)",
        """INSERT INTO cadastros_auxiliares (tipo,nome,ativo) VALUES (?,?,1)
           ON CONFLICT (tipo,nome) DO NOTHING""",
        (tipo, nome)
    )
    db.execute("UPDATE cadastros_auxiliares SET ativo=1 WHERE tipo=? AND nome=?", (tipo, nome))

def _fg(servidor): return bool(servidor["funcao_gratificada"])

def get_vinculos(obj):
    """Retorna lista de vínculos (secretarias/setores) de usuário ou pré-autorização."""
    if not obj:
        return []

    def valor(campo):
        try:
            return obj[campo]
        except Exception:
            try:
                return obj.get(campo)
            except Exception:
                return ""

    try:
        lst = json.loads(valor('vinculos') or '[]')
        if isinstance(lst, list) and lst:
            return [str(v).strip() for v in lst if str(v).strip()]
    except Exception:
        pass

    # Fallback para bases antigas que ainda usavam uma única secretaria/setor.
    v = (valor('secretaria') or valor('setor') or '').strip()
    return [v] if v else []

def _usuario_tem_servidor_ativo(db, usuario):
    """Usuários não-master só ficam ativos nas telas comuns se o servidor vinculado existir e estiver ativo."""
    if not usuario:
        return False
    try:
        nivel = usuario["nivel"]
        cpf = (usuario["cpf"] or "").strip()
        matricula = (usuario["matricula"] or "").strip()
    except Exception:
        nivel = usuario.get("nivel")
        cpf = (usuario.get("cpf") or "").strip()
        matricula = (usuario.get("matricula") or "").strip()

    if nivel == "master":
        return True
    if not cpf and not matricula:
        return False
    return bool(db.execute("""
        SELECT 1
        FROM servidores
        WHERE arquivado=0
          AND ((?<>'' AND cpf=?) OR (?<>'' AND matricula=?))
        LIMIT 1
    """, (cpf, cpf, matricula, matricula)).fetchone())

def _inativar_vinculos_servidor(db, servidor):
    """Inativa contas e liberações administrativas associadas a um servidor arquivado/excluído."""
    if not servidor:
        return 0
    matricula = (servidor["matricula"] or "").strip()
    cpf = (servidor["cpf"] or "").strip()
    params = [matricula, matricula, cpf, cpf]
    cur = db.execute("""
        UPDATE usuarios
        SET ativo=0
        WHERE ((?<>'' AND matricula=?) OR (?<>'' AND cpf=?))
          AND nivel<>'master'
    """, params)
    db.execute("""
        DELETE FROM pre_autorizacoes
        WHERE (?<>'' AND matricula=?) OR (?<>'' AND cpf=?)
    """, params)
    return cur.rowcount or 0

def usuario_pode_ver_matricula(db, matricula):
    """Valida acesso de leitura ao histórico conforme nível e vínculos da sessão."""
    nivel = session.get('nivel')
    if nivel == 'master':
        return True
    if session.get('matricula') == matricula:
        return bool(db.execute(
            "SELECT 1 FROM servidores WHERE matricula=? AND arquivado=0",
            (matricula,)
        ).fetchone())
    if nivel == 'servidor':
        return False
    if nivel in ('secretario', 'chefia'):
        vinculos = session.get('vinculos') or []
        if not vinculos:
            return False
        srv = db.execute(
            "SELECT secretaria,setor FROM servidores WHERE matricula=? AND arquivado=0",
            (matricula,)
        ).fetchone()
        if not srv:
            return False
        campo = srv['secretaria'] if nivel == 'secretario' else srv['setor']
        return campo in vinculos
    return False

def filtro_consulta_vinculos(nivel, vinculos, matricula_propria, alias='s'):
    """Retorna filtro para consulta: vínculos concedidos ou próprio servidor logado."""
    partes, params = [], []
    if nivel in ('secretario', 'chefia') and vinculos:
        ph = ','.join('?' * len(vinculos))
        campo = 'secretaria' if nivel == 'secretario' else 'setor'
        partes.append(f"{alias}.{campo} IN ({ph})")
        params.extend(vinculos)
    if matricula_propria:
        partes.append(f"{alias}.matricula=?")
        params.append(matricula_propria)
    if not partes:
        return " AND 1=0", []
    return " AND (" + " OR ".join(partes) + ")", params

def ultimos_n_meses(n=6):
    hoje = date.today()
    res = []
    for i in range(n-1, -1, -1):
        off = hoje.month - 1 - i
        ano = hoje.year + off // 12
        mes = off % 12 + 1
        res.append({"ano": ano, "mes": mes,
                    "label": f"{MESES_PT[mes-1]}/{str(ano)[2:]}",
                    "ym": f"{ano:04d}-{mes:02d}"})
    return res

# â”€â”€â”€ Dashboard â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/")
@master_required
def dashboard():
    db = get_db()
    cache_key = f"dashboard:{date.today().isoformat()}"
    dados_cache = cache_get(cache_key)
    if dados_cache is not None:
        return render_template("dashboard.html", **dados_cache, fmt=minutos_para_horas)

    meses6 = ultimos_n_meses(6)
    servidores_atalho = db.execute(
        "SELECT matricula,nome,secretaria,setor FROM servidores WHERE arquivado=0 ORDER BY nome"
    ).fetchall()

    # Lançamentos por mês (horas base)
    lanc_mes = []
    comp_mes = []
    pag_mes  = []
    for m in meses6:
        v = db.execute("""SELECT COALESCE(SUM(l.minutos_base),0) FROM lancamentos l
                          JOIN servidores s ON s.matricula=l.matricula
                          WHERE substr(l.data,1,7)=? AND s.arquivado=0""", (m["ym"],)).fetchone()[0]
        lanc_mes.append(round(minutos_num(v)/60, 1))
        v2 = db.execute("""SELECT COALESCE(SUM(c2.minutos_compensados),0) FROM compensacoes c2
                           JOIN servidores s ON s.matricula=c2.matricula
                           WHERE substr(c2.data,1,7)=? AND s.arquivado=0""", (m["ym"],)).fetchone()[0]
        comp_mes.append(round(minutos_num(v2)/60, 1))
        v3 = db.execute("""SELECT COALESCE(SUM(ROUND(c.minutos*l.minutos_base*1.0/l.minutos_creditados)),0)
                           FROM pagamentos p
                           JOIN consumos c ON c.referencia_id=p.id AND c.tipo='pagamento'
                           JOIN lancamentos l ON l.id=c.lancamento_id
                           JOIN servidores s ON s.matricula=p.matricula
                           WHERE substr(p.data_pagamento,1,7)=? AND s.arquivado=0""", (m["ym"],)).fetchone()[0]
        pag_mes.append(round(minutos_num(v3)/60, 1))

    # KPIs
    saldo_total = db.execute("""
        SELECT COALESCE(SUM(
            (SELECT COALESCE(SUM(minutos_creditados),0) FROM lancamentos WHERE matricula=s.matricula)
            -(SELECT COALESCE(SUM(c.minutos),0) FROM consumos c JOIN lancamentos l ON l.id=c.lancamento_id WHERE l.matricula=s.matricula)
        ),0) FROM servidores s WHERE s.arquivado=0""").fetchone()[0]

    total_serv  = db.execute("SELECT COUNT(*) FROM servidores WHERE arquivado=0").fetchone()[0]
    serv_fg     = db.execute("SELECT COUNT(*) FROM servidores WHERE funcao_gratificada=1 AND arquivado=0").fetchone()[0]
    venc_count  = db.execute("""
        SELECT COUNT(DISTINCT l.matricula) FROM lancamentos l
        JOIN servidores s ON s.matricula=l.matricula
        WHERE l.data<=? AND s.arquivado=0
          AND l.minutos_creditados>(SELECT COALESCE(SUM(c.minutos),0) FROM consumos c WHERE c.lancamento_id=l.id)
    """, (six_months_ago(),)).fetchone()[0]

    # Próximos vencimentos (horas que vencerão em 30 dias)
    prox_venc = db.execute("""
        SELECT COUNT(DISTINCT l.matricula) FROM lancamentos l
        JOIN servidores s ON s.matricula=l.matricula
        WHERE l.data<=? AND l.data>? AND s.arquivado=0
          AND l.minutos_creditados>(SELECT COALESCE(SUM(c.minutos),0) FROM consumos c WHERE c.lancamento_id=l.id)
    """, (five_months_ago(), six_months_ago())).fetchone()[0]

    # Top 5 saldo
    top5 = db.execute("""
        SELECT s.matricula, s.nome, s.secretaria, s.setor,
            (SELECT COALESCE(SUM(minutos_creditados),0) FROM lancamentos WHERE matricula=s.matricula)
            -(SELECT COALESCE(SUM(c.minutos),0) FROM consumos c JOIN lancamentos l ON l.id=c.lancamento_id WHERE l.matricula=s.matricula)
            AS saldo
        FROM servidores s WHERE s.arquivado=0 ORDER BY saldo DESC LIMIT 5""").fetchall()

    # Top 5 departamentos por saldo agregado
    top5_deptos = db.execute("""
        SELECT COALESCE(NULLIF(s.setor,''),'Sem Departamento') AS setor,
               COALESCE(NULLIF(s.secretaria,''),'Sem Secretaria') AS secretaria,
               COUNT(*) AS qtd_servidores,
               SUM(
                   (SELECT COALESCE(SUM(minutos_creditados),0) FROM lancamentos WHERE matricula=s.matricula)
                   -(SELECT COALESCE(SUM(c.minutos),0) FROM consumos c JOIN lancamentos l ON l.id=c.lancamento_id WHERE l.matricula=s.matricula)
               ) AS saldo
        FROM servidores s
        WHERE s.arquivado=0
        GROUP BY COALESCE(NULLIF(s.setor,''),'Sem Departamento'),
                 COALESCE(NULLIF(s.secretaria,''),'Sem Secretaria')
        ORDER BY saldo DESC
        LIMIT 5
    """).fetchall()
    top5_deptos = [r for r in top5_deptos if minutos_num(r["saldo"]) > 0]

    dados = {
        "meses_labels": json.dumps([m["label"] for m in meses6]),
        "lanc_mes": json.dumps(lanc_mes),
        "comp_mes": json.dumps(comp_mes),
        "pag_mes": json.dumps(pag_mes),
        "saldo_total": saldo_total,
        "total_serv": total_serv,
        "serv_fg": serv_fg,
        "venc_count": venc_count,
        "prox_venc": prox_venc,
        "top5": top5,
        "top5_deptos": top5_deptos,
        "servidores_atalho": servidores_atalho,
    }
    cache_set(cache_key, dados, ttl=30)
    return render_template("dashboard.html", **dados, fmt=minutos_para_horas)

# â”€â”€â”€ Servidores â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/servidores")
@master_required
def servidores():
    PER_PAGE = 20
    db = get_db()
    busca = request.args.get("busca","").strip()
    sec   = request.args.get("secretaria","").strip()
    set_  = request.args.get("setor","").strip()
    apenas_com_saldo = request.args.get("saldo") == "com_saldo"
    fg_sel = request.args.get("fg","").strip()
    vencimento_sel = request.args.get("vencimento","").strip()
    page  = max(1, request.args.get("page", 1, type=int))

    f, p  = _filtro_servidores(busca, sec, set_, arquivado=0)
    if fg_sel == "1":
        f += " AND s.funcao_gratificada=1"
    elif fg_sel == "0":
        f += " AND COALESCE(s.funcao_gratificada,0)=0"
    if vencimento_sel == "vencidas":
        f += """ AND EXISTS (
            SELECT 1 FROM lancamentos l
            WHERE l.matricula=s.matricula
              AND l.data<=?
              AND l.minutos_creditados>(SELECT COALESCE(SUM(c.minutos),0) FROM consumos c WHERE c.lancamento_id=l.id)
        )"""
        p.append(six_months_ago())
    elif vencimento_sel == "proximos":
        f += """ AND EXISTS (
            SELECT 1 FROM lancamentos l
            WHERE l.matricula=s.matricula
              AND l.data<=? AND l.data>?
              AND l.minutos_creditados>(SELECT COALESCE(SUM(c.minutos),0) FROM consumos c WHERE c.lancamento_id=l.id)
        )"""
        p.extend([five_months_ago(), six_months_ago()])
    saldo_expr = """
        (SELECT COALESCE(SUM(minutos_creditados),0) FROM lancamentos WHERE matricula=s.matricula)
        -(SELECT COALESCE(SUM(c.minutos),0) FROM consumos c JOIN lancamentos l ON l.id=c.lancamento_id WHERE l.matricula=s.matricula)
    """
    lista_total = db.execute(f"""
        SELECT s.*,
            {saldo_expr} AS saldo_minutos
        FROM servidores s {f} ORDER BY s.nome""", p).fetchall()
    if apenas_com_saldo:
        lista_total = [s for s in lista_total if minutos_num(s["saldo_minutos"]) > 0]

    total = len(lista_total)
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page = min(page, total_pages)
    lista = lista_total[(page-1)*PER_PAGE : page*PER_PAGE]

    secs, sets = _listas_filtro(db, 0)
    return render_template("servidores.html", servidores=lista, fmt=minutos_para_horas,
                            secretarias=secs, setores=sets,
                            busca=busca, secretaria_sel=sec, setor_sel=set_,
                            saldo_sel="com_saldo" if apenas_com_saldo else "",
                            fg_sel=fg_sel, vencimento_sel=vencimento_sel,
                            page=page, total_pages=total_pages, total=total,
                            per_page=PER_PAGE)

@app.route("/api/historico/<matricula>")
def api_historico(matricula):
    db = get_db()
    if not usuario_pode_ver_matricula(db, matricula):
        return jsonify({"erro": "Acesso não autorizado para esta matrícula."}), 403

    lancs = db.execute("""
        SELECT l.data, l.horas_base, l.percentual, l.minutos_creditados, l.descricao,
               COALESCE((SELECT SUM(c.minutos) FROM consumos c WHERE c.lancamento_id=l.id),0) AS consumido
        FROM lancamentos l WHERE l.matricula=? ORDER BY l.data DESC LIMIT 30""", (matricula,)).fetchall()
    comps = db.execute("SELECT data,tipo,minutos_compensados,descricao FROM compensacoes WHERE matricula=? ORDER BY data DESC LIMIT 30", (matricula,)).fetchall()
    pags  = db.execute("""
        SELECT p.data_pagamento AS data, p.descricao,
               COALESCE(SUM(ROUND(c.minutos*l.minutos_base*1.0/l.minutos_creditados)),0) AS base_paga,
               COALESCE(SUM(c.minutos),0) AS minutos_pagos
        FROM pagamentos p JOIN consumos c ON c.referencia_id=p.id AND c.tipo='pagamento'
        JOIN lancamentos l ON l.id=c.lancamento_id WHERE p.matricula=?
        GROUP BY p.id,p.data_pagamento,p.descricao ORDER BY p.data_pagamento DESC LIMIT 20""", (matricula,)).fetchall()
    eleicao_creditos = db.execute(
        "SELECT referencia_eleicao, quantidade_dias, observacao, criado_em FROM eleicao_creditos WHERE matricula=? ORDER BY criado_em DESC",
        (matricula,)).fetchall()
    eleicao_baixas = db.execute(
        "SELECT data, observacao, criado_em FROM eleicao_baixas WHERE matricula=? ORDER BY data DESC",
        (matricula,)).fetchall()
    saldo_eleicao = calcular_saldo_eleicao(db, matricula)
    return jsonify({
        "lancamentos": [{**dict(r),"minutos_fmt":minutos_para_horas(r["minutos_creditados"]),"saldo_fmt":minutos_para_horas(r["minutos_creditados"]-r["consumido"])} for r in lancs],
        "compensacoes": [{**dict(r),"minutos_fmt":minutos_para_horas(r["minutos_compensados"])} for r in comps],
        "pagamentos":   [{**dict(r),"base_fmt":minutos_para_horas(int(r["base_paga"])),"banco_fmt":minutos_para_horas(r["minutos_pagos"])} for r in pags],
        "eleicao_creditos": [dict(r) for r in eleicao_creditos],
        "eleicao_baixas":   [dict(r) for r in eleicao_baixas],
        "saldo_eleicao":    saldo_eleicao,
    })

@app.route("/servidores/novo", methods=["GET","POST"])
@master_required
def novo_servidor():
    db = get_db()
    if request.method == "POST":
        mat = request.form["matricula"].strip()
        if db.execute("SELECT 1 FROM servidores WHERE matricula=?", (mat,)).fetchone():
            flash("Matrícula já cadastrada.", "danger")
        else:
            fg = 1 if request.form.get("funcao_gratificada") else 0
            cargo = request.form["cargo"].strip()
            setor = request.form["setor"].strip()
            secretaria = request.form["secretaria"].strip()
            _garantir_cadastro_auxiliar(db, "cargo", cargo)
            _garantir_cadastro_auxiliar(db, "departamento", setor)
            _garantir_cadastro_auxiliar(db, "secretaria", secretaria)
            db.execute(
                "INSERT INTO servidores (matricula,nome,cpf,email,cargo,setor,secretaria,funcao_gratificada) VALUES (?,?,?,?,?,?,?,?)",
                (mat, request.form["nome"].strip(), request.form["cpf"].strip(),
                 request.form["email"].strip(), cargo, setor, secretaria, fg))
            db.commit()
            flash("Servidor cadastrado!", "success")
            return redirect(url_for("servidores"))
    return render_template("servidor_form.html", servidor=None, **_opcoes_servidor_form(db))

@app.route("/servidores/<matricula>/editar", methods=["GET","POST"])
@master_required
def editar_servidor(matricula):
    db  = get_db()
    srv = db.execute("SELECT * FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    if not srv: flash("Não encontrado.", "danger"); return redirect(url_for("servidores"))
    if request.method == "POST":
        fg = 1 if request.form.get("funcao_gratificada") else 0
        cargo = request.form["cargo"].strip()
        setor = request.form["setor"].strip()
        secretaria = request.form["secretaria"].strip()
        _garantir_cadastro_auxiliar(db, "cargo", cargo)
        _garantir_cadastro_auxiliar(db, "departamento", setor)
        _garantir_cadastro_auxiliar(db, "secretaria", secretaria)
        db.execute("UPDATE servidores SET nome=?,cpf=?,email=?,cargo=?,setor=?,secretaria=?,funcao_gratificada=? WHERE matricula=?",
                   (request.form["nome"].strip(), request.form["cpf"].strip(), request.form["email"].strip(),
                    cargo, setor, secretaria, fg, matricula))
        db.commit()
        flash("Dados atualizados.", "success")
        return redirect(url_for("servidores"))
    return render_template("servidor_form.html", servidor=srv, **_opcoes_servidor_form(db))

@app.route("/servidores/<matricula>/arquivar", methods=["POST"])
@master_required
def arquivar_servidor(matricula):
    db = get_db()
    srv = db.execute("SELECT * FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    if not srv:
        flash("Servidor não encontrado.", "danger")
        return redirect(url_for("servidores"))
    db.execute("UPDATE servidores SET arquivado=1 WHERE matricula=?", (matricula,))
    inativados = _inativar_vinculos_servidor(db, srv)
    db.commit()
    extra = f" {inativados} acesso(s) vinculado(s) foram inativados." if inativados else ""
    flash(f"Servidor arquivado. Dados preservados para consulta.{extra}", "warning")
    return redirect(url_for("servidores"))

@app.route("/servidores/<matricula>/restaurar", methods=["POST"])
@master_required
def restaurar_servidor(matricula):
    db = get_db()
    db.execute("UPDATE servidores SET arquivado=0 WHERE matricula=?", (matricula,))
    db.commit()
    flash("Servidor restaurado com sucesso. Acessos permanecem inativos até reativação pelo RH.", "success")
    return redirect(url_for("arquivados"))

# â”€â”€â”€ Arquivados â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/servidores/<matricula>/excluir", methods=["POST"])
@master_required
def excluir_servidor(matricula):
    db = get_db()
    srv = db.execute("SELECT * FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    if not srv:
        flash("Servidor não encontrado.", "danger")
        return redirect(url_for("servidores"))
    payload = _snapshot_servidor_exclusao(db, matricula)
    payload["servidor"] = dict(srv)
    criado_em = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    auditoria_id = registrar_auditoria(
        db, "Excluiu cadastro de servidor", "servidor", matricula, matricula, srv["nome"],
        f"Cadastro e movimentações removidos definitivamente; secretaria {srv['secretaria'] or '-'}; departamento {srv['setor'] or '-'}"
    )
    db.insert("""
        INSERT INTO exclusoes_servidores (auditoria_id,matricula,servidor_nome,payload,criado_em,restaurado)
        VALUES (?,?,?,?,?,0)
    """, (auditoria_id, matricula, srv["nome"], json.dumps(payload, default=str), criado_em))
    db.execute("DELETE FROM consumos WHERE lancamento_id IN (SELECT id FROM lancamentos WHERE matricula=?)", (matricula,))
    db.execute("DELETE FROM consumos WHERE tipo='compensacao' AND referencia_id IN (SELECT id FROM compensacoes WHERE matricula=?)", (matricula,))
    db.execute("DELETE FROM consumos WHERE tipo='pagamento' AND referencia_id IN (SELECT id FROM pagamentos WHERE matricula=?)", (matricula,))
    db.execute("DELETE FROM pagamentos WHERE matricula=?", (matricula,))
    db.execute("DELETE FROM compensacoes WHERE matricula=?", (matricula,))
    db.execute("DELETE FROM lancamentos WHERE matricula=?", (matricula,))
    _inativar_vinculos_servidor(db, srv)
    db.execute("""
        UPDATE usuarios
        SET matricula=NULL, vinculos='[]'
        WHERE ((?<>'' AND matricula=?) OR (?<>'' AND cpf=?))
          AND nivel<>'master'
    """, (matricula, matricula, (srv["cpf"] or "").strip(), (srv["cpf"] or "").strip()))
    db.execute("DELETE FROM servidores WHERE matricula=?", (matricula,))
    db.commit()
    flash("Cadastro do servidor excluído definitivamente.", "warning")
    return redirect(url_for("servidores"))

@app.route("/arquivados")
@master_required
def arquivados():
    db    = get_db()
    busca = request.args.get("busca","").strip()
    sec   = request.args.get("secretaria","").strip()
    set_  = request.args.get("setor","").strip()
    f, p  = _filtro_servidores(busca, sec, set_, arquivado=1)
    lista = db.execute(f"""
        SELECT s.*,
            (SELECT COALESCE(SUM(minutos_creditados),0) FROM lancamentos WHERE matricula=s.matricula)
            -(SELECT COALESCE(SUM(c.minutos),0) FROM consumos c JOIN lancamentos l ON l.id=c.lancamento_id WHERE l.matricula=s.matricula)
            AS saldo_minutos
        FROM servidores s {f} ORDER BY s.nome""", p).fetchall()
    secs, sets = _listas_filtro(db, 1)
    return render_template("arquivados.html", servidores=lista, fmt=minutos_para_horas,
                           secretarias=secs, setores=sets,
                           busca=busca, secretaria_sel=sec, setor_sel=set_)

# â”€â”€â”€ Lançamentos â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/lancamentos/<matricula>", methods=["GET","POST"])
@master_required
def lancamentos(matricula):
    db  = get_db()
    srv = db.execute("SELECT * FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    if not srv: flash("Não encontrado.","danger"); return redirect(url_for("servidores"))
    if request.method == "POST":
        data = request.form["data"]; hrs = request.form["horas"].strip()
        pct  = int(request.form["percentual"]); desc = request.form["descricao"].strip()
        mb   = horas_para_minutos(hrs)
        if mb <= 0: flash("Formato HH:MM inválido.","danger")
        else:
            mc = mb + mb*pct//100
            lid = db.insert("INSERT INTO lancamentos (matricula,data,horas_base,minutos_base,percentual,minutos_creditados,descricao) VALUES (?,?,?,?,?,?,?)",
                            (matricula, data, hrs, mb, pct, mc, desc))
            registrar_auditoria(
                db, "Criou lançamento", "lancamento", lid, matricula, srv["nome"],
                f"Data {data}; horas base {hrs}; adicional {pct}%; crédito {minutos_para_horas(mc)}; descrição: {desc or '-'}"
            )
            db.commit()
            flash(f"{srv['nome']} ({matricula})\nLançamento de {hrs} + {pct}% = {minutos_para_horas(mc)} creditado no banco.","success")
            if request.form.get('_source') == 'dashboard':
                return redirect(url_for("dashboard"))
            return redirect(url_for("lancamentos", matricula=matricula))
    hist = db.execute("""
        SELECT l.*, COALESCE((SELECT SUM(c.minutos) FROM consumos c WHERE c.lancamento_id=l.id),0) AS consumido
        FROM lancamentos l WHERE l.matricula=? ORDER BY l.data DESC""", (matricula,)).fetchall()
    return render_template("lancamentos.html", servidor=srv, historico=hist,
                           saldo=calcular_saldo(db, matricula), fmt=minutos_para_horas, fg=_fg(srv))

@app.route("/lancamentos/<int:id>/excluir", methods=["POST"])
@master_required
def excluir_lancamento(id):
    db = get_db()
    r  = db.execute("""
        SELECT l.*, s.nome AS servidor_nome FROM lancamentos l
        LEFT JOIN servidores s ON s.matricula=l.matricula WHERE l.id=?
    """, (id,)).fetchone()
    if r:
        registrar_auditoria(
            db, "Excluiu lançamento", "lancamento", id, r["matricula"], r["servidor_nome"],
            f"Data {r['data']}; horas base {r['horas_base']}; adicional {r['percentual']}%; crédito {minutos_para_horas(r['minutos_creditados'])}; descrição: {r['descricao'] or '-'}"
        )
        db.execute("DELETE FROM consumos WHERE lancamento_id=?", (id,))
        db.execute("DELETE FROM lancamentos WHERE id=?", (id,))
        db.commit(); flash("Lançamento excluído.","warning")
        return redirect(url_for("lancamentos", matricula=r["matricula"]))
    return redirect(url_for("servidores"))

# â”€â”€â”€ Compensações â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/compensacoes/<matricula>", methods=["GET","POST"])
@master_required
def compensacoes(matricula):
    db  = get_db()
    srv = db.execute("SELECT * FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    if not srv: flash("Não encontrado.","danger"); return redirect(url_for("servidores"))
    saldo = calcular_saldo(db, matricula)
    if request.method == "POST":
        data = request.form["data"]; tipo = "parcial"; desc = request.form["descricao"].strip()
        mc   = horas_para_minutos(request.form.get("horas","").strip())
        if mc <= 0: flash("Valor inválido.","danger")
        else:
            cid = db.insert("INSERT INTO compensacoes (matricula,data,tipo,minutos_compensados,descricao) VALUES (?,?,?,?,?)",
                            (matricula, data, tipo, mc, desc))
            _consumir_fifo_raw(db, matricula, mc, "compensacao", cid)
            registrar_auditoria(
                db, "Criou compensação", "compensacao", cid, matricula, srv["nome"],
                f"Data {data}; tipo {tipo}; compensado {minutos_para_horas(mc)}; saldo anterior {minutos_para_horas(saldo)}; descrição: {desc or '-'}"
            )
            db.commit()
            novo_saldo = saldo - mc
            if novo_saldo < 0:
                flash(f"{srv['nome']} ({matricula})\nCompensação de {minutos_para_horas(mc)} registrada. Atenção: saldo ficou negativo em {minutos_para_horas(novo_saldo)}.", "warning")
            else:
                flash(f"{srv['nome']} ({matricula})\nCompensação de {minutos_para_horas(mc)} registrada com sucesso.", "success")
            if request.form.get('_source') == 'dashboard':
                return redirect(url_for("dashboard"))
            return redirect(url_for("compensacoes", matricula=matricula))
    hist = db.execute("SELECT * FROM compensacoes WHERE matricula=? ORDER BY data DESC", (matricula,)).fetchall()
    return render_template("compensacoes.html", servidor=srv, historico=hist,
                           saldo=saldo, fmt=minutos_para_horas, fg=_fg(srv))

@app.route("/compensacoes/<int:id>/excluir", methods=["POST"])
@master_required
def excluir_compensacao(id):
    db = get_db()
    r  = db.execute("""
        SELECT c.*, s.nome AS servidor_nome FROM compensacoes c
        LEFT JOIN servidores s ON s.matricula=c.matricula WHERE c.id=?
    """, (id,)).fetchone()
    if r:
        registrar_auditoria(
            db, "Excluiu compensação", "compensacao", id, r["matricula"], r["servidor_nome"],
            f"Data {r['data']}; tipo {r['tipo']}; compensado {minutos_para_horas(r['minutos_compensados'])}; descrição: {r['descricao'] or '-'}"
        )
        db.execute("DELETE FROM consumos WHERE tipo='compensacao' AND referencia_id=?", (id,))
        db.execute("DELETE FROM compensacoes WHERE id=?", (id,))
        db.commit(); flash("Compensação excluída.","warning")
        return redirect(url_for("compensacoes", matricula=r["matricula"]))
    return redirect(url_for("servidores"))

# â”€â”€â”€ Pagamentos â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/pagamentos")
@master_required
def pagamentos_index():
    db   = get_db()
    busca      = request.args.get('busca', '').strip()
    sec_sel    = request.args.get('secretaria', '').strip()
    set_sel    = request.args.get('setor', '').strip()
    status_sel = request.args.get('status', '').strip()       # 'vencidas' | ''
    # ocultar_fg: default '1' (ocultar) na primeira visita sem params
    ocultar_fg = request.args.get('ocultar_fg', '1' if not request.args else '0')

    srvs = db.execute("SELECT * FROM servidores WHERE arquivado=0 ORDER BY nome").fetchall()
    pend_todos = []
    for s in srvs:
        itens_v = lancamentos_com_saldo(db, s["matricula"], apenas_vencidos=True)
        itens_t = lancamentos_com_saldo(db, s["matricula"], apenas_vencidos=False)
        if not itens_t: continue
        tb_v = sum(i["saldo_base_minutos"] for i in itens_v)
        tb_t = sum(i["saldo_base_minutos"] for i in itens_t)
        pend_todos.append({"matricula":s["matricula"],"nome":s["nome"],"cargo":s["cargo"],
                           "setor":s["setor"],"secretaria":s["secretaria"],
                           "funcao_gratificada":bool(s["funcao_gratificada"]),
                           "total_base_vencidas":tb_v,"total_base_todos":tb_t,
                           "qtd_vencidas":len(itens_v),"qtd_total":len(itens_t),
                           "acima_limite":tb_t>LIMITE_PAGAMENTO_MINUTOS})

    # Listas para filtros (antes de filtrar)
    secretarias = sorted({p["secretaria"] for p in pend_todos if p["secretaria"]})
    setores     = sorted({p["setor"]      for p in pend_todos if p["setor"]})

    # Aplicar filtros
    pend = pend_todos
    if ocultar_fg == '1':
        pend = [p for p in pend if not p["funcao_gratificada"]]
    if busca:
        bl = busca.lower()
        pend = [p for p in pend if bl in p["nome"].lower() or bl in p["matricula"].lower()]
    if sec_sel:
        pend = [p for p in pend if p["secretaria"] == sec_sel]
    if set_sel:
        pend = [p for p in pend if p["setor"] == set_sel]
    if status_sel == 'vencidas':
        pend = [p for p in pend if p["qtd_vencidas"] > 0]

    return render_template("pagamentos_index.html", pendentes=pend,
                           fmt=minutos_para_horas,
                           limite=LIMITE_PAGAMENTO_MINUTOS,
                           limite_fmt=minutos_para_horas(LIMITE_PAGAMENTO_MINUTOS),
                           secretarias=secretarias, setores=setores,
                           busca=busca, sec_sel=sec_sel, set_sel=set_sel,
                           status_sel=status_sel, ocultar_fg=ocultar_fg,
                           total_sem_fg=len([p for p in pend_todos if not p["funcao_gratificada"]]))

@app.route("/api/servidores-lista")
@login_required
def api_servidores_lista():
    """Lista de servidores para autocomplete de busca — filtrada pelo nível do usuário."""
    db    = get_db()
    nivel = session.get('nivel')
    vinculos = session.get('vinculos', [])

    if nivel == 'master':
        rows = db.execute(
            "SELECT matricula, nome, secretaria, setor FROM servidores WHERE arquivado=0 ORDER BY nome"
        ).fetchall()
    elif nivel in ('secretario', 'chefia'):
        acesso_sql, acesso_params = filtro_consulta_vinculos(nivel, vinculos, session.get('matricula') or '')
        rows = db.execute(
            f"SELECT matricula, nome, secretaria, setor FROM servidores "
            f"WHERE arquivado=0 {acesso_sql} ORDER BY nome", acesso_params
        ).fetchall()
    elif nivel == 'servidor':
        mat = session.get('matricula')
        rows = db.execute(
            "SELECT matricula, nome, secretaria, setor FROM servidores WHERE matricula=? AND arquivado=0",
            (mat,)
        ).fetchall() if mat else []
    else:
        rows = []

    return jsonify([{
        "m": r['matricula'],
        "n": r['nome'],
        "i": ' · '.join(filter(None, [r['secretaria'], r['setor']]))
    } for r in rows])


@app.route("/api/servidor-info/<matricula>")
@master_required
def api_servidor_info(matricula):
    db = get_db()
    srv = db.execute(
        "SELECT nome, cpf, cargo, secretaria, setor FROM servidores WHERE matricula=? AND arquivado=0",
        (matricula,)).fetchone()
    if not srv:
        return jsonify({"encontrado": False})
    ja_tem_conta = bool(db.execute(
        "SELECT 1 FROM usuarios WHERE cpf=? AND ativo=1", (srv['cpf'] or '',)).fetchone())
    return jsonify({
        "encontrado": True,
        "nome": srv['nome'],
        "cargo": srv['cargo'] or '',
        "secretaria": srv['secretaria'] or '',
        "setor": srv['setor'] or '',
        "ja_tem_conta": ja_tem_conta,
        "tem_cpf": bool((srv['cpf'] or '').strip()),
    })


@app.route("/api/saldo-servidor/<matricula>")
@master_required
def api_saldo_servidor(matricula):
    db = get_db()
    return jsonify({"saldo": calcular_saldo(db, matricula),
                    "saldo_fmt": minutos_para_horas(calcular_saldo(db, matricula))})


@app.route("/api/pagamentos-itens/<matricula>")
@master_required
def api_pagamentos_itens(matricula):
    db = get_db()
    itens_v = lancamentos_com_saldo(db, matricula, apenas_vencidos=True)
    itens_t = lancamentos_com_saldo(db, matricula, apenas_vencidos=False)
    return jsonify({
        "vencidos": [dict(i) for i in itens_v],
        "todos":    [dict(i) for i in itens_t],
        "limite":   LIMITE_PAGAMENTO_MINUTOS,
    })


@app.route("/pagamentos/<matricula>")
@master_required
def pagamentos_servidor(matricula):
    db  = get_db()
    srv = db.execute("SELECT * FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    if not srv: flash("Não encontrado.","danger"); return redirect(url_for("pagamentos_index"))
    itens = lancamentos_com_saldo(db, matricula, apenas_vencidos=False)
    hist  = db.execute("SELECT * FROM pagamentos WHERE matricula=? ORDER BY data_pagamento DESC", (matricula,)).fetchall()
    return render_template("pagamentos_servidor.html", servidor=srv, itens=itens, historico=hist,
                           fmt=minutos_para_horas, limite=LIMITE_PAGAMENTO_MINUTOS,
                           today=date.today().isoformat(), fg=_fg(srv))

@app.route("/pagamentos/<matricula>/registrar", methods=["POST"])
@master_required
def registrar_pagamento(matricula):
    db = get_db()
    srv = db.execute("SELECT nome FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    data_pag = request.form["data_pagamento"]
    descricao = request.form.get("descricao","").strip()
    just      = request.form.get("justificativa","").strip()
    itens     = lancamentos_com_saldo(db, matricula, apenas_vencidos=False)
    if not itens: flash("Sem horas com saldo.","warning"); return redirect(url_for("pagamentos_servidor", matricula=matricula))

    pag_itens=[]; total_base=0; tem_nao_venc=False
    for item in itens:
        v = request.form.get(f"horas_{item['id']}","").strip()
        if not v: continue
        mb_pagar = horas_para_minutos(v)
        if mb_pagar <= 0: continue
        if not item["vencido"]: tem_nao_venc = True
        mc = round(mb_pagar * item["minutos_creditados"] / item["minutos_base"]) if item["minutos_base"]>0 else mb_pagar
        mc = min(mc, item["saldo_minutos"])
        if mc <= 0: continue
        br = round(mc * item["minutos_base"] / item["minutos_creditados"]) if item["minutos_creditados"]>0 else 0
        total_base += br; pag_itens.append((item["id"], mc))

    if (tem_nao_venc or total_base>LIMITE_PAGAMENTO_MINUTOS) and not just:
        flash("Justificativa obrigatória para horas não vencidas ou acima de 45h.","danger")
        return redirect(url_for("pagamentos_servidor", matricula=matricula))
    if not pag_itens: flash("Nenhuma hora informada.","warning"); return redirect(url_for("pagamentos_servidor", matricula=matricula))

    desc_final = descricao + (f" | Justificativa: {just}" if just else "")
    pid = db.insert("INSERT INTO pagamentos (matricula,data_pagamento,descricao) VALUES (?,?,?)", (matricula,data_pag,desc_final))
    for lid, mins in pag_itens:
        db.execute("INSERT INTO consumos (lancamento_id,tipo,referencia_id,minutos) VALUES (?,?,?,?)", (lid,"pagamento",pid,mins))
    registrar_auditoria(
        db, "Criou pagamento", "pagamento", pid, matricula, srv["nome"] if srv else "",
        f"Data {data_pag}; horas base para folha {minutos_para_horas(total_base)}; itens pagos {len(pag_itens)}; descrição: {desc_final or '-'}"
    )
    db.commit()
    aviso = f"\n⚠️ Total {minutos_para_horas(total_base)} ultrapassa o limite de 45h." if total_base>LIMITE_PAGAMENTO_MINUTOS else ""
    nome_srv = srv["nome"] if srv else matricula
    flash(f"{nome_srv} ({matricula})\nPagamento de {minutos_para_horas(total_base)} horas base registrado no banco.{aviso}","success")
    source = request.form.get('_source', 'servidor')
    if source == 'index':
        return redirect(url_for("pagamentos_index"))
    return redirect(url_for("pagamentos_servidor", matricula=matricula))

@app.route("/pagamentos/<int:id>/estornar", methods=["POST"])
@master_required
def estornar_pagamento(id):
    db = get_db()
    r  = db.execute("""
        SELECT p.*, s.nome AS servidor_nome,
               COALESCE(SUM(ROUND(c.minutos*l.minutos_base*1.0/l.minutos_creditados)),0) AS base_paga
        FROM pagamentos p
        LEFT JOIN servidores s ON s.matricula=p.matricula
        LEFT JOIN consumos c ON c.referencia_id=p.id AND c.tipo='pagamento'
        LEFT JOIN lancamentos l ON l.id=c.lancamento_id
        WHERE p.id=?
        GROUP BY p.id,p.matricula,p.data_pagamento,p.descricao,p.criado_em,s.nome
    """, (id,)).fetchone()
    if r:
        registrar_auditoria(
            db, "Estornou pagamento", "pagamento", id, r["matricula"], r["servidor_nome"],
            f"Data {r['data_pagamento']}; horas base {minutos_para_horas(r['base_paga'])}; descrição: {r['descricao'] or '-'}"
        )
        db.execute("DELETE FROM consumos WHERE tipo='pagamento' AND referencia_id=?", (id,))
        db.execute("DELETE FROM pagamentos WHERE id=?", (id,))
        db.commit(); flash("Pagamento estornado.","warning")
        return redirect(url_for("pagamentos_servidor", matricula=r["matricula"]))
    return redirect(url_for("pagamentos_index"))

@app.route("/admin/auditoria")
@master_required
def admin_auditoria():
    db = get_db()
    desde = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")
    eventos = db.execute("""
        SELECT a.*, ex.id AS exclusao_id, ex.restaurado, ex.criado_em AS exclusao_criada
        FROM auditoria a
        LEFT JOIN exclusoes_servidores ex ON ex.auditoria_id=a.id
        WHERE a.criado_em >= ?
        ORDER BY a.criado_em ASC, a.id ASC
    """, (desde,)).fetchall()
    return render_template("admin/auditoria.html", eventos=eventos, desde=desde)

@app.route("/admin/auditoria/exclusao/<int:exclusao_id>/estornar", methods=["POST"])
@master_required
def admin_estornar_exclusao_servidor(exclusao_id):
    db = get_db()
    ex = db.execute("SELECT * FROM exclusoes_servidores WHERE id=?", (exclusao_id,)).fetchone()
    if not ex:
        flash("Registro de exclusão não encontrado.", "danger")
        return redirect(url_for("admin_auditoria"))
    if ex["restaurado"]:
        flash("Esta exclusão já foi estornada.", "warning")
        return redirect(url_for("admin_auditoria"))
    if ex["criado_em"] < (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S"):
        flash("Prazo de 30 dias para estorno expirado.", "danger")
        return redirect(url_for("admin_auditoria"))
    if db.execute("SELECT 1 FROM servidores WHERE matricula=?", (ex["matricula"],)).fetchone():
        flash("Não foi possível estornar: já existe servidor com esta matrícula.", "danger")
        return redirect(url_for("admin_auditoria"))

    payload = json.loads(ex["payload"])
    srv = payload.get("servidor", {})
    db.execute("""
        INSERT INTO servidores (matricula,nome,cargo,setor,secretaria,cpf,funcao_gratificada,email,arquivado)
        VALUES (?,?,?,?,?,?,?,?,?)
    """, (
        srv.get("matricula"), srv.get("nome"), srv.get("cargo"), srv.get("setor"),
        srv.get("secretaria"), srv.get("cpf"), srv.get("funcao_gratificada", 0),
        srv.get("email"), srv.get("arquivado", 0)
    ))
    for l in payload.get("lancamentos", []):
        db.execute("""INSERT INTO lancamentos
            (id,matricula,data,horas_base,minutos_base,percentual,minutos_creditados,descricao,criado_em)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            (l.get("id"), l.get("matricula"), l.get("data"), l.get("horas_base"), l.get("minutos_base"),
             l.get("percentual"), l.get("minutos_creditados"), l.get("descricao"), l.get("criado_em")))
    for c in payload.get("compensacoes", []):
        db.execute("""INSERT INTO compensacoes
            (id,matricula,data,tipo,minutos_compensados,descricao,criado_em)
            VALUES (?,?,?,?,?,?,?)""",
            (c.get("id"), c.get("matricula"), c.get("data"), c.get("tipo"), c.get("minutos_compensados"), c.get("descricao"), c.get("criado_em")))
    for p in payload.get("pagamentos", []):
        db.execute("""INSERT INTO pagamentos
            (id,matricula,data_pagamento,descricao,criado_em)
            VALUES (?,?,?,?,?)""",
            (p.get("id"), p.get("matricula"), p.get("data_pagamento"), p.get("descricao"), p.get("criado_em")))
    vistos = set()
    for c in payload.get("consumos", []):
        if c.get("id") in vistos:
            continue
        vistos.add(c.get("id"))
        db.execute("""INSERT INTO consumos
            (id,lancamento_id,tipo,referencia_id,minutos,criado_em)
            VALUES (?,?,?,?,?,?)""",
            (c.get("id"), c.get("lancamento_id"), c.get("tipo"), c.get("referencia_id"), c.get("minutos"), c.get("criado_em")))
    db.execute("UPDATE exclusoes_servidores SET restaurado=1, restaurado_em=? WHERE id=?",
               (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), exclusao_id))
    registrar_auditoria(db, "Estornou exclusão de servidor", "servidor", ex["matricula"], ex["matricula"], ex["servidor_nome"], "Cadastro restaurado pela auditoria.")
    db.commit()
    flash("Exclusão estornada. Cadastro e movimentações foram restaurados.", "success")
    return redirect(url_for("admin_auditoria"))

# â”€â”€â”€ Relatórios â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/admin/importacao")
@master_required
def admin_importacao():
    db = get_db()
    historico = db.execute("SELECT * FROM importacoes ORDER BY criado_em DESC, id DESC LIMIT 50").fetchall()
    relatorio = session.pop("import_report", None)
    return render_template("admin/importacao.html", historico=historico, relatorio=relatorio)

@app.route("/admin/importacao/modelo")
@master_required
def admin_importacao_modelo():
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["matricula","nome","cpf","email","cargo","secretaria","setor","funcao_gratificada"])
    w.writerow(["12345","Nome do Servidor","000.000.000-00","servidor@ibipora.pr.gov.br","Cargo","Secretaria","Departamento",""])
    w.writerow(["67890","Servidor com FG","111.111.111-11","fg@ibipora.pr.gov.br","Cargo","Secretaria","Departamento","FG-1"])
    r = make_response("\ufeff" + out.getvalue())
    r.headers["Content-Type"] = "text/csv; charset=utf-8"
    r.headers["Content-Disposition"] = "attachment; filename=modelo_importacao_servidores.csv"
    return r

@app.route("/admin/importacao/modelo-cadastros")
@master_required
def admin_importacao_modelo_cadastros():
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["tipo","nome"])
    w.writerow(["secretaria","Secretaria Municipal de Administração"])
    w.writerow(["departamento","Departamento de Recursos Humanos"])
    w.writerow(["cargo","Assistente Administrativo"])
    r = make_response("\ufeff" + out.getvalue())
    r.headers["Content-Type"] = "text/csv; charset=utf-8"
    r.headers["Content-Disposition"] = "attachment; filename=modelo_importacao_cadastros.csv"
    return r

@app.route("/admin/importacao/modelo-banco-horas")
@master_required
def admin_importacao_modelo_banco_horas():
    out = io.StringIO()
    w = csv.writer(out, delimiter=";")
    w.writerow(["matricula","nome","saldo_horas","percentual","competencia","data_lancamento","justificativa","observacoes"])
    w.writerow(["12345","Nome do Servidor","12:30","50","04/2026","2026-04-30","Importação de saldo inicial","Conferido pela unidade"])
    w.writerow(["67890","Outro Servidor","08:00","100","05/2026","","Ajuste autorizado","Data em branco usa a data atual"])
    r = make_response("\ufeff" + out.getvalue())
    r.headers["Content-Type"] = "text/csv; charset=utf-8"
    r.headers["Content-Disposition"] = "attachment; filename=modelo_importacao_banco_horas.csv"
    return r

@app.route("/admin/importacao/modelo-eleicao")
@master_required
def admin_importacao_modelo_eleicao():
    out = io.StringIO()
    w = csv.writer(out, delimiter=";")
    w.writerow(["matricula","nome","quantidade_dias","referencia","data_lancamento","justificativa","observacoes"])
    w.writerow(["12345","Nome do Servidor","3","Eleições 2024","2026-04-30","Importação de saldo eleitoral","Portaria/registro interno"])
    w.writerow(["67890","Outro Servidor","2","Convocação TRE 2024","","Ajuste autorizado","Data em branco usa a data atual"])
    r = make_response("\ufeff" + out.getvalue())
    r.headers["Content-Type"] = "text/csv; charset=utf-8"
    r.headers["Content-Disposition"] = "attachment; filename=modelo_importacao_dias_eleicao.csv"
    return r

@app.route("/admin/importacao/servidores", methods=["POST"])
@master_required
def admin_importar_servidores():
    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        flash("Selecione um arquivo CSV.", "danger")
        return redirect(url_for("admin_importacao"))
    try:
        leitor = _csv_reader_upload(arquivo)
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("admin_importacao"))
    campos = {str(c or "").strip().lower() for c in (leitor.fieldnames or [])}
    if not {"matricula", "nome"}.issubset(campos):
        flash("CSV inválido. Campos obrigatórios: matricula e nome.", "danger")
        return redirect(url_for("admin_importacao"))

    db = get_db()
    total = 0
    payload, sucessos, erros_lista = [], [], []
    for row in leitor:
        total += 1
        linha = total + 1
        matricula = _valor_csv(row, "matricula", "matrícula")
        nome = _valor_csv(row, "nome")
        if not matricula or not nome:
            _erro_importacao(erros_lista, linha, matricula, "Campo obrigatório não preenchido: matrícula e nome.")
            continue
        atual = db.execute("SELECT * FROM servidores WHERE matricula=?", (matricula,)).fetchone()
        if atual:
            _erro_importacao(erros_lista, linha, matricula, "Servidor já cadastrado. O registro existente não foi alterado.")
            continue
        dados = {
            "matricula": matricula, "nome": nome,
            "cpf": _valor_csv(row, "cpf"),
            "email": _valor_csv(row, "email", "e-mail"),
            "cargo": _valor_csv(row, "cargo"),
            "secretaria": _valor_csv(row, "secretaria", "secretária"),
            "setor": _valor_csv(row, "setor", "departamento"),
            "funcao_gratificada": 1 if _valor_csv(row, "funcao_gratificada", "fg", "função gratificada") else 0,
            "arquivado": 0,
        }
        _garantir_cadastro_auxiliar(db, "cargo", dados["cargo"])
        _garantir_cadastro_auxiliar(db, "secretaria", dados["secretaria"])
        _garantir_cadastro_auxiliar(db, "departamento", dados["setor"])
        db.execute("""INSERT INTO servidores
                      (matricula,nome,cpf,email,cargo,secretaria,setor,funcao_gratificada,arquivado)
                      VALUES (?,?,?,?,?,?,?,?,0)""",
                   (dados["matricula"], dados["nome"], dados["cpf"], dados["email"], dados["cargo"],
                    dados["secretaria"], dados["setor"], dados["funcao_gratificada"]))
        payload.append({"acao": "criado", "antes": None, "depois": dados})
        _sucesso_importacao(sucessos, linha, matricula, f"Servidor criado: {nome}")

    _registrar_relatorio_importacao("servidores", arquivo.filename, total, sucessos, erros_lista, payload)
    limpar_cache()
    return redirect(url_for("admin_importacao"))

def _normalizar_tipo_cadastro(valor):
    v = (valor or "").strip().lower()
    mapa = {
        "secretaria": "secretaria",
        "secretarias": "secretaria",
        "departamento": "departamento",
        "departamentos": "departamento",
        "local": "departamento",
        "local de trabalho": "departamento",
        "locais de trabalho": "departamento",
        "setor": "departamento",
        "setores": "departamento",
        "cargo": "cargo",
        "cargos": "cargo",
    }
    return mapa.get(v, "")

def _data_importacao(valor):
    valor = (valor or "").strip()
    if not valor:
        return date.today().isoformat()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(valor, fmt).date().isoformat()
        except ValueError:
            pass
    return ""

def _texto_importacao(*partes):
    return " | ".join(str(p).strip() for p in partes if str(p or "").strip())

@app.route("/admin/importacao/banco-horas", methods=["POST"])
@master_required
def admin_importar_banco_horas():
    arquivo = request.files.get("arquivo")
    competencia_padrao = request.form.get("competencia", "").strip()
    justificativa_padrao = request.form.get("justificativa", "").strip()
    if not arquivo or not arquivo.filename:
        flash("Selecione um arquivo CSV.", "danger")
        return redirect(url_for("admin_importacao"))
    try:
        leitor = _csv_reader_upload(arquivo)
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("admin_importacao"))
    campos = {str(c or "").strip().lower() for c in (leitor.fieldnames or [])}
    if "matricula" not in campos or not ({"saldo_horas", "horas", "saldo"} & campos):
        flash("CSV inválido. Campos obrigatórios: matricula e saldo_horas.", "danger")
        return redirect(url_for("admin_importacao"))

    db = get_db()
    total = 0
    payload, sucessos, erros_lista = [], [], []
    for row in leitor:
        total += 1
        linha = total + 1
        matricula = _valor_csv(row, "matricula", "matrícula")
        nome_csv = _valor_csv(row, "nome")
        horas_txt = _valor_csv(row, "saldo_horas", "horas", "saldo")
        percentual_txt = _valor_csv(row, "percentual") or "0"
        competencia = _valor_csv(row, "competencia", "competência") or competencia_padrao
        justificativa = _valor_csv(row, "justificativa") or justificativa_padrao
        obs = _valor_csv(row, "observacoes", "observações", "obs")
        data_lanc = _data_importacao(_valor_csv(row, "data_lancamento", "data"))
        if not matricula or not horas_txt or not competencia or not justificativa:
            _erro_importacao(erros_lista, linha, matricula, "Campo obrigatório não preenchido: matrícula, saldo_horas, competência e justificativa.")
            continue
        try:
            pct = int(str(percentual_txt).strip().replace("%", ""))
            if pct not in (0, 50, 100):
                raise ValueError()
        except ValueError:
            _erro_importacao(erros_lista, linha, matricula, "Percentual inválido. Use 0, 50 ou 100.")
            continue
        minutos = horas_para_minutos(horas_txt)
        if minutos <= 0 or not data_lanc:
            _erro_importacao(erros_lista, linha, matricula, "Formato inválido em saldo_horas ou data_lancamento.")
            continue
        minutos_creditados = minutos + minutos * pct // 100
        srv = db.execute("SELECT nome FROM servidores WHERE matricula=? AND arquivado=0", (matricula,)).fetchone()
        if not srv:
            _erro_importacao(erros_lista, linha, matricula, "Matrícula não localizada em servidores ativos.")
            continue
        duplicado = db.execute("""
            SELECT 1 FROM lancamentos
            WHERE matricula=? AND descricao LIKE ? AND minutos_base=?
            LIMIT 1
        """, (matricula, f"%Competência: {competencia}%", minutos)).fetchone()
        if duplicado:
            _erro_importacao(erros_lista, linha, matricula, "Lançamento de banco de horas já existente para esta competência.")
            continue
        desc = _texto_importacao("Importação Banco de Horas", f"Competência: {competencia}", f"Justificativa: {justificativa}", obs)
        lid = db.insert("""INSERT INTO lancamentos
            (matricula,data,horas_base,minutos_base,percentual,minutos_creditados,descricao)
            VALUES (?,?,?,?,?,?,?)""",
            (matricula, data_lanc, minutos_para_horas(minutos), minutos, pct, minutos_creditados, desc))
        payload.append({"acao": "criado_lancamento", "id": lid, "matricula": matricula})
        _sucesso_importacao(sucessos, linha, matricula, f"{srv['nome'] or nome_csv}: {minutos_para_horas(minutos)} base ({pct}%) = {minutos_para_horas(minutos_creditados)} creditado(s) na competência {competencia}")

    _registrar_relatorio_importacao("banco_horas", arquivo.filename, total, sucessos, erros_lista, payload)
    return redirect(url_for("admin_importacao"))

@app.route("/admin/importacao/eleicao", methods=["POST"])
@master_required
def admin_importar_eleicao():
    arquivo = request.files.get("arquivo")
    referencia_padrao = request.form.get("referencia", "").strip()
    justificativa_padrao = request.form.get("justificativa", "").strip()
    if not arquivo or not arquivo.filename:
        flash("Selecione um arquivo CSV.", "danger")
        return redirect(url_for("admin_importacao"))
    try:
        leitor = _csv_reader_upload(arquivo)
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("admin_importacao"))
    campos = {str(c or "").strip().lower() for c in (leitor.fieldnames or [])}
    if "matricula" not in campos or not ({"quantidade_dias", "dias"} & campos):
        flash("CSV inválido. Campos obrigatórios: matricula e quantidade_dias.", "danger")
        return redirect(url_for("admin_importacao"))

    db = get_db()
    total = 0
    payload, sucessos, erros_lista = [], [], []
    for row in leitor:
        total += 1
        linha = total + 1
        matricula = _valor_csv(row, "matricula", "matrícula")
        nome_csv = _valor_csv(row, "nome")
        dias_txt = _valor_csv(row, "quantidade_dias", "dias")
        referencia = _valor_csv(row, "referencia", "referência", "competencia", "competência") or referencia_padrao
        justificativa = _valor_csv(row, "justificativa") or justificativa_padrao
        obs = _valor_csv(row, "observacoes", "observações", "obs")
        if not matricula or not dias_txt or not referencia or not justificativa:
            _erro_importacao(erros_lista, linha, matricula, "Campo obrigatório não preenchido: matrícula, quantidade_dias, referência e justificativa.")
            continue
        try:
            dias = int(str(dias_txt).strip())
        except ValueError:
            dias = 0
        if dias <= 0:
            _erro_importacao(erros_lista, linha, matricula, "Formato inválido em quantidade_dias.")
            continue
        srv = db.execute("SELECT nome FROM servidores WHERE matricula=? AND arquivado=0", (matricula,)).fetchone()
        if not srv:
            _erro_importacao(erros_lista, linha, matricula, "Matrícula não localizada em servidores ativos.")
            continue
        duplicado = db.execute("""
            SELECT 1 FROM eleicao_creditos
            WHERE matricula=? AND referencia_eleicao=? AND quantidade_dias=?
            LIMIT 1
        """, (matricula, referencia, dias)).fetchone()
        if duplicado:
            _erro_importacao(erros_lista, linha, matricula, "Lançamento de banco de dias de eleição já existente para esta referência.")
            continue
        observacao = _texto_importacao("Importação Banco de Dias de Eleição", f"Justificativa: {justificativa}", obs)
        cid = db.insert("""INSERT INTO eleicao_creditos
            (matricula,referencia_eleicao,quantidade_dias,observacao,criado_por)
            VALUES (?,?,?,?,?)""",
            (matricula, referencia, dias, observacao, session.get("nome")))
        payload.append({"acao": "criado_eleicao_credito", "id": cid, "matricula": matricula})
        _sucesso_importacao(sucessos, linha, matricula, f"{srv['nome'] or nome_csv}: {dias} dia(s) importado(s) na referência {referencia}")

    _registrar_relatorio_importacao("banco_dias_eleicao", arquivo.filename, total, sucessos, erros_lista, payload)
    return redirect(url_for("admin_importacao"))

@app.route("/admin/backup")
@master_required
def admin_backup():
    return render_template("admin/backup.html")

@app.route("/admin/backup/excel")
@master_required
def admin_backup_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_db()
    wb = Workbook()
    azul = "1A3A6B"
    claro = "D9EAF7"
    verde = "E7F5EE"
    amarelo = "FFF4D6"
    vermelho = "FDE2E1"
    cinza = "F8FAFC"
    borda = Side(style="thin", color="D0D7DE")

    def v(row, key, default=""):
        try:
            return row[key]
        except Exception:
            try:
                return row.get(key, default)
            except Exception:
                return default

    def add_sheet(nome, headers, rows):
        ws = wb.active if len(wb.sheetnames) == 1 and wb.active.max_row == 1 and wb.active["A1"].value is None else wb.create_sheet()
        ws.title = nome[:31]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
        title = ws.cell(1, 1, nome)
        title.font = Font(bold=True, color="FFFFFF", size=14)
        title.fill = PatternFill("solid", fgColor=azul)
        title.alignment = Alignment(horizontal="center")
        ws.append([])
        ws.append(headers)
        for c in ws[3]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=azul)
            c.alignment = Alignment(horizontal="center")
            c.border = Border(top=borda, left=borda, right=borda, bottom=borda)
        for row in rows:
            ws.append([_safe_excel(v) for v in row])
            row_idx = ws.max_row
            tipo = str(row[2] if len(row) > 2 else "")
            fill = None
            if tipo == "Lançamento" or tipo == "Crédito":
                fill = verde
            elif tipo in ("Compensação", "Pagamento", "Baixa"):
                fill = amarelo if tipo != "Pagamento" else vermelho
            elif row_idx % 2 == 0:
                fill = cinza
            for c in ws[row_idx]:
                c.border = Border(top=borda, left=borda, right=borda, bottom=borda)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                if fill:
                    c.fill = PatternFill("solid", fgColor=fill)
                if str(c.value or "").startswith("TOTAL"):
                    c.font = Font(bold=True)
                    c.fill = PatternFill("solid", fgColor=claro)
        ws.freeze_panes = "A4"
        if ws.max_row >= 3 and headers:
            ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{ws.max_row}"
        for col in range(1, len(headers) + 1):
            mx = max(len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max(mx + 2, 12), 55)
        return ws

    servidores_headers = ["Matrícula","Nome","CPF","E-mail","Cargo","Secretaria","Departamento","FG","Arquivado"]
    servidores_rows = lambda arq: [
        [s["matricula"], s["nome"], v(s, "cpf"), v(s, "email"), v(s, "cargo"),
         v(s, "secretaria"), v(s, "setor"), "Sim" if v(s, "funcao_gratificada", 0) else "Não", v(s, "arquivado", 0)]
        for s in db.execute("SELECT * FROM servidores WHERE arquivado=? ORDER BY nome", (arq,)).fetchall()
    ]
    add_sheet("Servidores Ativos", servidores_headers, servidores_rows(0))
    add_sheet("Servidores Arquivados", servidores_headers, servidores_rows(1))
    add_sheet("Cargos", ["Tipo","Nome","Ativo","Criado em"], [
        [r["tipo"], r["nome"], "Sim" if r["ativo"] else "Não", v(r, "criado_em")]
        for r in db.execute("SELECT * FROM cadastros_auxiliares ORDER BY tipo,nome").fetchall()
    ])
    add_sheet("Acessos concedidos", ["Nome","CPF","E-mail","Nível","Matrícula","Vínculos","Ativo","Último acesso"], [
        [u["nome"], u["cpf"], v(u, "email"), u["nivel"], v(u, "matricula"), ", ".join(get_vinculos(u)),
         "Sim" if u["ativo"] else "Não", v(u, "ultimo_acesso")]
        for u in db.execute("SELECT * FROM usuarios ORDER BY nivel,nome").fetchall()
    ])

    saldo_rows = []
    for s in db.execute("SELECT * FROM servidores ORDER BY arquivado,nome").fetchall():
        credito = db.execute("SELECT COALESCE(SUM(minutos_creditados),0) v FROM lancamentos WHERE matricula=?", (s["matricula"],)).fetchone()["v"] or 0
        usado = db.execute("""SELECT COALESCE(SUM(c.minutos),0) v FROM consumos c
                              JOIN lancamentos l ON l.id=c.lancamento_id WHERE l.matricula=?""", (s["matricula"],)).fetchone()["v"] or 0
        saldo_rows.append([s["matricula"], s["nome"], v(s, "secretaria"), v(s, "setor"),
                           "Arquivado" if v(s, "arquivado", 0) else "Ativo",
                           minutos_para_horas(credito), minutos_para_horas(usado), minutos_para_horas(credito - usado)])
    add_sheet("Banco de Horas", ["Matrícula","Servidor","Secretaria","Departamento","Status","Total creditado","Total baixado","Saldo"], saldo_rows)

    hist_bh = []
    for r in db.execute("""SELECT l.*,s.nome,s.secretaria,s.setor,s.arquivado FROM lancamentos l
                           LEFT JOIN servidores s ON s.matricula=l.matricula ORDER BY l.data,l.id""").fetchall():
        hist_bh.append([r["matricula"], v(r, "nome"), "Lançamento", r["data"], r["horas_base"], r["percentual"],
                        minutos_para_horas(r["minutos_creditados"]), "", "", v(r, "descricao"), v(r, "criado_em"),
                        "Arquivado" if v(r, "arquivado", 0) else "Ativo"])
    for r in db.execute("""SELECT c.*,s.nome,s.secretaria,s.setor,s.arquivado FROM compensacoes c
                           LEFT JOIN servidores s ON s.matricula=c.matricula ORDER BY c.data,c.id""").fetchall():
        hist_bh.append([r["matricula"], v(r, "nome"), "Compensação", r["data"], "", "",
                        "", minutos_para_horas(r["minutos_compensados"]), "", v(r, "descricao"), v(r, "criado_em"),
                        "Arquivado" if v(r, "arquivado", 0) else "Ativo"])
    for r in db.execute("""SELECT p.*,s.nome,s.arquivado,
                                  COALESCE(SUM(c.minutos),0) AS minutos_banco,
                                  COALESCE(SUM(ROUND(c.minutos*l.minutos_base*1.0/l.minutos_creditados)),0) AS base_paga
                           FROM pagamentos p
                           LEFT JOIN servidores s ON s.matricula=p.matricula
                           LEFT JOIN consumos c ON c.referencia_id=p.id AND c.tipo='pagamento'
                           LEFT JOIN lancamentos l ON l.id=c.lancamento_id
                           GROUP BY p.id,p.matricula,p.data_pagamento,p.descricao,p.criado_em,s.nome,s.arquivado
                           ORDER BY p.data_pagamento,p.id""").fetchall():
        hist_bh.append([r["matricula"], v(r, "nome"), "Pagamento", r["data_pagamento"], "", "",
                        "", minutos_para_horas(r["minutos_banco"] or 0), minutos_para_horas(r["base_paga"] or 0),
                        v(r, "descricao"), v(r, "criado_em"), "Arquivado" if v(r, "arquivado", 0) else "Ativo"])
    add_sheet("Histórico Banco de Horas", ["Matrícula","Servidor","Tipo","Data","H.Base","%","Crédito no Banco","Débito no Banco","H.Base Pagas","Justificativa/Descrição","Criado em","Status"], hist_bh)

    eleicao_saldo = []
    for s in db.execute("SELECT * FROM servidores ORDER BY arquivado,nome").fetchall():
        cred = db.execute("SELECT COALESCE(SUM(quantidade_dias),0) v FROM eleicao_creditos WHERE matricula=?", (s["matricula"],)).fetchone()["v"] or 0
        baix = db.execute("SELECT COUNT(*) v FROM eleicao_baixas WHERE matricula=?", (s["matricula"],)).fetchone()["v"] or 0
        eleicao_saldo.append([s["matricula"], s["nome"], v(s, "secretaria"), v(s, "setor"),
                              "Arquivado" if v(s, "arquivado", 0) else "Ativo", cred, baix, cred - baix])
    add_sheet("Banco Dias Eleição", ["Matrícula","Servidor","Secretaria","Departamento","Status","Dias concedidos","Dias utilizados","Saldo"], eleicao_saldo)

    hist_el = []
    for r in db.execute("""SELECT e.*,s.nome,s.arquivado FROM eleicao_creditos e
                           LEFT JOIN servidores s ON s.matricula=e.matricula ORDER BY e.criado_em,e.id""").fetchall():
        hist_el.append([r["matricula"], v(r, "nome"), "Crédito", r["referencia_eleicao"], r["quantidade_dias"], "",
                        v(r, "observacao"), v(r, "criado_por"), v(r, "criado_em"),
                        "Arquivado" if v(r, "arquivado", 0) else "Ativo"])
    for r in db.execute("""SELECT b.*,s.nome,s.arquivado FROM eleicao_baixas b
                           LEFT JOIN servidores s ON s.matricula=b.matricula ORDER BY b.data,b.id""").fetchall():
        hist_el.append([r["matricula"], v(r, "nome"), "Baixa", r["data"], "", 1,
                        v(r, "observacao"), v(r, "criado_por"), v(r, "criado_em"),
                        "Arquivado" if v(r, "arquivado", 0) else "Ativo"])
    add_sheet("Histórico Dias Eleição", ["Matrícula","Servidor","Tipo","Referência/Data","Dias concedidos","Dias utilizados","Justificativa/Observação","Usuário","Criado em","Status"], hist_el)

    add_sheet("Importações", ["ID","Tipo","Arquivo","Usuário","Total","Importados","Atualizados","Erros","Criado em","Estornado"], [
        [i["id"], i["tipo"], v(i, "arquivo"), v(i, "usuario_nome"), i["total_linhas"], i["criados"], i["atualizados"], i["erros"], i["criado_em"], "Sim" if i["estornado"] else "Não"]
        for i in db.execute("SELECT * FROM importacoes ORDER BY criado_em,id").fetchall()
    ])
    add_sheet("Auditoria", ["Data","Usuário","CPF","Ação","Entidade","ID","Matrícula","Servidor","Detalhe"], [
        [a["criado_em"], v(a, "usuario_nome"), v(a, "usuario_cpf"), a["acao"], a["entidade"],
         v(a, "entidade_id"), v(a, "matricula"), v(a, "servidor_nome"), v(a, "detalhe")]
        for a in db.execute("SELECT * FROM auditoria ORDER BY criado_em,id").fetchall()
    ])
    add_sheet("Visualizações", ["Data","Usuário","CPF","Tela","Caminho","Título"], [
        [vis["criado_em"], v(vis, "usuario_nome"), v(vis, "usuario_cpf"), v(vis, "endpoint"), v(vis, "caminho"), v(vis, "titulo")]
        for vis in db.execute("SELECT * FROM visualizacoes ORDER BY criado_em,id").fetchall()
    ])

    registrar_auditoria(db, "Gerou backup completo em Excel", "backup", date.today().isoformat(), None, None,
                        "Exportação completa do sistema em memória, sem arquivo temporário persistente.")
    db.commit()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f"attachment; filename=backup_banco_horas_{date.today().isoformat()}.xlsx"
    return resp

@app.route("/admin/importacao/cadastros", methods=["POST"])
@master_required
def admin_importar_cadastros_auxiliares():
    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        flash("Selecione um arquivo CSV.", "danger")
        return redirect(url_for("admin_importacao"))
    try:
        leitor = _csv_reader_upload(arquivo)
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("admin_importacao"))
    campos = {str(c or "").strip().lower() for c in (leitor.fieldnames or [])}
    if not {"tipo", "nome"}.issubset(campos):
        flash("CSV inválido. Campos obrigatórios: tipo e nome.", "danger")
        return redirect(url_for("admin_importacao"))

    db = get_db()
    total = 0
    payload, sucessos, erros_lista = [], [], []
    vistos = set()
    for row in leitor:
        total += 1
        linha = total + 1
        tipo = _normalizar_tipo_cadastro(_valor_csv(row, "tipo"))
        nome = _valor_csv(row, "nome", "descrição", "descricao")
        if tipo not in CADASTROS_TIPOS or not nome:
            _erro_importacao(erros_lista, linha, nome, "Campo obrigatório não preenchido ou tipo inválido.")
            continue
        chave = (tipo, nome.strip().lower())
        if chave in vistos:
            _erro_importacao(erros_lista, linha, nome, "Registro duplicado na própria planilha.")
            continue
        vistos.add(chave)

        atual = db.execute("SELECT * FROM cadastros_auxiliares WHERE tipo=? AND nome=?", (tipo, nome)).fetchone()
        if atual:
            _erro_importacao(erros_lista, linha, nome, f"{CADASTROS_TIPOS[tipo]['label'][:-1]} já existente. O cadastro não foi alterado.")
            continue
        cid = db.insert(
            "INSERT INTO cadastros_auxiliares (tipo,nome,ativo) VALUES (?,?,1)",
            (tipo, nome)
        )
        payload.append({"acao": "criado", "antes": None, "depois": {"id": cid, "tipo": tipo, "nome": nome, "ativo": 1}})
        _sucesso_importacao(sucessos, linha, nome, f"Cadastro criado em {tipo}: {nome}")

    _registrar_relatorio_importacao("cadastros_auxiliares", arquivo.filename, total, sucessos, erros_lista, payload)
    limpar_cache()
    return redirect(url_for("admin_importacao"))

@app.route("/admin/importacao/<int:importacao_id>/estornar", methods=["POST"])
@master_required
def admin_estornar_importacao(importacao_id):
    db = get_db()
    imp = db.execute("SELECT * FROM importacoes WHERE id=?", (importacao_id,)).fetchone()
    if not imp:
        flash("Importação não encontrada.", "danger")
        return redirect(url_for("admin_importacao"))
    if imp["estornado"]:
        flash("Esta importação já foi estornada.", "warning")
        return redirect(url_for("admin_importacao"))
    payload = json.loads(imp["payload"] or "[]")
    restaurados = removidos = 0
    if imp["tipo"] == "cadastros_auxiliares":
        for item in reversed(payload):
            acao = item.get("acao")
            depois = item.get("depois") or {}
            antes = item.get("antes")
            tipo = depois.get("tipo") or (antes or {}).get("tipo")
            nome = depois.get("nome") or (antes or {}).get("nome")
            if not tipo or not nome:
                continue
            if acao == "criado":
                row = db.execute("SELECT id FROM cadastros_auxiliares WHERE tipo=? AND nome=?", (tipo, nome)).fetchone()
                if row and _cadastro_auxiliar_em_uso(db, tipo, nome):
                    db.execute("UPDATE cadastros_auxiliares SET ativo=0 WHERE id=?", (row["id"],))
                elif row:
                    db.execute("DELETE FROM cadastros_auxiliares WHERE id=?", (row["id"],))
                removidos += 1
            elif acao == "reativado" and antes:
                db.execute("UPDATE cadastros_auxiliares SET ativo=? WHERE tipo=? AND nome=?",
                           (antes.get("ativo", 0), tipo, nome))
                restaurados += 1
    elif imp["tipo"] == "banco_horas":
        for item in reversed(payload):
            if item.get("acao") == "criado_lancamento" and item.get("id"):
                db.execute("DELETE FROM consumos WHERE lancamento_id=?", (item["id"],))
                db.execute("DELETE FROM lancamentos WHERE id=?", (item["id"],))
                removidos += 1
    elif imp["tipo"] == "banco_dias_eleicao":
        for item in reversed(payload):
            if item.get("acao") == "criado_eleicao_credito" and item.get("id"):
                db.execute("DELETE FROM eleicao_creditos WHERE id=?", (item["id"],))
                removidos += 1
    else:
        for item in reversed(payload):
            acao = item.get("acao")
            depois = item.get("depois") or {}
            antes = item.get("antes")
            matricula = depois.get("matricula") or (antes or {}).get("matricula")
            if not matricula:
                continue
            if acao == "criado":
                db.execute("DELETE FROM servidores WHERE matricula=?", (matricula,))
                removidos += 1
            elif acao == "atualizado" and antes:
                db.execute("""UPDATE servidores
                              SET nome=?,cpf=?,email=?,cargo=?,secretaria=?,setor=?,funcao_gratificada=?,arquivado=?
                              WHERE matricula=?""",
                           (antes.get("nome"), antes.get("cpf"), antes.get("email"), antes.get("cargo"),
                            antes.get("secretaria"), antes.get("setor"), antes.get("funcao_gratificada", 0),
                            antes.get("arquivado", 0), matricula))
                restaurados += 1
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    db.execute("UPDATE importacoes SET estornado=1, estornado_em=? WHERE id=?", (agora, importacao_id))
    registrar_auditoria(db, "Estornou importação", "importacao", importacao_id, None, None,
                        f"Importação {importacao_id}; removidos {removidos}; restaurados {restaurados}.")
    db.commit()
    limpar_cache()
    flash(f"Importação estornada: {removidos} removido(s), {restaurados} restaurado(s).", "warning")
    return redirect(url_for("admin_importacao"))

# ─── Admin: Cadastros auxiliares ─────────────────────────────────────────────

CADASTROS_TIPOS = {
    "secretaria": {"label": "Secretarias", "icone": "bi-building", "campo_servidor": "secretaria"},
    "departamento": {"label": "Locais de trabalho / Departamentos", "icone": "bi-diagram-3", "campo_servidor": "setor"},
    "cargo": {"label": "Cargos", "icone": "bi-briefcase", "campo_servidor": "cargo"},
}

def _sincronizar_cadastro_auxiliar(db, tipo, antigo, novo):
    if antigo == novo:
        return
    campo = CADASTROS_TIPOS[tipo]["campo_servidor"]
    db.execute(f"UPDATE servidores SET {campo}=? WHERE {campo}=?", (novo, antigo))
    if tipo == "secretaria":
        db.execute("UPDATE usuarios SET secretaria=? WHERE secretaria=?", (novo, antigo))
        db.execute("UPDATE pre_autorizacoes SET secretaria=? WHERE secretaria=?", (novo, antigo))
    elif tipo == "departamento":
        db.execute("UPDATE usuarios SET setor=? WHERE setor=?", (novo, antigo))
        db.execute("UPDATE pre_autorizacoes SET setor=? WHERE setor=?", (novo, antigo))

    if tipo in ("secretaria", "departamento"):
        for tabela in ("usuarios", "pre_autorizacoes"):
            for row in db.execute(f"SELECT id,vinculos FROM {tabela} WHERE vinculos LIKE ?", (f"%{antigo}%",)).fetchall():
                vinculos = get_vinculos(row)
                novos = [novo if v == antigo else v for v in vinculos]
                if novos != vinculos:
                    db.execute(f"UPDATE {tabela} SET vinculos=? WHERE id=?", (json.dumps(novos, ensure_ascii=False), row["id"]))

def _cadastro_auxiliar_em_uso(db, tipo, nome):
    campo = CADASTROS_TIPOS[tipo]["campo_servidor"]
    return bool(db.execute(f"SELECT 1 FROM servidores WHERE {campo}=? LIMIT 1", (nome,)).fetchone())

@app.route("/admin/cadastros")
@master_required
def admin_cadastros():
    db = get_db()
    dados = {}
    for tipo in CADASTROS_TIPOS:
        rows = db.execute(
            "SELECT * FROM cadastros_auxiliares WHERE tipo=? ORDER BY ativo DESC, nome",
            (tipo,)
        ).fetchall()
        dados[tipo] = rows
    return render_template("admin/cadastros.html", tipos=CADASTROS_TIPOS, dados=dados)

@app.route("/admin/cadastros/novo", methods=["POST"])
@master_required
def admin_cadastro_aux_novo():
    db = get_db()
    tipo = request.form.get("tipo", "").strip()
    nome = request.form.get("nome", "").strip()
    if tipo not in CADASTROS_TIPOS or not nome:
        flash("Informe um cadastro válido.", "danger")
        return redirect(url_for("admin_cadastros"))
    _garantir_cadastro_auxiliar(db, tipo, nome)
    registrar_auditoria(db, "Criou cadastro auxiliar", "cadastro_auxiliar", tipo, None, None, nome)
    db.commit()
    limpar_cache()
    flash("Cadastro incluído.", "success")
    return redirect(url_for("admin_cadastros"))

@app.route("/admin/cadastros/<int:cid>/editar", methods=["POST"])
@master_required
def admin_cadastro_aux_editar(cid):
    db = get_db()
    row = db.execute("SELECT * FROM cadastros_auxiliares WHERE id=?", (cid,)).fetchone()
    if not row:
        flash("Cadastro não encontrado.", "danger")
        return redirect(url_for("admin_cadastros"))
    novo = request.form.get("nome", "").strip()
    if not novo:
        flash("Nome obrigatório.", "danger")
        return redirect(url_for("admin_cadastros"))
    existente = db.execute(
        "SELECT id FROM cadastros_auxiliares WHERE tipo=? AND nome=? AND id<>?",
        (row["tipo"], novo, cid)
    ).fetchone()
    if existente:
        flash("Já existe um cadastro com este nome.", "warning")
        return redirect(url_for("admin_cadastros"))
    antigo = row["nome"]
    db.execute("UPDATE cadastros_auxiliares SET nome=?, ativo=1 WHERE id=?", (novo, cid))
    _sincronizar_cadastro_auxiliar(db, row["tipo"], antigo, novo)
    registrar_auditoria(db, "Editou cadastro auxiliar", "cadastro_auxiliar", cid, None, None, f"{antigo} -> {novo}")
    db.commit()
    limpar_cache()
    flash("Cadastro atualizado e servidores vinculados sincronizados.", "success")
    return redirect(url_for("admin_cadastros"))

@app.route("/admin/cadastros/<int:cid>/toggle", methods=["POST"])
@master_required
def admin_cadastro_aux_toggle(cid):
    db = get_db()
    row = db.execute("SELECT * FROM cadastros_auxiliares WHERE id=?", (cid,)).fetchone()
    if not row:
        flash("Cadastro não encontrado.", "danger")
        return redirect(url_for("admin_cadastros"))
    novo = 0 if row["ativo"] else 1
    db.execute("UPDATE cadastros_auxiliares SET ativo=? WHERE id=?", (novo, cid))
    registrar_auditoria(db, "Alterou status de cadastro auxiliar", "cadastro_auxiliar", cid, None, None,
                        f"{row['nome']} => {'ativo' if novo else 'inativo'}")
    db.commit()
    limpar_cache()
    flash(f"Cadastro {'ativado' if novo else 'inativado'}.", "success" if novo else "warning")
    return redirect(url_for("admin_cadastros"))

@app.route("/admin/cadastros/<int:cid>/excluir", methods=["POST"])
@master_required
def admin_cadastro_aux_excluir(cid):
    db = get_db()
    row = db.execute("SELECT * FROM cadastros_auxiliares WHERE id=?", (cid,)).fetchone()
    if not row:
        flash("Cadastro não encontrado.", "danger")
        return redirect(url_for("admin_cadastros"))
    if _cadastro_auxiliar_em_uso(db, row["tipo"], row["nome"]):
        db.execute("UPDATE cadastros_auxiliares SET ativo=0 WHERE id=?", (cid,))
        msg = "Cadastro está em uso e foi apenas inativado."
        cat = "warning"
    else:
        db.execute("DELETE FROM cadastros_auxiliares WHERE id=?", (cid,))
        msg = "Cadastro excluído."
        cat = "success"
    registrar_auditoria(db, "Excluiu cadastro auxiliar", "cadastro_auxiliar", cid, None, None, row["nome"])
    db.commit()
    limpar_cache()
    flash(msg, cat)
    return redirect(url_for("admin_cadastros"))

@app.route("/relatorios")
@master_required
def relatorios():
    db  = get_db()
    aba = request.args.get("aba","saldos")
    fmt_out = request.args.get("fmt","html")
    mat = request.args.get("matricula","").strip()
    servidor_busca = request.args.get("servidor","").strip()
    sec = request.args.get("secretaria","").strip()
    set_= request.args.get("setor","").strip()
    cargo = request.args.get("cargo","").strip()
    di  = request.args.get("data_ini","")
    df  = request.args.get("data_fim","")
    mes = request.args.get("mes","")
    ano = request.args.get("ano",str(date.today().year))
    agr = request.args.get("agrupar","servidor")

    if servidor_busca and " - " in servidor_busca:
        mat = servidor_busca.split(" - ", 1)[0].strip()
        servidor_busca = ""

    filt_srv, params_srv = _filtro_servidores(servidor_busca, sec, set_, cargo, arquivado=0)
    if mat: filt_srv += " AND s.matricula=?"; params_srv.append(mat)

    secs, sets = _listas_filtro(db, 0)
    cargos = _cadastros_nomes(db, "cargo")
    srvs_lista = db.execute("SELECT * FROM servidores WHERE arquivado=0 ORDER BY nome").fetchall()
    filtros = {"matricula":mat,"servidor":servidor_busca,"secretaria":sec,"setor":set_,"cargo":cargo,"data_ini":di,"data_fim":df,"mes":mes,"ano":ano,"agrupar":agr}
    filtros_qs = urlencode({k:v for k,v in filtros.items() if v})
    data = {}

    if aba == "saldos":
        data["servidores"] = db.execute(f"""
            SELECT s.*,
                (SELECT COALESCE(SUM(minutos_creditados),0) FROM lancamentos WHERE matricula=s.matricula) AS total_credito,
                (SELECT COALESCE(SUM(c.minutos),0) FROM consumos c JOIN lancamentos l ON l.id=c.lancamento_id WHERE l.matricula=s.matricula AND c.tipo='compensacao') AS total_compensado,
                (SELECT COALESCE(SUM(c.minutos),0) FROM consumos c JOIN lancamentos l ON l.id=c.lancamento_id WHERE l.matricula=s.matricula AND c.tipo='pagamento') AS total_pago,
                (SELECT COALESCE(SUM(minutos_creditados),0) FROM lancamentos WHERE matricula=s.matricula)
                -(SELECT COALESCE(SUM(c.minutos),0) FROM consumos c JOIN lancamentos l ON l.id=c.lancamento_id WHERE l.matricula=s.matricula) AS saldo
            FROM servidores s {filt_srv} ORDER BY s.nome""", params_srv).fetchall()
        data["grupos"] = _agrupar_itens(data["servidores"], agr)
        if fmt_out in ("csv", "xlsx", "pdf"):
            headers = ["Grupo","Matrícula","Nome","CPF","Email","Secretaria","Departamento","Cargo","Total Creditado","Total Compensado","Total Pago","Saldo","FG"]
            rows = []
            geral = {"credito": 0, "comp": 0, "pago": 0, "saldo": 0, "qtd": 0}
            for grupo, itens in data["grupos"].items():
                subt = {"credito": 0, "comp": 0, "pago": 0, "saldo": 0}
                for s in itens:
                    credito = minutos_num(s["total_credito"]); comp = minutos_num(s["total_compensado"])
                    pago = minutos_num(s["total_pago"]); saldo = minutos_num(s["saldo"])
                    subt["credito"] += credito; subt["comp"] += comp; subt["pago"] += pago; subt["saldo"] += saldo
                    geral["credito"] += credito; geral["comp"] += comp; geral["pago"] += pago; geral["saldo"] += saldo; geral["qtd"] += 1
                    rows.append([grupo, s["matricula"], s["nome"], s["cpf"] or "", s["email"] or "", s["secretaria"] or "", s["setor"] or "", s["cargo"] or "",
                                 minutos_para_horas(credito), minutos_para_horas(comp), minutos_para_horas(pago), minutos_para_horas(saldo), "Sim" if s["funcao_gratificada"] else "Não"])
                rows.append([f"TOTAL DO GRUPO: {grupo}", "", f"{len(itens)} servidor(es)", "", "", "", "", "",
                             minutos_para_horas(subt["credito"]), minutos_para_horas(subt["comp"]), minutos_para_horas(subt["pago"]), minutos_para_horas(subt["saldo"]), ""])
                rows.append(["", "", "", "", "", "", "", "", "", "", "", "", ""])
            rows.append(["TOTAL GERAL", "", f"{geral['qtd']} servidor(es)", "", "", "", "", "",
                         minutos_para_horas(geral["credito"]), minutos_para_horas(geral["comp"]), minutos_para_horas(geral["pago"]), minutos_para_horas(geral["saldo"]), ""])
            return _export_response(fmt_out, "saldos", "Relatório de Saldos em Banco de Horas", headers, rows)

    elif aba == "historico":
        mats=[r["matricula"] for r in db.execute(f"SELECT s.matricula FROM servidores s {filt_srv}",params_srv).fetchall()]
        if not mats: data["eventos"]=[]
        else:
            ph=",".join("?"*len(mats)); ev=[]
            def apd(f,p,c):
                if di: f+=f" AND {c}>=?"; p.append(di)
                if df: f+=f" AND {c}<=?"; p.append(df)
                return f,p
            fl,pl=apd(f"WHERE l.matricula IN ({ph})",list(mats),"l.data")
            for r in db.execute(f"SELECT l.*,s.nome,s.secretaria,s.setor,s.cargo,s.funcao_gratificada FROM lancamentos l JOIN servidores s ON s.matricula=l.matricula {fl} ORDER BY l.data DESC",pl).fetchall():
                ev.append({**dict(r),"tipo_evento":"lancamento","data_ord":r["data"]})
            fc,pc=apd(f"WHERE c.matricula IN ({ph})",list(mats),"c.data")
            for r in db.execute(f"SELECT c.*,s.nome,s.secretaria,s.setor,s.cargo,s.funcao_gratificada FROM compensacoes c JOIN servidores s ON s.matricula=c.matricula {fc} ORDER BY c.data DESC",pc).fetchall():
                ev.append({**dict(r),"tipo_evento":"compensacao","data_ord":r["data"]})
            fp,pp=apd(f"WHERE p.matricula IN ({ph})",list(mats),"p.data_pagamento")
            for r in db.execute(f"""SELECT p.*,s.nome,s.secretaria,s.setor,s.cargo,s.funcao_gratificada,
                COALESCE(SUM(ROUND(c.minutos*l.minutos_base*1.0/l.minutos_creditados)),0) AS base_paga,
                COALESCE(SUM(c.minutos),0) AS minutos_pagos
                FROM pagamentos p JOIN servidores s ON s.matricula=p.matricula
                JOIN consumos c ON c.referencia_id=p.id AND c.tipo='pagamento'
                JOIN lancamentos l ON l.id=c.lancamento_id {fp}
                GROUP BY p.id,p.matricula,p.data_pagamento,p.descricao,p.criado_em,
                         s.nome,s.secretaria,s.setor,s.cargo,s.funcao_gratificada
                ORDER BY p.data_pagamento DESC""",pp).fetchall():
                ev.append({**dict(r),"tipo_evento":"pagamento","data_ord":r["data_pagamento"]})
            ev.sort(key=lambda x:x["data_ord"],reverse=True); data["eventos"]=ev
        data["grupos"] = _agrupar_itens(data["eventos"], agr)
        data["totais_historico"] = {}
        for grupo, eventos in data["grupos"].items():
            data["totais_historico"][grupo] = {
                "qtd": len(eventos),
                "lanc": sum(minutos_num(e.get("minutos_creditados")) for e in eventos if e.get("tipo_evento") == "lancamento"),
                "comp": sum(minutos_num(e.get("minutos_compensados")) for e in eventos if e.get("tipo_evento") == "compensacao"),
                "pago": sum(minutos_num(e.get("base_paga")) for e in eventos if e.get("tipo_evento") == "pagamento"),
            }
        if fmt_out in ("csv", "xlsx", "pdf"):
            headers = ["Grupo","Data","Matrícula","Servidor","Secretaria","Departamento","Cargo","FG","Tipo","Detalhes","H. Lançadas","H. Compensadas","H. Pagas Base"]
            rows = []
            geral = {"lanc": 0, "comp": 0, "pago": 0, "qtd": 0}
            for grupo, eventos in data["grupos"].items():
                subt = {"lanc": 0, "comp": 0, "pago": 0}
                for e in eventos:
                    fg_s="Sim" if e.get("funcao_gratificada") else "Não"
                    data_ev = e["data_pagamento"] if e["tipo_evento"] == "pagamento" else e["data"]
                    lanc = comp = pago = 0
                    if e["tipo_evento"]=="lancamento":
                        lanc = minutos_num(e["minutos_creditados"]); detalhes = f"{e['horas_base']} + {e['percentual']}%"; tipo_ev = "Lançamento"
                    elif e["tipo_evento"]=="compensacao":
                        comp = minutos_num(e["minutos_compensados"]); detalhes = "Horas informadas"; tipo_ev = "Compensação"
                    else:
                        pago = minutos_num(e["base_paga"]); detalhes = e.get("descricao",""); tipo_ev = "Pagamento Folha"
                    subt["lanc"] += lanc; subt["comp"] += comp; subt["pago"] += pago
                    geral["lanc"] += lanc; geral["comp"] += comp; geral["pago"] += pago; geral["qtd"] += 1
                    rows.append([grupo, data_ev, e["matricula"], e["nome"], e.get("secretaria",""), e.get("setor",""), e.get("cargo",""), fg_s, tipo_ev, detalhes,
                                 minutos_para_horas(lanc) if lanc else "", minutos_para_horas(comp) if comp else "", minutos_para_horas(pago) if pago else ""])
                rows.append([f"TOTAL DO GRUPO: {grupo}", "", f"{len(eventos)} evento(s)", "", "", "", "", "", "", "",
                             minutos_para_horas(subt["lanc"]), minutos_para_horas(subt["comp"]), minutos_para_horas(subt["pago"])])
                rows.append(["", "", "", "", "", "", "", "", "", "", "", "", ""])
            rows.append(["TOTAL GERAL", "", f"{geral['qtd']} evento(s)", "", "", "", "", "", "", "",
                         minutos_para_horas(geral["lanc"]), minutos_para_horas(geral["comp"]), minutos_para_horas(geral["pago"])])
            return _export_response(fmt_out, "historico", "Relatório de Histórico Completo", headers, rows)

    elif aba == "pagamentos":
        fp=f"WHERE p.matricula IN (SELECT matricula FROM servidores s {filt_srv})"; pp=list(params_srv)
        if di: fp+=" AND p.data_pagamento>=?"; pp.append(di)
        if df: fp+=" AND p.data_pagamento<=?"; pp.append(df)
        pags=db.execute(f"""SELECT p.*,s.nome,s.secretaria,s.setor,s.cargo,s.funcao_gratificada,
            COALESCE(SUM(ROUND(c.minutos*l.minutos_base*1.0/l.minutos_creditados)),0) AS base_paga,
            COALESCE(SUM(c.minutos),0) AS minutos_pagos
            FROM pagamentos p JOIN servidores s ON s.matricula=p.matricula
            LEFT JOIN consumos c ON c.referencia_id=p.id AND c.tipo='pagamento'
            LEFT JOIN lancamentos l ON l.id=c.lancamento_id {fp}
            GROUP BY p.id,p.matricula,p.data_pagamento,p.descricao,p.criado_em,
                     s.nome,s.secretaria,s.setor,s.cargo,s.funcao_gratificada
            ORDER BY p.data_pagamento DESC""",pp).fetchall()
        dets={p["id"]:db.execute("""SELECT l.data AS data_hora,l.horas_base,l.minutos_base,l.percentual,
            l.minutos_creditados,c.minutos AS minutos_consumidos,
            ROUND(c.minutos*l.minutos_base*1.0/l.minutos_creditados) AS base_paga
            FROM consumos c JOIN lancamentos l ON l.id=c.lancamento_id
            WHERE c.tipo='pagamento' AND c.referencia_id=? ORDER BY l.data ASC""",(p["id"],)).fetchall() for p in pags}
        data["pagamentos"]=pags; data["detalhes"]=dets
        data["grupos"] = _agrupar_itens(pags, agr)
        data["totais_pagamentos"] = {}
        for grupo, pags_grupo in data["grupos"].items():
            total = sum(minutos_num(d["base_paga"]) for p in pags_grupo for d in dets[p["id"]])
            data["totais_pagamentos"][grupo] = {"qtd": len(pags_grupo), "base": total}
        if fmt_out in ("csv", "xlsx", "pdf"):
            headers = ["Grupo","Pag.ID","Matrícula","Servidor","Secretaria","Departamento","Cargo","FG","Data Pagamento","Referência","Data Realização","H.Base Realizadas","%","H.Base Pagas"]
            rows = []
            total_geral = 0
            qtd_geral = 0
            for grupo, pags_grupo in data["grupos"].items():
                subtotal = 0
                qtd = 0
                for p in pags_grupo:
                    fg_s="Sim" if p.get("funcao_gratificada") else "Não"
                    for d in dets[p["id"]]:
                        base_paga = minutos_num(d["base_paga"])
                        subtotal += base_paga; total_geral += base_paga; qtd += 1; qtd_geral += 1
                        rows.append([grupo,p["id"],p["matricula"],p["nome"],p.get("secretaria",""),p.get("setor",""),p.get("cargo",""),fg_s,p["data_pagamento"],p.get("descricao",""),d["data_hora"],d["horas_base"],f"{d['percentual']}%",minutos_para_horas(base_paga)])
                rows.append([f"TOTAL DO GRUPO: {grupo}", "", f"{qtd} item(ns)", "", "", "", "", "", "", "", "", "", "", minutos_para_horas(subtotal)])
                rows.append(["", "", "", "", "", "", "", "", "", "", "", "", "", ""])
            rows.append(["TOTAL GERAL", "", f"{qtd_geral} item(ns)", "", "", "", "", "", "", "", "", "", "", minutos_para_horas(total_geral)])
            return _export_response(fmt_out, "pagamentos", "Relatório de Pagamentos Realizados", headers, rows)

    elif aba == "competencia":
        data.update({"grupos":{},"mes":mes,"ano":ano,"agrupar":agr,"meses":MESES_FULL})
        if mes and ano:
            mats=[r["matricula"] for r in db.execute(f"SELECT s.matricula FROM servidores s {filt_srv}",params_srv).fetchall()]
            if mats:
                ph=",".join("?"*len(mats))
                rows=db.execute(f"""SELECT l.*,s.nome,s.secretaria,s.setor,s.cargo,s.matricula AS mat,s.funcao_gratificada
                    FROM lancamentos l JOIN servidores s ON s.matricula=l.matricula
                    WHERE substr(l.data,6,2)=? AND substr(l.data,1,4)=? AND l.matricula IN ({ph})
                    ORDER BY s.secretaria,s.setor,s.nome,l.data""",[mes.zfill(2),ano]+mats).fetchall()
                grps={}
                for r in rows:
                    chave=(r["secretaria"] or "Sem Secretaria") if agr=="secretaria" else (r["setor"] or "Sem Departamento") if agr=="departamento" else (r["cargo"] or "Sem Cargo") if agr=="cargo" else f"{r['nome']} ({r['mat']})"
                    grps.setdefault(chave,[]).append(dict(r))
                data["grupos"]=grps
                data["totais_competencia"] = {
                    "base": sum(minutos_num(r["minutos_base"]) for r in rows),
                    "creditadas": sum(minutos_num(r["minutos_creditados"]) for r in rows),
                    "qtd": len(rows),
                }
        if fmt_out in ("csv", "xlsx", "pdf") and data["grupos"]:
            headers = ["Grupo","Matrícula","Servidor","Data","H.Base","%","H.Creditadas","Descrição"]
            rows = []
            total_geral_base = 0
            total_geral_creditado = 0
            total_geral_lancamentos = 0
            for g,its in data["grupos"].items():
                total_grupo_base = 0
                total_grupo_creditado = 0
                for r in its:
                    base = minutos_num(r["minutos_base"])
                    creditado = minutos_num(r["minutos_creditados"])
                    total_grupo_base += base
                    total_grupo_creditado += creditado
                    total_geral_base += base
                    total_geral_creditado += creditado
                    total_geral_lancamentos += 1
                    rows.append([g,r["matricula"],r["nome"],r["data"],r["horas_base"],f"{r['percentual']}%",minutos_para_horas(creditado),r["descricao"] or ""])
                rows.append([
                    f"TOTAL DO GRUPO: {g}", "", f"{len(its)} lançamento(s)", "",
                    minutos_para_horas(total_grupo_base), "", minutos_para_horas(total_grupo_creditado), ""
                ])
                rows.append(["", "", "", "", "", "", "", ""])
            rows.append([
                "TOTAL GERAL", "", f"{total_geral_lancamentos} lançamento(s)", "",
                minutos_para_horas(total_geral_base), "", minutos_para_horas(total_geral_creditado), ""
            ])
            return _export_response(fmt_out, f"competencia_{mes}_{ano}", f"Relatório de Horas por Competência {mes}/{ano}", headers, rows)

    elif aba == "vencimentos":
        rows_v = db.execute(f"""
            SELECT l.*, s.nome, s.secretaria, s.setor, s.cargo, s.matricula AS mat, s.funcao_gratificada,
                   l.minutos_creditados - COALESCE((SELECT SUM(c.minutos) FROM consumos c WHERE c.lancamento_id=l.id),0) AS saldo_creditado
            FROM lancamentos l
            JOIN servidores s ON s.matricula=l.matricula
            {filt_srv}
            ORDER BY l.data ASC, s.nome
        """, params_srv).fetchall()
        itens = []
        for r in rows_v:
            saldo_creditado = minutos_num(r["saldo_creditado"])
            if saldo_creditado <= 0:
                continue
            venc = _somar_meses_iso(r["data"], 6)
            if di and venc < di:
                continue
            if df and venc > df:
                continue
            d = dict(r)
            d["matricula"] = r["mat"]
            d["vencimento"] = venc
            d["saldo_base"] = round(saldo_creditado * minutos_num(r["minutos_base"]) / minutos_num(r["minutos_creditados"])) if minutos_num(r["minutos_creditados"]) else saldo_creditado
            itens.append(d)
        data["vencimentos"] = itens
        data["grupos"] = _agrupar_itens(itens, agr)
        if fmt_out in ("csv", "xlsx", "pdf"):
            headers = ["Grupo","Matrícula","Servidor","Secretaria","Departamento","Cargo","Data Realização","Vencimento","H.Base Original","%","Saldo Creditado","Saldo Base"]
            rows = []
            total_creditado = 0
            total_base = 0
            total_qtd = 0
            for grupo, grupo_itens in data["grupos"].items():
                sub_creditado = 0
                sub_base = 0
                for r in grupo_itens:
                    saldo_creditado = minutos_num(r["saldo_creditado"]); saldo_base = minutos_num(r["saldo_base"])
                    sub_creditado += saldo_creditado; sub_base += saldo_base
                    total_creditado += saldo_creditado; total_base += saldo_base; total_qtd += 1
                    rows.append([grupo,r["matricula"],r["nome"],r.get("secretaria",""),r.get("setor",""),r.get("cargo",""),r["data"],r["vencimento"],r["horas_base"],f"{r['percentual']}%",minutos_para_horas(saldo_creditado),minutos_para_horas(saldo_base)])
                rows.append([f"TOTAL DO GRUPO: {grupo}", "", f"{len(grupo_itens)} lançamento(s)", "", "", "", "", "", "", "", minutos_para_horas(sub_creditado), minutos_para_horas(sub_base)])
                rows.append(["", "", "", "", "", "", "", "", "", "", "", ""])
            rows.append(["TOTAL GERAL", "", f"{total_qtd} lançamento(s)", "", "", "", "", "", "", "", minutos_para_horas(total_creditado), minutos_para_horas(total_base)])
            return _export_response(fmt_out, "vencimentos", "Relatório de Horas a Vencer/Vencidas", headers, rows)

    return render_template("relatorios.html", aba=aba, data=data, fmt=minutos_para_horas,
                           secretarias=secs, setores=sets, cargos=cargos, servidores_lista=srvs_lista,
                           meses=MESES_FULL, filtros=filtros, filtros_qs=filtros_qs)

# â”€â”€â”€ Auth: Setup / Login / Logout â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route('/setup', methods=['GET','POST'])
def setup():
    db = get_db()
    if db.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0] > 0:
        return redirect(url_for('login'))
    if request.method == 'POST':
        cpf   = request.form['cpf'].strip()
        nome  = request.form['nome'].strip()
        email = request.form['email'].strip()
        senha = request.form['senha']
        conf  = request.form['confirmar']
        if senha != conf:
            flash("As senhas não coincidem.", "danger")
        elif len(senha) < 8:
            flash("A senha deve ter pelo menos 8 caracteres.", "danger")
        else:
            db.execute("""INSERT INTO usuarios (cpf,nome,email,senha_hash,nivel,ativo,senha_temporaria)
                          VALUES (?,?,?,?,'master',1,0)""",
                       (cpf, nome, email, generate_password_hash(senha)))
            db.commit()
            flash("Usuário master criado com sucesso! Faça login.", "success")
            return redirect(url_for('login'))
    return render_template('setup.html')


@app.route('/login', methods=['GET','POST'])
def login():
    db = get_db()
    # Sem usuários → setup
    if db.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0] == 0:
        return redirect(url_for('setup'))

    if request.method == 'POST':
        cpf   = request.form['cpf'].strip()
        senha = request.form['senha']
        u = db.execute("SELECT * FROM usuarios WHERE cpf=? AND ativo=1", (cpf,)).fetchone()

        # ── Primeiro acesso: auto-cria conta para servidores cadastrados ─────
        if not u:
            srv = db.execute(
                "SELECT * FROM servidores WHERE cpf=? AND arquivado=0", (cpf,)).fetchone()
            if srv:
                try:
                    pre = db.execute(
                        "SELECT * FROM pre_autorizacoes WHERE cpf=?", (cpf,)).fetchone()
                    nivel_auto     = pre['nivel']     if pre else 'servidor'
                    vinculos_auto  = get_vinculos(pre) if pre else []
                    matricula_auto = (pre['matricula'] or srv['matricula']) if pre else srv['matricula']
                    db.insert("""
                        INSERT INTO usuarios
                            (cpf, nome, email, senha_hash, nivel, matricula, vinculos, ativo, senha_temporaria)
                        VALUES (?,?,?,?,?,?,?,1,1)
                    """, (cpf, srv['nome'], srv.get('email') or '',
                          generate_password_hash('123456'),
                          nivel_auto, matricula_auto, json.dumps(vinculos_auto)))
                    db.commit()
                    u = db.execute(
                        "SELECT * FROM usuarios WHERE cpf=? AND ativo=1", (cpf,)).fetchone()
                except Exception:
                    db.commit()
                    u = db.execute(
                        "SELECT * FROM usuarios WHERE cpf=? AND ativo=1", (cpf,)).fetchone()

        if not u or not check_password_hash(u['senha_hash'], senha):
            msg_senha = (
                "CPF ou senha inválidos. Caso não tenha acesso ou não se recorde da senha, "
                "contate o Departamento de Gestão de Pessoas para emissão de senha temporária."
            )
            try:
                registrar_auditoria(db, "Falha de login", "seguranca", None, None, None,
                                    f"Tentativa de acesso com CPF {cpf}")
                db.commit()
            except Exception:
                pass
            if not u:
                flash(msg_senha, "danger")
            else:
                flash(msg_senha, "danger")
            return render_template('login.html')
        if not _usuario_tem_servidor_ativo(db, u):
            flash("Acesso inativo ou sem servidor ativo vinculado. Contate o RH.", "warning")
            return render_template('login.html')

        # Registra último acesso
        db.execute("UPDATE usuarios SET ultimo_acesso=? WHERE id=?",
                   (datetime.now().strftime('%Y-%m-%d %H:%M'), u['id']))
        db.commit()

        session.clear()
        session['uid']   = u['id']
        session['nivel'] = u['nivel']
        session['nome']  = u['nome']
        session['cpf']   = u['cpf']
        session['temp']  = bool(u['senha_temporaria'])
        session['vinculos'] = get_vinculos(u)
        if u['nivel'] == 'servidor': session['matricula'] = u['matricula']


        if u['senha_temporaria']:
            flash("Sua senha é temporária. Defina uma nova senha.", "warning")
            return redirect(url_for('trocar_senha'))

        return redirect(url_for('portal'))

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash("Sessão encerrada.", "info")
    return redirect(url_for('login'))


@app.route('/trocar-senha', methods=['GET','POST'])
def trocar_senha():
    if 'uid' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        atual = request.form['senha_atual']
        nova  = request.form['senha_nova']
        conf  = request.form['confirmar']
        db = get_db()
        u  = db.execute("SELECT * FROM usuarios WHERE id=?", (session['uid'],)).fetchone()
        if not check_password_hash(u['senha_hash'], atual):
            msg_atual = ("Senha temporária incorreta. Use <strong>123456</strong> no campo acima."
                         if u.get('senha_temporaria') else "Senha atual incorreta.")
            flash(msg_atual, "danger")
        elif nova != conf:
            flash("As novas senhas não coincidem.", "danger")
        elif nova == '123456':
            flash("Escolha uma senha diferente da senha temporária padrão.", "danger")
        elif len(nova) < 6:
            flash("A senha deve ter pelo menos 6 caracteres.", "danger")
        else:
            db.execute("UPDATE usuarios SET senha_hash=?, senha_temporaria=0 WHERE id=?",
                       (generate_password_hash(nova), session['uid']))
            db.commit()
            session['temp'] = False
            flash("Senha alterada com sucesso!", "success")
            return redirect(url_for('portal'))
    return render_template('trocar_senha.html')


@app.route('/meu-cadastro', methods=['GET','POST'])
def meu_cadastro():
    if 'uid' not in session:
        return redirect(url_for('login'))

    db = get_db()
    u = db.execute("SELECT * FROM usuarios WHERE id=?", (session['uid'],)).fetchone()
    if not u:
        session.clear()
        flash("Usuário não encontrado. Faça login novamente.", "warning")
        return redirect(url_for('login'))

    if request.method == 'POST':
        email = request.form.get('email','').strip()
        senha_atual = request.form.get('senha_atual','')
        senha_nova = request.form.get('senha_nova','')
        confirmar = request.form.get('confirmar','')

        alterar_senha = bool(senha_atual or senha_nova or confirmar)
        if alterar_senha:
            if not check_password_hash(u['senha_hash'], senha_atual):
                flash("Senha atual incorreta.", "danger")
                return render_template('meu_cadastro.html', usuario=u)
            if senha_nova != confirmar:
                flash("A nova senha e a confirmação não coincidem.", "danger")
                return render_template('meu_cadastro.html', usuario=u)
            if len(senha_nova) < 8:
                flash("A nova senha deve ter pelo menos 8 caracteres.", "danger")
                return render_template('meu_cadastro.html', usuario=u)
            db.execute(
                "UPDATE usuarios SET email=?, senha_hash=?, senha_temporaria=0 WHERE id=?",
                (email, generate_password_hash(senha_nova), session['uid'])
            )
            session['temp'] = False
            flash("Cadastro e senha atualizados com sucesso.", "success")
        else:
            db.execute("UPDATE usuarios SET email=? WHERE id=?", (email, session['uid']))
            flash("E-mail atualizado com sucesso.", "success")

        db.commit()
        return redirect(url_for('meu_cadastro'))

    return render_template('meu_cadastro.html', usuario=u)


@app.route('/recuperar-senha', methods=['GET','POST'])
def recuperar_senha():
    if request.method == 'POST':
        cpf = request.form['cpf'].strip()
        db  = get_db()
        u   = db.execute("SELECT * FROM usuarios WHERE cpf=? AND ativo=1", (cpf,)).fetchone()
        if u and u['email']:
            token  = secrets.token_urlsafe(32)
            expiry = (datetime.now() + timedelta(hours=1)).isoformat()
            db.execute("UPDATE usuarios SET reset_token=?, reset_expiry=? WHERE id=?",
                       (token, expiry, u['id']))
            db.commit()
            link = url_for('recuperar_senha_token', token=token, _external=True)
            html = f"""<p>Olá, <b>{u['nome']}</b>.</p>
            <p>Clique no link abaixo para redefinir sua senha (válido por 1 hora):</p>
            <p><a href="{link}">{link}</a></p>"""
            ok, err = enviar_email_smtp(u['email'], "Redefinição de senha — Banco de Horas Ibiporã", html)
            if ok:
                flash("E-mail de recuperação enviado! Verifique sua caixa de entrada.", "success")
            else:
                if err == "SMTP_NAO_CONFIGURADO":
                    flash(
                        "O envio de e-mail ainda não foi configurado. Solicite ao RH a redefinição de senha pelo painel "
                        "Admin > Usuários, ou configure o SMTP em Admin > E-mail.",
                        "warning"
                    )
                else:
                    flash(f"Não foi possível enviar o e-mail ({err}). Solicite ao administrador.", "warning")
        else:
            # CPF não encontrado ou sem e-mail — não revelar
            flash("Se o CPF estiver cadastrado com e-mail, você receberá as instruções.", "info")
        return redirect(url_for('login'))
    return render_template('recuperar_senha.html')


@app.route('/recuperar-senha/<token>', methods=['GET','POST'])
def recuperar_senha_token(token):
    db = get_db()
    u  = db.execute("SELECT * FROM usuarios WHERE reset_token=? AND ativo=1", (token,)).fetchone()
    if not u:
        flash("Link inválido ou já utilizado.", "danger")
        return redirect(url_for('login'))
    if datetime.now() > datetime.fromisoformat(u['reset_expiry']):
        flash("Link expirado. Solicite um novo.", "danger")
        return redirect(url_for('recuperar_senha'))
    if request.method == 'POST':
        nova = request.form['senha_nova']
        conf = request.form['confirmar']
        if nova != conf:
            flash("As senhas não coincidem.", "danger")
        elif len(nova) < 8:
            flash("Mínimo de 8 caracteres.", "danger")
        else:
            db.execute("UPDATE usuarios SET senha_hash=?, senha_temporaria=0, reset_token=NULL, reset_expiry=NULL WHERE id=?",
                       (generate_password_hash(nova), u['id']))
            db.commit()
            flash("Senha redefinida! Faça login.", "success")
            return redirect(url_for('login'))
    return render_template('recuperar_senha_token.html', token=token)


# â”€â”€â”€ Portal (redireciona por nível) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route('/portal')
def portal():
    nivel = session.get('nivel')
    if nivel == 'master':    return redirect(url_for('dashboard'))
    if nivel == 'servidor':  return redirect(url_for('meu_banco'))
    return redirect(url_for('consulta'))


# â”€â”€â”€ Meu Banco (Servidor) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route('/meu-banco')
def meu_banco():
    if 'uid' not in session or session.get('nivel') != 'servidor':
        return redirect(url_for('login'))
    mat = session.get('matricula')
    if not mat:
        flash("Matrícula não vinculada ao seu usuário. Contate o RH.", "danger")
        return redirect(url_for('login'))
    db  = get_db()
    srv = db.execute("SELECT * FROM servidores WHERE matricula=? AND arquivado=0", (mat,)).fetchone()
    if not srv:
        flash("Servidor não encontrado.", "danger")
        return redirect(url_for('login'))
    saldo = calcular_saldo(db, mat)
    lancs = db.execute("""
        SELECT l.*, COALESCE((SELECT SUM(c.minutos) FROM consumos c WHERE c.lancamento_id=l.id),0) AS consumido
        FROM lancamentos l WHERE l.matricula=? ORDER BY l.data DESC""", (mat,)).fetchall()
    comps = db.execute("SELECT * FROM compensacoes WHERE matricula=? ORDER BY data DESC", (mat,)).fetchall()
    pags  = db.execute("""
        SELECT p.*, COALESCE(SUM(ROUND(c.minutos*l.minutos_base*1.0/l.minutos_creditados)),0) AS base_paga
        FROM pagamentos p
        LEFT JOIN consumos c ON c.referencia_id=p.id AND c.tipo='pagamento'
        LEFT JOIN lancamentos l ON l.id=c.lancamento_id
        WHERE p.matricula=?
        GROUP BY p.id,p.matricula,p.data_pagamento,p.descricao,p.criado_em
        ORDER BY p.data_pagamento DESC""", (mat,)).fetchall()
    el_creditos = db.execute(
        "SELECT * FROM eleicao_creditos WHERE matricula=? ORDER BY criado_em DESC", (mat,)).fetchall()
    el_baixas = db.execute(
        "SELECT * FROM eleicao_baixas WHERE matricula=? ORDER BY data DESC", (mat,)).fetchall()
    saldo_eleicao = calcular_saldo_eleicao(db, mat)
    p_sol = max(1, int(request.args.get('p_sol', 1) or 1))
    sol_total = db.execute("SELECT COUNT(*) FROM solicitacoes WHERE matricula=?", (mat,)).fetchone()[0] or 0
    p_sol, sol_pages, sol_offset = _paginar(sol_total, p_sol)
    solicitacoes_srv = db.execute(
        "SELECT * FROM solicitacoes WHERE matricula=? ORDER BY criado_em DESC LIMIT 20 OFFSET ?",
        (mat, sol_offset)
    ).fetchall()
    sol_pendentes_comp = db.execute(
        "SELECT * FROM solicitacoes WHERE matricula=? AND status IN ('solicitado','autorizado') ORDER BY criado_em ASC",
        (mat,)
    ).fetchall()
    return render_template('meu_banco.html', servidor=srv, saldo=saldo,
                           lancamentos=lancs, compensacoes=comps, pagamentos=pags,
                           el_creditos=el_creditos, el_baixas=el_baixas,
                           saldo_eleicao=saldo_eleicao,
                           solicitacoes_srv=solicitacoes_srv,
                           sol_pendentes_comp=sol_pendentes_comp,
                           sol_total=sol_total, p_sol=p_sol, sol_pages=sol_pages,
                           solicitacoes_habilitado=_solicitacoes_habilitado(db),
                           status_display=_status_display,
                           fmt=minutos_para_horas)


# â”€â”€â”€ Consulta (Secretário / Chefia) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route('/consulta')
def consulta():
    if 'uid' not in session or session.get('nivel') not in ('secretario','chefia'):
        return redirect(url_for('login'))
    db    = get_db()
    nivel = session.get('nivel')
    busca = request.args.get('busca','').strip()
    sec = request.args.get('secretaria','').strip()
    set_ = request.args.get('setor','').strip()

    vinculos = session.get('vinculos', [])
    matricula_propria = session.get('matricula') or ''
    filtro = "WHERE s.arquivado=0"
    params = []

    acesso_sql, acesso_params = filtro_consulta_vinculos(nivel, vinculos, matricula_propria)
    filtro += acesso_sql
    params.extend(acesso_params)

    if busca:
        filtro += " AND (s.matricula LIKE ? OR s.nome LIKE ?)"; params += [f"%{busca}%",f"%{busca}%"]
    if sec:
        filtro += " AND s.secretaria=?"; params.append(sec)
    if set_:
        filtro += " AND s.setor=?"; params.append(set_)

    lista = db.execute(f"""
        SELECT s.*,
            (SELECT COALESCE(SUM(minutos_creditados),0) FROM lancamentos WHERE matricula=s.matricula)
            -(SELECT COALESCE(SUM(c.minutos),0) FROM consumos c JOIN lancamentos l ON l.id=c.lancamento_id WHERE l.matricula=s.matricula)
            AS saldo_minutos,
            (SELECT COALESCE(SUM(quantidade_dias),0) FROM eleicao_creditos WHERE matricula=s.matricula)
            -(SELECT COUNT(*) FROM eleicao_baixas WHERE matricula=s.matricula)
            AS saldo_eleicao
        FROM servidores s {filtro} ORDER BY s.nome""", params).fetchall()
    filtro_base = "WHERE s.arquivado=0"
    params_base = []
    acesso_base_sql, acesso_base_params = filtro_consulta_vinculos(nivel, vinculos, matricula_propria)
    filtro_base += acesso_base_sql
    params_base.extend(acesso_base_params)
    secretarias = [r[0] for r in db.execute(f"SELECT DISTINCT s.secretaria FROM servidores s {filtro_base} AND s.secretaria IS NOT NULL AND s.secretaria!='' ORDER BY s.secretaria", params_base).fetchall()]
    setores = [r[0] for r in db.execute(f"SELECT DISTINCT s.setor FROM servidores s {filtro_base} AND s.setor IS NOT NULL AND s.setor!='' ORDER BY s.setor", params_base).fetchall()]

    uid = session.get('uid')
    campo = 'secretaria' if nivel == 'secretario' else 'setor'

    # Solicitações pendentes de autorização (paginadas)
    p_pend = max(1, int(request.args.get('p_pend', 1) or 1))
    sol_pendentes = []
    sol_pendentes_total = 0
    sol_pendentes_pages = 1
    if vinculos:
        ph = ','.join('?' * len(vinculos))
        sol_pendentes_total = db.execute(
            f"""SELECT COUNT(*) FROM solicitacoes sol JOIN servidores s ON s.matricula=sol.matricula
                WHERE sol.status='solicitado' AND s.{campo} IN ({ph}) AND sol.criado_por_uid != ?""",
            vinculos + [uid]
        ).fetchone()[0] or 0
        p_pend, sol_pendentes_pages, offset_pend = _paginar(sol_pendentes_total, p_pend)
        sol_pendentes = db.execute(
            f"""SELECT sol.*, s.nome AS servidor_nome, s.secretaria, s.setor
                FROM solicitacoes sol JOIN servidores s ON s.matricula=sol.matricula
                WHERE sol.status='solicitado' AND s.{campo} IN ({ph}) AND sol.criado_por_uid != ?
                ORDER BY sol.criado_em ASC LIMIT 20 OFFSET ?""",
            vinculos + [uid, offset_pend]
        ).fetchall()

    # Solicitações autorizadas nos últimos 60 dias (paginadas)
    p_aut = max(1, int(request.args.get('p_aut', 1) or 1))
    sol_autorizadas = []
    sol_aut_total = 0
    sol_aut_pages = 1
    data_60_dias = (date.today() - timedelta(days=60)).isoformat()
    if vinculos:
        ph = ','.join('?' * len(vinculos))
        sol_aut_total = db.execute(
            f"""SELECT COUNT(*) FROM solicitacoes sol JOIN servidores s ON s.matricula=sol.matricula
                WHERE sol.status IN ('autorizado','lancado','indeferido')
                AND s.{campo} IN ({ph}) AND sol.data_autorizacao >= ?""",
            vinculos + [data_60_dias]
        ).fetchone()[0] or 0
        p_aut, sol_aut_pages, offset_aut = _paginar(sol_aut_total, p_aut)
        sol_autorizadas = db.execute(
            f"""SELECT sol.*, s.nome AS servidor_nome, s.secretaria, s.setor
                FROM solicitacoes sol JOIN servidores s ON s.matricula=sol.matricula
                WHERE sol.status IN ('autorizado','lancado','indeferido')
                AND s.{campo} IN ({ph}) AND sol.data_autorizacao >= ?
                ORDER BY sol.data_autorizacao DESC LIMIT 20 OFFSET ?""",
            vinculos + [data_60_dias, offset_aut]
        ).fetchall()

    aba = request.args.get('aba', '')
    titulo_vinculos = ' | '.join(vinculos) if vinculos else '(sem vínculo)'
    return render_template('consulta.html', servidores=lista, fmt=minutos_para_horas,
                           busca=busca, nivel=nivel, secretarias=secretarias, setores=setores,
                           secretaria_sel=sec, setor_sel=set_,
                           sol_pendentes=sol_pendentes,
                           sol_pendentes_total=sol_pendentes_total,
                           sol_pendentes_pages=sol_pendentes_pages,
                           p_pend=p_pend,
                           sol_autorizadas=sol_autorizadas,
                           sol_aut_total=sol_aut_total,
                           sol_aut_pages=sol_aut_pages,
                           p_aut=p_aut,
                           aba=aba,
                           status_display=_status_display,
                           titulo=f"Secretaria(s): {titulo_vinculos}" if nivel=='secretario'
                                  else f"Departamento(s): {titulo_vinculos}")


# â”€â”€â”€ Admin: Usuários â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route('/admin/usuarios')
@master_required
def admin_usuarios():
    db   = get_db()
    page = request.args.get('nivel','')
    busca = request.args.get('busca','').strip()
    q    = "WHERE u.ativo=1 AND (u.nivel='master' OR s.matricula IS NOT NULL)"
    p    = [page] if page else []
    if page:
        q += " AND u.nivel=?"
    if busca:
        q += " AND (u.nome LIKE ? OR u.cpf LIKE ? OR u.matricula LIKE ? OR s.nome LIKE ?)"
        like = f"%{busca}%"
        p.extend([like, like, like, like])
    rows = db.execute(f"""
        SELECT u.*, s.nome AS servidor_nome
        FROM usuarios u
        LEFT JOIN servidores s ON s.arquivado=0
          AND ((u.matricula IS NOT NULL AND u.matricula!='' AND s.matricula=u.matricula)
               OR (u.cpf IS NOT NULL AND u.cpf!='' AND s.cpf=u.cpf))
        {q}
        ORDER BY u.nivel,u.nome
    """, p).fetchall()
    usrs = []
    for u in rows:
        d = dict(u)
        d["vinculos_lista"] = get_vinculos(u)
        d["ultimas_visualizacoes"] = db.execute("""
            SELECT titulo,caminho,criado_em FROM visualizacoes
            WHERE usuario_id=?
            ORDER BY criado_em DESC, id DESC
            LIMIT 3
        """, (u["id"],)).fetchall()
        usrs.append(d)
    return render_template('admin/usuarios.html', usuarios=usrs, filtro_nivel=page, busca=busca)


@app.route('/admin/usuarios/novo', methods=['GET','POST'])
@master_required
def admin_novo_usuario():
    db = get_db()
    if request.method == 'POST':
        cpf   = request.form['cpf'].strip()
        nome  = request.form['nome'].strip()
        email = request.form['email'].strip()
        nivel = request.form['nivel']
        vinculos = request.form.getlist('vinculos')
        sec   = vinculos[0] if nivel == 'secretario' and vinculos else request.form.get('secretaria','').strip()
        set_  = vinculos[0] if nivel == 'chefia' and vinculos else request.form.get('setor','').strip()
        mat   = request.form.get('matricula','').strip()

        if db.execute("SELECT 1 FROM usuarios WHERE cpf=?", (cpf,)).fetchone():
            flash("CPF já cadastrado.", "danger")
        else:
            senha_temp = gerar_senha_temp()
            db.execute("""INSERT INTO usuarios (cpf,nome,email,senha_hash,nivel,secretaria,setor,matricula,vinculos,ativo,senha_temporaria)
                          VALUES (?,?,?,?,?,?,?,?,?,1,1)""",
                       (cpf, nome, email, generate_password_hash(senha_temp), nivel, sec, set_, mat, json.dumps(vinculos)))
            db.commit()
            flash(f"Usuário criado! Senha temporária: {senha_temp}", "success")
            if email:
                html = f"<p>Olá, <b>{nome}</b>!</p><p>Seu acesso foi criado no sistema Banco de Horas de Ibiporã.</p><p><b>CPF:</b> {cpf}<br><b>Senha temporária:</b> {senha_temp}</p><p>Altere a senha no primeiro acesso.</p>"
                ok, _ = enviar_email_smtp(email, "Acesso ao Banco de Horas — Ibiporã", html)
                if ok: flash("E-mail enviado ao usuário.", "info")
            return redirect(url_for('admin_usuarios'))

    srvs = db.execute("SELECT matricula,nome FROM servidores WHERE arquivado=0 ORDER BY nome").fetchall()
    secs = _cadastros_nomes(db, "secretaria")
    sets = _cadastros_nomes(db, "departamento")
    return render_template('admin/usuario_form.html', usuario=None,
                           servidores=srvs, secretarias=secs, setores=sets)


@app.route('/admin/usuarios/<int:uid>/editar', methods=['GET','POST'])
@master_required
def admin_editar_usuario(uid):
    db = get_db()
    u  = db.execute("SELECT * FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not u: flash("Usuário não encontrado.", "danger"); return redirect(url_for('admin_usuarios'))
    if request.method == 'POST':
        nome  = request.form['nome'].strip()
        email = request.form['email'].strip()
        nivel = request.form['nivel']
        vinculos = request.form.getlist('vinculos')
        sec   = vinculos[0] if nivel == 'secretario' and vinculos else request.form.get('secretaria','').strip()
        set_  = vinculos[0] if nivel == 'chefia' and vinculos else request.form.get('setor','').strip()
        mat   = request.form.get('matricula','').strip()
        db.execute("UPDATE usuarios SET nome=?,email=?,nivel=?,secretaria=?,setor=?,matricula=?,vinculos=? WHERE id=?",
                   (nome, email, nivel, sec, set_, mat, json.dumps(vinculos), uid))
        db.commit()
        flash("Usuário atualizado.", "success")
        return redirect(url_for('admin_usuarios'))
    srvs = db.execute("SELECT matricula,nome FROM servidores WHERE arquivado=0 ORDER BY nome").fetchall()
    secs = _cadastros_nomes(db, "secretaria")
    sets = _cadastros_nomes(db, "departamento")
    d = dict(u)
    d["vinculos_lista"] = get_vinculos(u)
    return render_template('admin/usuario_form.html', usuario=d,
                           servidores=srvs, secretarias=secs, setores=sets)


@app.route('/admin/usuarios/<int:uid>/toggle', methods=['POST'])
@master_required
def admin_toggle_usuario(uid):
    db = get_db()
    u  = db.execute("SELECT * FROM usuarios WHERE id=?", (uid,)).fetchone()
    if u:
        novo = 0 if u['ativo'] else 1
        if novo and not _usuario_tem_servidor_ativo(db, u):
            flash("Não é possível ativar este usuário sem servidor ativo vinculado.", "danger")
            return redirect(url_for('admin_usuarios'))
        db.execute("UPDATE usuarios SET ativo=? WHERE id=?", (novo, uid))
        db.commit()
        flash(f"Usuário {'ativado' if novo else 'desativado'}.", "success")
    return redirect(url_for('admin_usuarios'))


@app.route('/admin/usuarios/<int:uid>/reset-senha', methods=['POST'])
@master_required
def admin_reset_senha(uid):
    db   = get_db()
    u    = db.execute("SELECT * FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not u: flash("Usuário não encontrado.", "danger"); return redirect(url_for('admin_usuarios'))
    nova = gerar_senha_temp()
    db.execute("UPDATE usuarios SET senha_hash=?, senha_temporaria=1 WHERE id=?",
               (generate_password_hash(nova), uid))
    db.commit()
    flash(f"Senha de '{u['nome']}' redefinida. Senha temporária: {nova}", "warning")
    if u['email']:
        html = f"<p>Olá, <b>{u['nome']}</b>.</p><p>Sua senha foi redefinida pelo administrador.</p><p><b>Nova senha temporária:</b> {nova}</p><p>Altere no próximo acesso.</p>"
        ok, err = enviar_email_smtp(u['email'], "Senha redefinida — Banco de Horas Ibiporã", html)
        if ok:
            flash("E-mail enviado ao usuário com a nova senha.", "info")
        elif err == "SMTP_NAO_CONFIGURADO":
            flash("E-mail não enviado: SMTP não configurado. A senha temporária foi exibida acima para repasse manual.", "warning")
        else:
            flash(f"E-mail não enviado: {err}. A senha temporária foi exibida acima para repasse manual.", "warning")
    return redirect(url_for('admin_usuarios'))


@app.route('/admin/config-email', methods=['GET','POST'])
@master_required
def admin_config_email():
    db = get_db()
    if request.method == 'POST':
        cfg_atual = obter_config_smtp()
        for chave in ['smtp_host','smtp_port','smtp_user','smtp_from','smtp_tls']:
            valor = request.form.get(chave,'').strip()
            db.upsert(
                "INSERT OR REPLACE INTO config (chave,valor) VALUES (?,?)",
                "INSERT INTO config (chave,valor) VALUES (?,?) ON CONFLICT (chave) DO UPDATE SET valor=EXCLUDED.valor",
                (chave, valor)
            )
        senha = request.form.get('smtp_pass','').strip()
        if senha:
            db.upsert(
                "INSERT OR REPLACE INTO config (chave,valor) VALUES (?,?)",
                "INSERT INTO config (chave,valor) VALUES (?,?) ON CONFLICT (chave) DO UPDATE SET valor=EXCLUDED.valor",
                ('smtp_pass', senha)
            )
        elif not cfg_atual.get('smtp_pass'):
            db.upsert(
                "INSERT OR REPLACE INTO config (chave,valor) VALUES (?,?)",
                "INSERT INTO config (chave,valor) VALUES (?,?) ON CONFLICT (chave) DO UPDATE SET valor=EXCLUDED.valor",
                ('smtp_pass', '')
            )
        db.commit()
        flash("Configurações de e-mail salvas.", "success")
        # Teste de envio
        email_teste = request.form.get('email_teste','').strip()
        if email_teste:
            ok, err = enviar_email_smtp(email_teste, "Teste SMTP — Banco de Horas Ibiporã",
                                        "<p>Configuração de e-mail funcionando corretamente!</p>")
            flash(f"Teste: {'✅ E-mail enviado com sucesso!' if ok else f'❌ Falha: {err}'}", "info" if ok else "danger")
        return redirect(url_for('admin_config_email'))
    cfg = obter_config_smtp()
    return render_template('admin/config_email.html', cfg=cfg)


# â”€â”€â”€ Auto-cadastro â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route('/api/verificar-cpf')
@master_required
def api_verificar_cpf():
    """AJAX: verifica CPF e retorna dados para o formulário de cadastro."""
    cpf = request.args.get('cpf','').strip()
    db  = get_db()

    # Já tem conta?
    u = db.execute("SELECT ativo FROM usuarios WHERE cpf=?", (cpf,)).fetchone()
    if u:
        msg = "Você já possui acesso. Faça login." if u['ativo'] else "Seu acesso está desativado. Contate o RH."
        return jsonify({"ok": False, "msg": msg})

    # Busca pre-autorização
    pre = db.execute("SELECT * FROM pre_autorizacoes WHERE cpf=?", (cpf,)).fetchone()

    # Busca servidor pelo CPF. Sem pré-autorização, o servidor pode criar conta
    # automaticamente como nível "servidor", usando a própria matrícula.
    srv = db.execute("SELECT * FROM servidores WHERE cpf=? AND arquivado=0", (cpf,)).fetchone()

    if not srv and not pre:
        return jsonify({"ok": False, "msg": "CPF não encontrado no sistema. Solicite liberação de acesso ao RH."})

    nivel     = pre['nivel']     if pre else 'servidor'
    matricula = pre['matricula'] if pre else (srv['matricula'] if srv else '')
    nome      = srv['nome']      if srv else ''

    # Vinculos (multiplos setores/secretarias)
    vinculos = get_vinculos(pre) if pre else []

    nivel_label = {'master':'Master (RH)','secretario':'Secretario',
                   'chefia':'Chefia Imediata','servidor':'Servidor'}.get(nivel, nivel)

    aviso = '' if pre else 'Nenhuma pre-autorizacao. Acesso como Servidor (padrao).'

    return jsonify({'ok': True, 'nome': nome, 'nivel': nivel, 'nivel_label': nivel_label,
                    'vinculos': vinculos, 'matricula': matricula, 'aviso': aviso})


@app.route('/criar-conta', methods=['GET','POST'])
@master_required
def criar_conta():
    if request.method == 'POST':
        cpf       = request.form['cpf'].strip()
        nome_conf = request.form.get('nome_confirmado','').strip()
        email     = request.form['email'].strip()
        senha     = request.form['senha']
        conf      = request.form['confirmar']
        nivel     = request.form.get('nivel','servidor')
        vinculos  = request.form.getlist('vinculos')
        matricula = request.form.get('matricula','').strip()

        db = get_db()

        if db.execute('SELECT 1 FROM usuarios WHERE cpf=?', (cpf,)).fetchone():
            flash('CPF ja possui acesso. Faca login.', 'warning')
            return redirect(url_for('login'))
        if senha != conf:
            flash('As senhas nao coincidem.', 'danger')
            return render_template('criar_conta.html')
        if len(senha) < 8:
            flash('Minimo 8 caracteres.', 'danger')
            return render_template('criar_conta.html')

        pre = db.execute("SELECT * FROM pre_autorizacoes WHERE cpf=?", (cpf,)).fetchone()
        srv = db.execute("SELECT * FROM servidores WHERE cpf=? AND arquivado=0", (cpf,)).fetchone()
        if not pre and not srv:
            flash('CPF não encontrado no cadastro de servidores. Contate o RH.', 'danger')
            return render_template('criar_conta.html')

        # O backend não confia nos campos hidden: reaplica a regra oficial.
        if pre:
            nivel = pre['nivel'] or 'servidor'
            vinculos = get_vinculos(pre)
            matricula = pre['matricula'] or (srv['matricula'] if srv else '')
        else:
            nivel = 'servidor'
            vinculos = []
            matricula = srv['matricula']

        nome = nome_conf or (srv['nome'] if srv else cpf)

        vj = json.dumps(vinculos)
        db.execute('''INSERT INTO usuarios (cpf,nome,email,senha_hash,nivel,matricula,vinculos,ativo,senha_temporaria)
                      VALUES (?,?,?,?,?,?,?,1,0)''',
                   (cpf, nome, email, generate_password_hash(senha), nivel, matricula, vj))
        db.commit()
        flash(f'Conta criada! Bem-vindo(a), {nome}.', 'success')
        return redirect(url_for('login'))

    return render_template('criar_conta.html')
# â”€â”€â”€ Admin: Gerenciamento de Acessos â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route('/admin/acessos')
@master_required
def admin_acessos():
    db = get_db()
    filtro_nivel = request.args.get('nivel','').strip()
    busca = request.args.get('busca','').strip()
    departamento = request.args.get('departamento','').strip()

    # Usuários por nível
    q = "WHERE u.ativo=1 AND (u.nivel='master' OR s.matricula IS NOT NULL)"
    params = []
    if filtro_nivel:
        q += " AND u.nivel=?"; params.append(filtro_nivel)
    if busca:
        q += " AND (u.nome LIKE ? OR u.cpf LIKE ? OR u.email LIKE ? OR u.matricula LIKE ? OR s.nome LIKE ?)"
        like = f"%{busca}%"; params.extend([like, like, like, like, like])
    if departamento:
        q += " AND (u.setor LIKE ? OR u.vinculos LIKE ? OR s.setor LIKE ?)"
        like_dep = f"%{departamento}%"; params.extend([like_dep, like_dep, like_dep])
    usuarios_rows = db.execute(f"""
        SELECT u.*, s.nome AS servidor_nome
        FROM usuarios u
        LEFT JOIN servidores s ON s.arquivado=0
          AND ((u.matricula IS NOT NULL AND u.matricula!='' AND s.matricula=u.matricula)
               OR (u.cpf IS NOT NULL AND u.cpf!='' AND s.cpf=u.cpf))
        {q}
        ORDER BY u.nivel, u.nome
    """, params).fetchall()
    usuarios = []
    for u in usuarios_rows:
        d = dict(u)
        d['vinculos_lista'] = get_vinculos(u)
        d['vinculos_json'] = json.dumps(d['vinculos_lista'])
        usuarios.append(d)

    # Contadores por nível
    contadores = {r['nivel']: r['qtd'] for r in db.execute("""
        SELECT u.nivel, COUNT(DISTINCT u.id) AS qtd
        FROM usuarios u
        LEFT JOIN servidores s ON s.arquivado=0
          AND ((u.matricula IS NOT NULL AND u.matricula!='' AND s.matricula=u.matricula)
               OR (u.cpf IS NOT NULL AND u.cpf!='' AND s.cpf=u.cpf))
        WHERE u.ativo=1 AND (u.nivel='master' OR s.matricula IS NOT NULL)
        GROUP BY u.nivel
    """).fetchall()}

    # Pré-autorizações pendentes (não cadastradas ainda)
    pre_where = """WHERE p.cpf NOT IN (SELECT cpf FROM usuarios)
                   AND EXISTS (SELECT 1 FROM servidores sx WHERE sx.arquivado=0 AND sx.cpf=p.cpf)"""
    pre_params = []
    if filtro_nivel:
        pre_where += " AND p.nivel=?"; pre_params.append(filtro_nivel)
    if busca:
        pre_where += " AND (p.cpf LIKE ? OR COALESCE((SELECT nome FROM servidores WHERE arquivado=0 AND cpf=p.cpf LIMIT 1),'') LIKE ?)"
        like = f"%{busca}%"; pre_params.extend([like, like])
    if departamento:
        pre_where += " AND (p.setor LIKE ? OR p.vinculos LIKE ?)"
        like_dep = f"%{departamento}%"; pre_params.extend([like_dep, like_dep])
    pre_rows = db.execute(f"""
        SELECT p.*,
               (SELECT nome FROM servidores WHERE arquivado=0 AND cpf=p.cpf LIMIT 1) AS nome_servidor
        FROM pre_autorizacoes p
        {pre_where}
        ORDER BY p.criado_em DESC
    """, pre_params).fetchall()
    pre = []
    for p in pre_rows:
        d = dict(p)
        d['vinculos_lista'] = get_vinculos(p)
        d['vinculos_json'] = json.dumps(d['vinculos_lista'])
        pre.append(d)

    # Log de últimos acessos
    ultimos = db.execute("""
        SELECT u.nome,u.cpf,u.nivel,u.ultimo_acesso
        FROM usuarios u
        LEFT JOIN servidores s ON s.arquivado=0
          AND ((u.matricula IS NOT NULL AND u.matricula!='' AND s.matricula=u.matricula)
               OR (u.cpf IS NOT NULL AND u.cpf!='' AND s.cpf=u.cpf))
        WHERE u.ultimo_acesso IS NOT NULL
          AND u.ativo=1
          AND (u.nivel='master' OR s.matricula IS NOT NULL)
        ORDER BY u.ultimo_acesso DESC LIMIT 10
    """).fetchall()

    secs = _cadastros_nomes(db, "secretaria")
    sets = _cadastros_nomes(db, "departamento")
    srvs = db.execute("SELECT matricula,nome,cpf FROM servidores WHERE arquivado=0 ORDER BY nome").fetchall()

    # ── Painel de vínculos ────────────────────────────────────────────────────
    # Carrega todos os usuários ativos de nível chefia/secretario com seus vínculos
    usuarios_vinc = db.execute(
        "SELECT id, nome, nivel, vinculos FROM usuarios WHERE ativo=1 AND nivel IN ('chefia','secretario')"
    ).fetchall()
    # mapa: setor → lista de chefias; secretaria → lista de secretarios
    _mapa_chefia = {}
    _mapa_secretario = {}
    for u in usuarios_vinc:
        try:
            vs = json.loads(u['vinculos'] or '[]')
        except Exception:
            vs = []
        entry = {'id': u['id'], 'nome': u['nome']}
        if u['nivel'] == 'chefia':
            for v in vs:
                _mapa_chefia.setdefault(v, []).append(entry)
        else:
            for v in vs:
                _mapa_secretario.setdefault(v, []).append(entry)

    # Para cada setor, descobre a secretaria pai (pelo cadastro de servidores)
    _setor_secretaria = {}
    for row in db.execute(
        "SELECT DISTINCT setor, secretaria FROM servidores WHERE setor IS NOT NULL AND setor!='' AND arquivado=0"
    ).fetchall():
        if row['setor'] and row['setor'] not in _setor_secretaria and row['secretaria']:
            _setor_secretaria[row['setor']] = row['secretaria']

    # Conta servidores por setor (para exibir no painel)
    _servidores_por_setor = {}
    for row in db.execute(
        "SELECT setor, COUNT(*) AS qtd FROM servidores WHERE setor IS NOT NULL AND setor!='' AND arquivado=0 GROUP BY setor"
    ).fetchall():
        _servidores_por_setor[row['setor']] = row['qtd']

    # Conta servidores por secretaria
    _servidores_por_sec = {}
    for row in db.execute(
        "SELECT secretaria, COUNT(*) AS qtd FROM servidores WHERE secretaria IS NOT NULL AND secretaria!='' AND arquivado=0 GROUP BY secretaria"
    ).fetchall():
        _servidores_por_sec[row['secretaria']] = row['qtd']

    painel_secretarias = []
    for sec in sorted(secs):
        painel_secretarias.append({
            'nome': sec,
            'secretarios': _mapa_secretario.get(sec, []),
            'qtd_servidores': _servidores_por_sec.get(sec, 0),
            'ok': bool(_mapa_secretario.get(sec)),
        })

    painel_setores = []
    for s in sorted(sets):
        secretaria_pai = _setor_secretaria.get(s, '')
        painel_setores.append({
            'nome': s,
            'secretaria': secretaria_pai,
            'chefias': _mapa_chefia.get(s, []),
            'qtd_servidores': _servidores_por_setor.get(s, 0),
            'sec_ok': bool(_mapa_secretario.get(secretaria_pai)) if secretaria_pai else False,
            'ok': bool(_mapa_chefia.get(s)),
        })
    # ──────────────────────────────────────────────────────────────────────────

    return render_template('admin/acessos.html',
                            usuarios=usuarios, contadores=contadores,
                            pre_autorizacoes=pre, ultimos=ultimos,
                            secretarias=secs, setores=sets, servidores=srvs,
                            busca=busca, filtro_nivel=filtro_nivel, departamento=departamento,
                            painel_secretarias=painel_secretarias,
                            painel_setores=painel_setores)


@app.route('/admin/acessos/pre/novo', methods=['POST'])
@master_required
def admin_nova_pre_autorizacao():
    db  = get_db()
    mat = request.form.get('matricula','').strip()
    if not mat:
        flash('Informe a matrícula do servidor.', 'danger')
        return redirect(url_for('admin_acessos'))

    # Busca o servidor pela matrícula para obter o CPF
    srv = db.execute("SELECT nome, cpf FROM servidores WHERE matricula=? AND arquivado=0", (mat,)).fetchone()
    if not srv:
        flash(f'Matrícula {mat} não encontrada no cadastro de servidores.', 'danger')
        return redirect(url_for('admin_acessos'))

    cpf = (srv['cpf'] or '').strip()
    if not cpf:
        flash(f'O servidor {srv["nome"]} não possui CPF cadastrado. Atualize o cadastro antes de pré-autorizar.', 'warning')
        return redirect(url_for('admin_acessos'))

    if db.execute('SELECT 1 FROM usuarios WHERE cpf=? AND ativo=1', (cpf,)).fetchone():
        flash(f'{srv["nome"]} já possui conta ativa no sistema.', 'warning')
        return redirect(url_for('admin_acessos'))

    nivel    = request.form.get('nivel','servidor')
    vinculos = request.form.getlist('vinculos')
    obs      = request.form.get('obs','').strip()
    vj = json.dumps(vinculos)
    db.upsert(
        '''INSERT OR REPLACE INTO pre_autorizacoes (cpf,nivel,matricula,obs,vinculos)
           VALUES (?,?,?,?,?)''',
        '''INSERT INTO pre_autorizacoes (cpf,nivel,matricula,obs,vinculos)
           VALUES (?,?,?,?,?)
           ON CONFLICT (cpf) DO UPDATE SET
             nivel=EXCLUDED.nivel,
             matricula=EXCLUDED.matricula,
             obs=EXCLUDED.obs,
             vinculos=EXCLUDED.vinculos''',
        (cpf, nivel, mat, obs, vj)
    )
    db.commit()
    flash(f'Pré-autorização criada para {srv["nome"]} (matrícula {mat}).', 'success')
    return redirect(url_for('admin_acessos'))


@app.route('/admin/acessos/pre/<int:pid>/excluir', methods=['POST'])
@master_required
def admin_excluir_pre(pid):
    db = get_db()
    db.execute("DELETE FROM pre_autorizacoes WHERE id=?", (pid,))
    db.commit()
    flash("Pré-autorização removida.", "warning")
    return redirect(url_for('admin_acessos'))


@app.route('/admin/acessos/usuario/<int:uid>/alterar-nivel', methods=['POST'])
@master_required
def admin_alterar_nivel(uid):
    db       = get_db()
    nivel    = request.form['nivel']
    vinculos = request.form.getlist('vinculos')
    mat      = request.form.get('matricula','').strip()
    vj = json.dumps(vinculos)
    db.execute('UPDATE usuarios SET nivel=?,matricula=?,vinculos=? WHERE id=?',
               (nivel, mat, vj, uid))
    db.commit()
    flash('Nivel de acesso atualizado.', 'success')
    return redirect(url_for('admin_acessos'))


@app.route('/admin/acessos/usuario/<int:uid>/revogar', methods=['POST'])
@master_required
def admin_revogar_acesso(uid):
    db = get_db()
    u  = db.execute("SELECT nome FROM usuarios WHERE id=?", (uid,)).fetchone()
    if u:
        db.execute("UPDATE usuarios SET ativo=0 WHERE id=?", (uid,))
        db.commit()
        flash(f"Acesso de '{u['nome']}' revogado.", "danger")
    return redirect(url_for('admin_acessos'))


# ─── Banco de Dias de Eleição ─────────────────────────────────────────────────

@app.route('/eleicao')
@login_required
def eleicao_index():
    db    = get_db()
    nivel = session.get('nivel')
    busca     = request.args.get('busca', '').strip()
    sec_sel   = request.args.get('secretaria', '').strip()
    set_sel   = request.args.get('setor', '').strip()
    saldo_sel = request.args.get('saldo', '').strip()

    # Restrição de vínculo para secretario/chefia
    vinculos = session.get('vinculos', []) if nivel in ('secretario', 'chefia') else []
    matricula_propria = session.get('matricula') or ''

    filtro = "WHERE s.arquivado=0"
    params = []

    if nivel in ('secretario', 'chefia'):
        acesso_sql, acesso_params = filtro_consulta_vinculos(nivel, vinculos, matricula_propria)
        filtro += acesso_sql
        params.extend(acesso_params)
    else:
        # master: só mostra quem tem registro de eleição
        filtro += """ AND (EXISTS (SELECT 1 FROM eleicao_creditos WHERE matricula=s.matricula)
                           OR EXISTS (SELECT 1 FROM eleicao_baixas WHERE matricula=s.matricula))"""

    if busca:
        filtro += " AND (s.matricula LIKE ? OR s.nome LIKE ?)"; params += [f"%{busca}%", f"%{busca}%"]
    if sec_sel:
        filtro += " AND s.secretaria=?"; params.append(sec_sel)
    if set_sel:
        filtro += " AND s.setor=?"; params.append(set_sel)

    servidores = db.execute(f"""
        SELECT s.matricula, s.nome, s.secretaria, s.setor, s.cargo,
               COALESCE((SELECT SUM(quantidade_dias) FROM eleicao_creditos WHERE matricula=s.matricula), 0) AS total_creditos,
               (SELECT COUNT(*) FROM eleicao_baixas WHERE matricula=s.matricula) AS total_baixas
        FROM servidores s {filtro} ORDER BY s.nome
    """, params).fetchall()

    if saldo_sel == 'com_saldo':
        servidores = [s for s in servidores if (int(s['total_creditos'] or 0) - int(s['total_baixas'] or 0)) > 0]
    elif saldo_sel == 'zerado':
        servidores = [s for s in servidores if (int(s['total_creditos'] or 0) - int(s['total_baixas'] or 0)) <= 0]

    # Listas de filtro
    base_filtro = "WHERE s.arquivado=0"
    base_params = []
    if nivel in ('secretario', 'chefia'):
        acesso_base_sql, acesso_base_params = filtro_consulta_vinculos(nivel, vinculos, matricula_propria)
        base_filtro += acesso_base_sql
        base_params.extend(acesso_base_params)
    secs  = [r[0] for r in db.execute(f"SELECT DISTINCT s.secretaria FROM servidores s {base_filtro} AND s.secretaria IS NOT NULL AND s.secretaria!='' ORDER BY s.secretaria", base_params).fetchall()]
    sets_ = [r[0] for r in db.execute(f"SELECT DISTINCT s.setor FROM servidores s {base_filtro} AND s.setor IS NOT NULL AND s.setor!='' ORDER BY s.setor", base_params).fetchall()]

    # Para modal de lançamento rápido (master apenas): lista todos os servidores ativos
    todos_servidores = []
    if nivel == 'master':
        todos_servidores = db.execute(
            "SELECT matricula, nome, secretaria, setor FROM servidores WHERE arquivado=0 ORDER BY nome"
        ).fetchall()

    return render_template('eleicao_index.html', servidores=servidores,
                           busca=busca, sec_sel=sec_sel, set_sel=set_sel, saldo_sel=saldo_sel,
                           secretarias=secs, setores=sets_,
                           todos_servidores=todos_servidores,
                           nivel=nivel)


@app.route('/eleicao/<matricula>', methods=['GET', 'POST'])
@login_required
def eleicao_servidor(matricula):
    db    = get_db()
    nivel = session.get('nivel')

    # Servidor só pode ver seus próprios dados
    if nivel == 'servidor' and session.get('matricula') != matricula:
        return redirect(url_for('meu_banco'))

    if not usuario_pode_ver_matricula(db, matricula):
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for('portal'))

    srv = db.execute("SELECT * FROM servidores WHERE matricula=? AND arquivado=0", (matricula,)).fetchone()
    if not srv:
        flash("Servidor não encontrado.", "danger")
        return redirect(url_for('eleicao_index'))

    if request.method == 'POST' and nivel != 'master':
        flash("Apenas o RH pode realizar lançamentos.", "danger")
        return redirect(url_for('eleicao_servidor', matricula=matricula))

    if request.method == 'POST':
        acao = request.form.get('acao')

        if acao == 'add_credito':
            ref = request.form.get('referencia_eleicao', '').strip()
            obs = request.form.get('observacao', '').strip()
            try:
                qtd = int(request.form.get('quantidade_dias', 0))
            except (ValueError, TypeError):
                qtd = 0
            if not ref:
                flash("Informe a referência da eleição.", "danger")
            elif qtd <= 0:
                flash("A quantidade de dias deve ser maior que zero.", "danger")
            else:
                db.insert("""
                    INSERT INTO eleicao_creditos (matricula, referencia_eleicao, quantidade_dias, observacao, criado_por)
                    VALUES (?,?,?,?,?)
                """, (matricula, ref, qtd, obs or None, session.get('nome')))
                db.commit()
                registrar_auditoria(db, "ELEICAO_CREDITO_ADD", "eleicao_creditos",
                                    matricula=matricula, servidor_nome=srv['nome'],
                                    detalhe=f"{qtd} dia(s) – {ref}")
                db.commit()
                flash(f"{srv['nome']} ({matricula})\n{qtd} dia(s) de eleição creditado(s) — {ref}", "success")

        elif acao == 'add_baixa':
            data_baixa = request.form.get('data', '').strip()
            obs = request.form.get('observacao', '').strip()
            if not data_baixa:
                flash("Informe a data da folga.", "danger")
            elif calcular_saldo_eleicao(db, matricula) <= 0:
                flash("Saldo insuficiente de dias de eleição.", "danger")
            else:
                db.insert("""
                    INSERT INTO eleicao_baixas (matricula, data, observacao, criado_por)
                    VALUES (?,?,?,?)
                """, (matricula, data_baixa, obs or None, session.get('nome')))
                db.commit()
                registrar_auditoria(db, "ELEICAO_BAIXA_ADD", "eleicao_baixas",
                                    matricula=matricula, servidor_nome=srv['nome'],
                                    detalhe=f"Folga em {data_baixa}")
                db.commit()
                flash(f"{srv['nome']} ({matricula})\nDia de folga eleitoral registrado em {data_baixa}.", "success")

        return redirect(url_for('eleicao_servidor', matricula=matricula))

    creditos = db.execute(
        "SELECT * FROM eleicao_creditos WHERE matricula=? ORDER BY criado_em DESC", (matricula,)).fetchall()
    baixas = db.execute(
        "SELECT * FROM eleicao_baixas WHERE matricula=? ORDER BY data DESC", (matricula,)).fetchall()
    saldo = calcular_saldo_eleicao(db, matricula)
    total_creditos = sum(int(c['quantidade_dias'] or 0) for c in creditos)
    return render_template('eleicao_servidor.html', servidor=srv,
                           creditos=creditos, baixas=baixas,
                           saldo=saldo, total_creditos=total_creditos,
                           total_baixas=len(baixas),
                           nivel=nivel)


@app.route('/eleicao/<matricula>/credito/<int:cid>/excluir', methods=['POST'])
@master_required
def eleicao_excluir_credito(matricula, cid):
    db = get_db()
    srv = db.execute("SELECT nome FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    credito = db.execute("SELECT * FROM eleicao_creditos WHERE id=? AND matricula=?", (cid, matricula)).fetchone()
    if not credito:
        flash("Crédito não encontrado.", "danger")
        return redirect(url_for('eleicao_servidor', matricula=matricula))
    total_dias = int(credito['quantidade_dias'])
    saldo_atual = calcular_saldo_eleicao(db, matricula)
    if saldo_atual - total_dias < 0:
        baixas_count = db.execute("SELECT COUNT(*) FROM eleicao_baixas WHERE matricula=?", (matricula,)).fetchone()[0]
        outros_creditos = db.execute(
            "SELECT COALESCE(SUM(quantidade_dias),0) FROM eleicao_creditos WHERE matricula=? AND id!=?",
            (matricula, cid)).fetchone()[0]
        if int(outros_creditos or 0) < int(baixas_count or 0):
            flash("Não é possível excluir: o saldo ficaria negativo. Exclua folgas primeiro.", "warning")
            return redirect(url_for('eleicao_servidor', matricula=matricula))
    db.execute("DELETE FROM eleicao_creditos WHERE id=? AND matricula=?", (cid, matricula))
    db.commit()
    registrar_auditoria(db, "ELEICAO_CREDITO_DEL", "eleicao_creditos",
                        matricula=matricula, servidor_nome=srv['nome'] if srv else None,
                        detalhe=f"{total_dias} dia(s) – {credito['referencia_eleicao']}")
    db.commit()
    flash("Crédito de eleição excluído.", "warning")
    return redirect(url_for('eleicao_servidor', matricula=matricula))


@app.route('/eleicao/<matricula>/credito/<int:cid>/editar', methods=['POST'])
@master_required
def eleicao_editar_credito(matricula, cid):
    db     = get_db()
    srv    = db.execute("SELECT nome FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    credito= db.execute("SELECT * FROM eleicao_creditos WHERE id=? AND matricula=?", (cid, matricula)).fetchone()
    if not credito:
        flash("Crédito não encontrado.", "danger")
        return redirect(url_for('eleicao_servidor', matricula=matricula))
    ref = request.form.get('referencia_eleicao', '').strip()
    obs = request.form.get('observacao', '').strip()
    try:
        qtd = int(request.form.get('quantidade_dias', 0))
    except (ValueError, TypeError):
        qtd = 0
    if not ref or qtd <= 0:
        flash("Preencha corretamente os campos.", "danger")
        return redirect(url_for('eleicao_servidor', matricula=matricula))
    # Valida saldo: nova quantidade vs baixas existentes
    outras_creds = db.execute(
        "SELECT COALESCE(SUM(quantidade_dias),0) FROM eleicao_creditos WHERE matricula=? AND id!=?",
        (matricula, cid)).fetchone()[0]
    baixas_count = db.execute("SELECT COUNT(*) FROM eleicao_baixas WHERE matricula=?", (matricula,)).fetchone()[0]
    if int(outras_creds or 0) + qtd < int(baixas_count or 0):
        flash("A nova quantidade deixaria o saldo negativo. Reduza as folgas primeiro.", "warning")
        return redirect(url_for('eleicao_servidor', matricula=matricula))
    db.execute(
        "UPDATE eleicao_creditos SET referencia_eleicao=?, quantidade_dias=?, observacao=? WHERE id=? AND matricula=?",
        (ref, qtd, obs or None, cid, matricula))
    db.commit()
    registrar_auditoria(db, "ELEICAO_CREDITO_EDIT", "eleicao_creditos",
                        matricula=matricula, servidor_nome=srv['nome'] if srv else None,
                        detalhe=f"{qtd} dia(s) – {ref}")
    db.commit()
    flash("Crédito atualizado com sucesso.", "success")
    return redirect(url_for('eleicao_servidor', matricula=matricula))


@app.route('/eleicao/<matricula>/exportar/<fmt>')
@login_required
def eleicao_exportar(matricula, fmt):
    db  = get_db()
    if not usuario_pode_ver_matricula(db, matricula):
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for('portal'))
    srv = db.execute("SELECT * FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    if not srv:
        flash("Servidor não encontrado.", "danger")
        return redirect(url_for('eleicao_index'))

    creditos = db.execute(
        "SELECT referencia_eleicao, quantidade_dias, observacao, criado_por, criado_em "
        "FROM eleicao_creditos WHERE matricula=? ORDER BY criado_em", (matricula,)).fetchall()
    baixas = db.execute(
        "SELECT data, observacao, criado_por, criado_em "
        "FROM eleicao_baixas WHERE matricula=? ORDER BY data", (matricula,)).fetchall()

    saldo = calcular_saldo_eleicao(db, matricula)
    total_cred = sum(int(c['quantidade_dias'] or 0) for c in creditos)
    total_baixas = len(baixas)

    title = (f"Banco de Dias de Eleição\n"
             f"{srv['nome']} – Matrícula {srv['matricula']}\n"
             f"Saldo: {saldo} dia(s)")
    filename = f"dias_eleicao_{matricula}"

    h_cred = ["Eleição / Referência", "Dias Creditados", "Trabalho Realizado", "Lançado por", "Data Lançamento"]
    rows_cred = [
        [c['referencia_eleicao'], c['quantidade_dias'], c['observacao'] or '', c['criado_por'] or '', (c['criado_em'] or '')[:10]]
        for c in creditos
    ]
    rows_cred.append(["TOTAL CRÉDITOS", total_cred, "", "", ""])

    h_baixas = ["Data da Folga", "Observação", "Registrado por", "Data Registro"]
    rows_baixas = [
        [b['data'], b['observacao'] or '', b['criado_por'] or '', (b['criado_em'] or '')[:10]]
        for b in baixas
    ]
    rows_baixas.append(["TOTAL FOLGAS", total_baixas, "", ""])

    if fmt == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        AMBER  = "78350F"
        AMBER2 = "D97706"
        LIGHT  = "FEF3C7"

        def _sheet(ws, titulo, headers, rows, totals_row=True):
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            c = ws.cell(1, 1, titulo)
            c.font = Font(bold=True, color="FFFFFF", size=13)
            c.fill = PatternFill("solid", fgColor=AMBER)
            c.alignment = Alignment(horizontal="center")
            ws.append([])
            ws.append(headers)
            for col_c in ws[3]:
                col_c.font = Font(bold=True, color="FFFFFF")
                col_c.fill = PatternFill("solid", fgColor=AMBER2)
                col_c.alignment = Alignment(horizontal="center")
            for row in rows:
                ws.append([_safe_excel(v) for v in row])
                label = str(row[0]) if row else ""
                if label.startswith("TOTAL"):
                    for col_c in ws[ws.max_row]:
                        col_c.font = Font(bold=True)
                        col_c.fill = PatternFill("solid", fgColor=LIGHT)
            for col in range(1, len(headers) + 1):
                mx = max(len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1))
                ws.column_dimensions[get_column_letter(col)].width = min(max(mx + 2, 12), 50)

        ws1 = wb.active
        ws1.title = "Créditos"
        _sheet(ws1, f"Créditos de Eleição — {srv['nome']} ({matricula})", h_cred, rows_cred)
        ws2 = wb.create_sheet("Folgas Tiradas")
        _sheet(ws2, f"Folgas Tiradas — {srv['nome']} ({matricula})", h_baixas, rows_baixas)
        ws3 = wb.create_sheet("Resumo")
        ws3.append(["Servidor", srv['nome']])
        ws3.append(["Matrícula", srv['matricula']])
        ws3.append(["Secretaria", srv.get('secretaria') or ''])
        ws3.append(["Departamento", srv.get('setor') or ''])
        ws3.append(["Total dias creditados", total_cred])
        ws3.append(["Total folgas tiradas", total_baixas])
        ws3.append(["Saldo atual (dias)", saldo])
        for r in ws3.iter_rows():
            r[0].font = Font(bold=True)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        resp = make_response(buf.getvalue())
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        resp.headers["Content-Disposition"] = f"attachment; filename={filename}.xlsx"
        return resp

    elif fmt == 'pdf':
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from xml.sax.saxutils import escape

        AMBER_RL = colors.HexColor("#78350F")
        AMBER2_RL = colors.HexColor("#D97706")
        LIGHT_RL  = colors.HexColor("#FEF3C7")

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        title_style  = ParagraphStyle("T", parent=styles["Title"], fontSize=15, leading=18, alignment=1, textColor=AMBER_RL)
        sub_style    = ParagraphStyle("S", parent=styles["Normal"], fontSize=9, leading=12, alignment=1, textColor=colors.HexColor("#6B7280"))
        section_style= ParagraphStyle("Sec", parent=styles["Heading2"], fontSize=10, leading=13, textColor=AMBER_RL)
        cell_s = ParagraphStyle("C", parent=styles["BodyText"], fontSize=7, leading=9, wordWrap="CJK")
        hdr_s  = ParagraphStyle("H", parent=cell_s, fontName="Helvetica-Bold", textColor=colors.white, alignment=1)

        def make_table(headers, rows, col_ratios=None):
            def pcell(v, s=cell_s): return Paragraph(escape(str(v or "")), s)
            data = [[pcell(h, hdr_s) for h in headers]] + [[pcell(v) for v in row] for row in rows]
            if not col_ratios:
                col_ratios = [1] * len(headers)
            total_r = sum(col_ratios)
            col_w = [doc.width * r / total_r for r in col_ratios]
            t = Table(data, colWidths=col_w, repeatRows=1)
            estilo = [
                ("BACKGROUND", (0, 0), (-1, 0), AMBER2_RL),
                ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID",       (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D7DE")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFFBEB")]),
                ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING",(0, 0), (-1, -1), 4),
                ("TOPPADDING",  (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
            ]
            for i, row in enumerate(rows, 1):
                if str(row[0]).startswith("TOTAL"):
                    estilo += [("BACKGROUND", (0, i), (-1, i), LIGHT_RL),
                               ("FONTNAME",   (0, i), (-1, i), "Helvetica-Bold")]
            t.setStyle(TableStyle(estilo))
            return t

        story = [
            Paragraph("Banco de Dias de Eleição", title_style),
            Paragraph(f"{escape(srv['nome'])} — Matrícula {srv['matricula']}"
                      + (f" | {escape(srv['secretaria'])}" if srv.get('secretaria') else "")
                      + (f" | {escape(srv['setor'])}" if srv.get('setor') else ""), sub_style),
            Spacer(1, 4),
            Paragraph(f"Saldo atual: <b>{saldo} dia(s)</b> &nbsp;|&nbsp; "
                      f"Creditados: <b>{total_cred}</b> &nbsp;|&nbsp; "
                      f"Folgas: <b>{total_baixas}</b>", sub_style),
            Spacer(1, 10),
            HRFlowable(color=AMBER_RL, thickness=1, width="100%"),
            Spacer(1, 8),
            Paragraph("Créditos de Eleição", section_style),
            Spacer(1, 4),
            make_table(h_cred, rows_cred, col_ratios=[4, 1, 4, 2, 2]),
            Spacer(1, 14),
            Paragraph("Folgas Tiradas", section_style),
            Spacer(1, 4),
            make_table(h_baixas, rows_baixas, col_ratios=[2, 4, 2, 2]),
        ]
        doc.build(story)
        buf.seek(0)
        resp = make_response(buf.getvalue())
        resp.headers["Content-Type"] = "application/pdf"
        resp.headers["Content-Disposition"] = f"attachment; filename={filename}.pdf"
        return resp

    flash("Formato inválido.", "danger")
    return redirect(url_for('eleicao_servidor', matricula=matricula))


@app.route('/eleicao/<matricula>/baixa/<int:bid>/excluir', methods=['POST'])
@master_required
def eleicao_excluir_baixa(matricula, bid):
    db = get_db()
    srv = db.execute("SELECT nome FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    baixa = db.execute("SELECT * FROM eleicao_baixas WHERE id=? AND matricula=?", (bid, matricula)).fetchone()
    if not baixa:
        flash("Folga não encontrada.", "danger")
        return redirect(url_for('eleicao_servidor', matricula=matricula))
    db.execute("DELETE FROM eleicao_baixas WHERE id=? AND matricula=?", (bid, matricula))
    db.commit()
    registrar_auditoria(db, "ELEICAO_BAIXA_DEL", "eleicao_baixas",
                        matricula=matricula, servidor_nome=srv['nome'] if srv else None,
                        detalhe=f"Folga em {baixa['data']}")
    db.commit()
    flash("Folga de eleição estornada.", "warning")
    return redirect(url_for('eleicao_servidor', matricula=matricula))


def _paginar(total, page, per_page=20):
    """Retorna (page_corrigida, total_pages, offset) para paginação server-side."""
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    return page, total_pages, (page - 1) * per_page

def _calcular_data_fim(data_pretendida, quantidade, tipo):
    """Calcula data final do período de compensação."""
    if tipo == 'eleicao' and quantidade > 1:
        try:
            d = date.fromisoformat(data_pretendida)
            return (d + timedelta(days=quantidade - 1)).isoformat()
        except Exception:
            pass
    return data_pretendida

# ─── Solicitações de Compensação ─────────────────────────────────────────────

def _solicitacoes_habilitado(db):
    r = db.execute("SELECT valor FROM config WHERE chave='solicitacoes_habilitado'").fetchone()
    return (r['valor'] if r else '1') == '1'

def _aprovadores_para_servidor(db, matricula):
    """Retorna lista de usuários (chefia + secretário) aptos a autorizar a solicitação."""
    srv = db.execute("SELECT secretaria, setor FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    if not srv:
        return []
    setor = srv['setor'] or ''
    secretaria = srv['secretaria'] or ''
    users = db.execute(
        "SELECT id, nome, nivel, vinculos FROM usuarios WHERE ativo=1 AND nivel IN ('chefia','secretario')"
    ).fetchall()
    result = []
    for u in users:
        try:
            vinculos = json.loads(u['vinculos'] or '[]')
        except Exception:
            vinculos = []
        if u['nivel'] == 'chefia' and setor and setor in vinculos:
            result.append(dict(u))
        elif u['nivel'] == 'secretario' and secretaria and secretaria in vinculos:
            result.append(dict(u))
    return result

def _secretarios_para_chefia(db, chefia_uid):
    """Retorna secretários que devem receber solicitação de uma chefia."""
    chefia = db.execute("SELECT secretaria, vinculos FROM usuarios WHERE id=?", (chefia_uid,)).fetchone()
    if not chefia:
        return []
    secretaria = (chefia['secretaria'] or '').strip()
    if not secretaria:
        try:
            vinculos = json.loads(chefia['vinculos'] or '[]')
        except Exception:
            vinculos = []
        for setor in vinculos:
            row = db.execute(
                "SELECT secretaria FROM servidores WHERE setor=? AND arquivado=0 LIMIT 1", (setor,)
            ).fetchone()
            if row and row['secretaria']:
                secretaria = row['secretaria']
                break
    if not secretaria:
        return []
    users = db.execute(
        "SELECT id, nome, nivel, vinculos FROM usuarios WHERE ativo=1 AND nivel='secretario'"
    ).fetchall()
    result = []
    for u in users:
        try:
            vinculos = json.loads(u['vinculos'] or '[]')
        except Exception:
            vinculos = []
        if secretaria in vinculos:
            result.append(dict(u))
    return result

def _pode_autorizar_solicitacao(db, sol):
    """Verifica se o usuário logado pode autorizar a solicitação."""
    nivel = session.get('nivel')
    uid = session.get('uid')
    if nivel == 'master':
        return True
    if nivel not in ('chefia', 'secretario'):
        return False
    if sol['criado_por_uid'] == uid:
        return False
    return usuario_pode_ver_matricula(db, sol['matricula'])

def _status_display(sol):
    status = sol['status']
    tipo = sol['tipo']
    if status == 'solicitado':
        return 'Horas Pendentes' if tipo == 'banco_horas' else 'Dias Pendentes'
    mapa = {
        'autorizado': 'Autorizado — Pendente de Lançamento pelo RH',
        'indeferido':  'Indeferido',
        'lancado':    'Lançado pelo RH',
        'cancelado':  'Cancelado',
        'estornado':  'Estornado pelo RH',
    }
    return mapa.get(status, status)

@app.route("/api/saldo-solicitacao/<matricula>/<tipo>")
@login_required
def api_saldo_solicitacao(matricula, tipo):
    db = get_db()
    uid = session.get('uid')
    # Valida que o usuário está consultando apenas o próprio saldo
    u_row = db.execute("SELECT matricula FROM usuarios WHERE id=?", (uid,)).fetchone()
    mat_vinculada = (u_row['matricula'] if u_row else None) or session.get('matricula')
    if mat_vinculada != matricula:
        return json.dumps({'erro': 'Acesso negado'}), 403
    if tipo == 'banco_horas':
        saldo = calcular_saldo(db, matricula)
        return json.dumps({'saldo': saldo, 'fmt': minutos_para_horas(saldo)})
    elif tipo == 'eleicao':
        saldo = calcular_saldo_eleicao(db, matricula)
        return json.dumps({'saldo': saldo, 'fmt': str(saldo)})
    return json.dumps({'erro': 'Tipo inválido'}), 400

@app.route("/solicitacoes/nova", methods=["POST"])
@login_required
def solicitacoes_nova():
    db = get_db()
    if not _solicitacoes_habilitado(db):
        return json.dumps({'erro': 'Funcionalidade desabilitada pelo RH.'}), 403, {'Content-Type': 'application/json'}
    uid  = session.get('uid')
    nome = session.get('nome')
    tipo = request.form.get('tipo', '').strip()
    data_pretendida = request.form.get('data_pretendida', '').strip()
    justificativa = request.form.get('justificativa', '').strip()

    # Determina matricula vinculada ao usuário logado
    u_row = db.execute("SELECT matricula FROM usuarios WHERE id=?", (uid,)).fetchone()
    matricula = (u_row['matricula'] if u_row else None) or session.get('matricula') or ''
    if not matricula:
        return json.dumps({'erro': 'Não foi localizado cadastro de servidor vinculado ao usuário logado.'}), 400, {'Content-Type': 'application/json'}

    if not tipo or not data_pretendida:
        return json.dumps({'erro': 'Preencha todos os campos obrigatórios.'}), 400, {'Content-Type': 'application/json'}
    if tipo not in ('banco_horas', 'eleicao'):
        return json.dumps({'erro': 'Tipo de solicitação inválido.'}), 400, {'Content-Type': 'application/json'}

    srv = db.execute("SELECT nome FROM servidores WHERE matricula=? AND arquivado=0", (matricula,)).fetchone()
    if not srv:
        return json.dumps({'erro': 'Não foi localizado cadastro de servidor vinculado ao usuário logado.'}), 404, {'Content-Type': 'application/json'}

    if tipo == 'banco_horas':
        horas_txt = request.form.get('quantidade', '').strip()
        quantidade = horas_para_minutos(horas_txt)
        if quantidade <= 0:
            return json.dumps({'erro': 'Informe uma quantidade válida de horas.'}), 400, {'Content-Type': 'application/json'}
        saldo_disponivel = calcular_saldo(db, matricula)
        if quantidade > saldo_disponivel:
            return json.dumps({'erro': 'Quantidade de horas solicitada superior ao saldo disponível.'}), 400, {'Content-Type': 'application/json'}
        data_fim = data_pretendida
    else:
        try:
            quantidade = int(request.form.get('quantidade', '0'))
        except (ValueError, TypeError):
            quantidade = 0
        if quantidade <= 0:
            return json.dumps({'erro': 'Informe uma quantidade válida de dias.'}), 400, {'Content-Type': 'application/json'}
        saldo_disponivel = calcular_saldo_eleicao(db, matricula)
        if quantidade > saldo_disponivel:
            return json.dumps({'erro': 'Quantidade de dias solicitada superior ao saldo disponível.'}), 400, {'Content-Type': 'application/json'}
        data_fim = _calcular_data_fim(data_pretendida, quantidade, tipo)

    sid = db.insert("""
        INSERT INTO solicitacoes (matricula, tipo, quantidade, data_pretendida, data_fim, justificativa, status, criado_por_uid, criado_por_nome)
        VALUES (?,?,?,?,?,?,?,?,?)
    """, (matricula, tipo, quantidade, data_pretendida, data_fim, justificativa or None, 'solicitado', uid, nome))
    registrar_auditoria(db, "Solicitação criada", "solicitacoes", sid, matricula, srv['nome'],
                        f"Tipo: {tipo}; Quantidade: {quantidade}; Período: {data_pretendida} a {data_fim}")
    db.commit()
    return json.dumps({'ok': True, 'id': sid}), 200, {'Content-Type': 'application/json'}

@app.route("/solicitacoes/<int:sid>/autorizar", methods=["POST"])
@login_required
def solicitacoes_autorizar(sid):
    db = get_db()
    sol = db.execute("SELECT * FROM solicitacoes WHERE id=?", (sid,)).fetchone()
    if not sol:
        return json.dumps({'erro': 'Solicitação não encontrada.'}), 404, {'Content-Type': 'application/json'}
    if sol['status'] not in ('solicitado', 'autorizado'):
        return json.dumps({'erro': 'Solicitação não pode ser autorizada neste status.'}), 400, {'Content-Type': 'application/json'}
    if not _pode_autorizar_solicitacao(db, sol):
        return json.dumps({'erro': 'Sem permissão para autorizar esta solicitação.'}), 403, {'Content-Type': 'application/json'}
    nivel = session.get('nivel')
    uid   = session.get('uid')
    nome  = session.get('nome')
    justificativa_rh = request.form.get('justificativa_rh', '').strip()
    if nivel == 'master' and not justificativa_rh:
        return json.dumps({'erro': 'Justificativa obrigatória para autorização pelo RH.'}), 400, {'Content-Type': 'application/json'}
    despacho = request.form.get('despacho_chefia', '').strip()
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    db.execute("""UPDATE solicitacoes SET status='autorizado', aprovador_uid=?, aprovador_nome=?,
                  data_autorizacao=?, justificativa_rh=?, despacho_chefia=? WHERE id=?""",
               (uid, nome, agora, justificativa_rh or None, despacho or None, sid))
    srv = db.execute("SELECT nome FROM servidores WHERE matricula=?", (sol['matricula'],)).fetchone()
    registrar_auditoria(db, "Solicitação autorizada", "solicitacoes", sid, sol['matricula'],
                        srv['nome'] if srv else None,
                        f"Autorizado por: {nome} ({nivel}){' | Despacho: ' + despacho if despacho else ''}{' | Justificativa RH: ' + justificativa_rh if justificativa_rh else ''}")
    db.commit()
    return json.dumps({'ok': True}), 200, {'Content-Type': 'application/json'}

@app.route("/solicitacoes/<int:sid>/indeferir", methods=["POST"])
@login_required
def solicitacoes_indeferir(sid):
    db = get_db()
    sol = db.execute("SELECT * FROM solicitacoes WHERE id=?", (sid,)).fetchone()
    if not sol:
        return json.dumps({'erro': 'Solicitação não encontrada.'}), 404, {'Content-Type': 'application/json'}
    if sol['status'] not in ('solicitado', 'autorizado'):
        return json.dumps({'erro': 'Solicitação não pode ser indeferida neste status.'}), 400, {'Content-Type': 'application/json'}
    if not _pode_autorizar_solicitacao(db, sol):
        return json.dumps({'erro': 'Sem permissão para indeferir esta solicitação.'}), 403, {'Content-Type': 'application/json'}
    motivo   = request.form.get('motivo', '').strip()
    despacho = request.form.get('despacho_chefia', '').strip()
    uid  = session.get('uid')
    nome = session.get('nome')
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    db.execute("""UPDATE solicitacoes SET status='indeferido', aprovador_uid=?, aprovador_nome=?,
                  data_autorizacao=?, motivo_indeferimento=?, despacho_chefia=? WHERE id=?""",
               (uid, nome, agora, motivo or None, despacho or None, sid))
    srv = db.execute("SELECT nome FROM servidores WHERE matricula=?", (sol['matricula'],)).fetchone()
    registrar_auditoria(db, "Solicitação indeferida", "solicitacoes", sid, sol['matricula'],
                        srv['nome'] if srv else None,
                        f"Indeferido por: {nome}; Motivo: {motivo or '-'}")
    db.commit()
    return json.dumps({'ok': True}), 200, {'Content-Type': 'application/json'}

@app.route("/solicitacoes/<int:sid>/cancelar", methods=["POST"])
@login_required
def solicitacoes_cancelar(sid):
    db = get_db()
    sol = db.execute("SELECT * FROM solicitacoes WHERE id=?", (sid,)).fetchone()
    if not sol:
        return json.dumps({'erro': 'Solicitação não encontrada.'}), 404, {'Content-Type': 'application/json'}
    nivel = session.get('nivel')
    uid   = session.get('uid')
    if nivel not in ('master',) and sol['criado_por_uid'] != uid:
        return json.dumps({'erro': 'Sem permissão para cancelar esta solicitação.'}), 403, {'Content-Type': 'application/json'}
    if sol['status'] in ('lancado', 'cancelado', 'indeferido'):
        return json.dumps({'erro': 'Solicitação não pode ser cancelada neste status.'}), 400, {'Content-Type': 'application/json'}
    db.execute("UPDATE solicitacoes SET status='cancelado' WHERE id=?", (sid,))
    srv = db.execute("SELECT nome FROM servidores WHERE matricula=?", (sol['matricula'],)).fetchone()
    registrar_auditoria(db, "Solicitação cancelada", "solicitacoes", sid, sol['matricula'],
                        srv['nome'] if srv else None, f"Cancelado por: {session.get('nome')}")
    db.commit()
    return json.dumps({'ok': True}), 200, {'Content-Type': 'application/json'}

@app.route("/admin/tarefas")
@master_required
def admin_tarefas():
    db = get_db()
    f_matricula  = request.args.get('matricula', '').strip()
    f_secretaria = request.args.get('secretaria', '').strip()
    f_setor      = request.args.get('setor', '').strip()
    f_tipo       = request.args.get('tipo', '').strip()
    f_status     = request.args.get('status', '').strip()
    f_dt_ini     = request.args.get('dt_ini', '').strip()
    f_dt_fim     = request.args.get('dt_fim', '').strip()

    where, params = [], []
    if f_matricula:
        where.append("(s.matricula LIKE ? OR s.nome LIKE ?)")
        params += [f"%{f_matricula}%", f"%{f_matricula}%"]
    if f_secretaria:
        where.append("s.secretaria LIKE ?"); params.append(f"%{f_secretaria}%")
    if f_setor:
        where.append("s.setor LIKE ?"); params.append(f"%{f_setor}%")
    if f_tipo:
        where.append("sol.tipo=?"); params.append(f_tipo)
    if f_status:
        where.append("sol.status=?"); params.append(f_status)
    if f_dt_ini:
        where.append("sol.criado_em >= ?"); params.append(f_dt_ini)
    if f_dt_fim:
        where.append("sol.criado_em <= ?"); params.append(f_dt_fim + " 23:59:59")

    clausula = ("WHERE " + " AND ".join(where)) if where else ""
    total = db.execute(f"""
        SELECT COUNT(*) FROM solicitacoes sol
        JOIN servidores s ON s.matricula = sol.matricula {clausula}
    """, params).fetchone()[0] or 0
    page = max(1, int(request.args.get('page', 1) or 1))
    page, total_pages, offset = _paginar(total, page)
    solicitacoes = db.execute(f"""
        SELECT sol.*, s.nome AS servidor_nome, s.secretaria, s.setor, s.cargo
        FROM solicitacoes sol
        JOIN servidores s ON s.matricula = sol.matricula
        {clausula}
        ORDER BY
            CASE sol.status WHEN 'autorizado' THEN 0 WHEN 'solicitado' THEN 1 ELSE 2 END,
            sol.criado_em DESC
        LIMIT 20 OFFSET ?
    """, params + [offset]).fetchall()

    habilitado = _solicitacoes_habilitado(db)
    secs, sets = _listas_filtro(db)
    return render_template("admin/tarefas.html",
                           solicitacoes=solicitacoes,
                           total=total, page=page, total_pages=total_pages,
                           habilitado=habilitado,
                           fmt=minutos_para_horas,
                           status_display=_status_display,
                           filtros={'matricula': f_matricula, 'secretaria': f_secretaria,
                                    'setor': f_setor, 'tipo': f_tipo, 'status': f_status,
                                    'dt_ini': f_dt_ini, 'dt_fim': f_dt_fim},
                           secretarias=secs, setores=sets)

@app.route("/admin/tarefas/<int:sid>/lancar", methods=["POST"])
@master_required
def admin_tarefas_lancar(sid):
    db = get_db()
    sol = db.execute("SELECT * FROM solicitacoes WHERE id=?", (sid,)).fetchone()
    if not sol:
        return json.dumps({'erro': 'Solicitação não encontrada.'}), 404, {'Content-Type': 'application/json'}
    if sol['status'] != 'autorizado':
        nivel = session.get('nivel')
        if nivel != 'master':
            return json.dumps({'erro': 'Solicitação precisa estar autorizada para ser lançada.'}), 400, {'Content-Type': 'application/json'}
        justificativa_rh = request.form.get('justificativa_rh', '').strip()
        if not justificativa_rh:
            return json.dumps({'erro': 'Para lançar sem autorização prévia da chefia, informe uma justificativa.'}), 400, {'Content-Type': 'application/json'}
        agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        db.execute("""UPDATE solicitacoes SET status='autorizado', aprovador_uid=?, aprovador_nome=?,
                      data_autorizacao=?, justificativa_rh=? WHERE id=?""",
                   (session.get('uid'), session.get('nome'), agora, justificativa_rh, sid))

    uid  = session.get('uid')
    nome = session.get('nome')
    matricula = sol['matricula']
    srv = db.execute("SELECT nome FROM servidores WHERE matricula=?", (matricula,)).fetchone()
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    data_lanc = sol['data_pretendida']
    ref_id = None

    if sol['tipo'] == 'banco_horas':
        saldo_real = calcular_saldo(db, matricula)
        # calcular_saldo já desconta pendentes; ao lançar, soma de volta a reserva e confere o real
        saldo_sem_esta = saldo_real + sol['quantidade']
        if saldo_sem_esta < sol['quantidade']:
            return json.dumps({'erro': 'Saldo insuficiente para lançar a compensação.'}), 400, {'Content-Type': 'application/json'}
        desc = f"Solicitação #{sid} — compensação autorizada por {sol['aprovador_nome'] or nome}"
        cid = db.insert(
            "INSERT INTO compensacoes (matricula,data,tipo,minutos_compensados,descricao) VALUES (?,?,?,?,?)",
            (matricula, data_lanc, 'parcial', sol['quantidade'], desc))
        _consumir_fifo_raw(db, matricula, sol['quantidade'], "compensacao", cid)
        ref_id = cid
    else:
        saldo_real = calcular_saldo_eleicao(db, matricula)
        saldo_sem_esta = saldo_real + sol['quantidade']
        if saldo_sem_esta < sol['quantidade']:
            return json.dumps({'erro': 'Saldo de dias insuficiente para lançar a fruição.'}), 400, {'Content-Type': 'application/json'}
        desc = f"Solicitação #{sid} — fruição autorizada por {sol['aprovador_nome'] or nome}"
        bid = db.insert(
            "INSERT INTO eleicao_baixas (matricula,data,observacao,criado_por) VALUES (?,?,?,?)",
            (matricula, data_lanc, desc, nome))
        ref_id = bid

    db.execute("""UPDATE solicitacoes SET status='lancado', rh_uid=?, rh_nome=?,
                  data_lancamento=?, referencia_id=? WHERE id=?""",
               (uid, nome, agora, ref_id, sid))
    registrar_auditoria(db, "Solicitação lançada pelo RH", "solicitacoes", sid, matricula,
                        srv['nome'] if srv else None,
                        f"Tipo: {sol['tipo']}; Ref lançamento: {ref_id}")
    db.commit()
    return json.dumps({'ok': True}), 200, {'Content-Type': 'application/json'}

@app.route("/admin/tarefas/toggle", methods=["POST"])
@master_required
def admin_tarefas_toggle():
    db = get_db()
    atual = _solicitacoes_habilitado(db)
    novo  = '0' if atual else '1'
    db.upsert(
        "INSERT OR REPLACE INTO config (chave,valor) VALUES (?,?)",
        "INSERT INTO config (chave,valor) VALUES (?,?) ON CONFLICT (chave) DO UPDATE SET valor=EXCLUDED.valor",
        ("solicitacoes_habilitado", novo)
    )
    db.commit()
    estado = 'habilitada' if novo == '1' else 'desabilitada'
    return json.dumps({'ok': True, 'habilitado': novo == '1', 'estado': estado}), 200, {'Content-Type': 'application/json'}

@app.route("/admin/tarefas/<int:sid>/estornar", methods=["POST"])
@master_required
def admin_tarefas_estornar(sid):
    db = get_db()
    sol = db.execute("SELECT * FROM solicitacoes WHERE id=?", (sid,)).fetchone()
    if not sol:
        return json.dumps({'erro': 'Solicitação não encontrada.'}), 404, {'Content-Type': 'application/json'}
    if sol['status'] != 'lancado':
        return json.dumps({'erro': 'Apenas solicitações já lançadas podem ser estornadas.'}), 400, {'Content-Type': 'application/json'}

    matricula = sol['matricula']
    ref_id = sol['referencia_id']
    srv = db.execute("SELECT nome FROM servidores WHERE matricula=?", (matricula,)).fetchone()

    if sol['tipo'] == 'banco_horas' and ref_id:
        # Remove consumos vinculados a essa compensação e a própria compensação
        db.execute("DELETE FROM consumos WHERE tipo='compensacao' AND referencia_id=?", (ref_id,))
        db.execute("DELETE FROM compensacoes WHERE id=?", (ref_id,))
    elif sol['tipo'] == 'eleicao' and ref_id:
        db.execute("DELETE FROM eleicao_baixas WHERE id=?", (ref_id,))

    db.execute("UPDATE solicitacoes SET status='estornado' WHERE id=?", (sid,))
    registrar_auditoria(db, "Solicitação estornada pelo RH", "solicitacoes", sid, matricula,
                        srv['nome'] if srv else None,
                        f"Tipo: {sol['tipo']}; Ref estornada: {ref_id}; Estornado por: {session.get('nome')}")
    db.commit()
    return json.dumps({'ok': True}), 200, {'Content-Type': 'application/json'}

@app.route("/api/solicitacoes-servidor")
@login_required
def api_solicitacoes_servidor():
    """Retorna solicitações do servidor logado (para exibição no meu_banco)."""
    db = get_db()
    nivel = session.get('nivel')
    uid   = session.get('uid')
    matricula = session.get('matricula')
    if nivel == 'servidor':
        rows = db.execute(
            "SELECT * FROM solicitacoes WHERE matricula=? ORDER BY criado_em DESC",
            (matricula,)).fetchall()
    elif nivel in ('chefia', 'secretario'):
        vinculos = session.get('vinculos') or []
        if not vinculos:
            return json.dumps([]), 200, {'Content-Type': 'application/json'}
        ph = ','.join('?' * len(vinculos))
        campo = 'secretaria' if nivel == 'secretario' else 'setor'
        rows = db.execute(
            f"""SELECT sol.*, s.nome AS servidor_nome FROM solicitacoes sol
                JOIN servidores s ON s.matricula=sol.matricula
                WHERE sol.status='solicitado' AND s.{campo} IN ({ph})
                AND sol.criado_por_uid != ?
                ORDER BY sol.criado_em ASC""",
            vinculos + [uid]
        ).fetchall()
    else:
        return json.dumps([]), 200, {'Content-Type': 'application/json'}
    result = []
    for r in rows:
        d = dict(r)
        d['status_display'] = _status_display(d)
        result.append(d)
    return json.dumps(result, default=str), 200, {'Content-Type': 'application/json'}


# No Render/Gunicorn o bloco __main__ não executa, então inicializamos o schema
# no import da aplicação. As migrações são idempotentes.
init_db()

if __name__ == "__main__":
    app.run(debug=True, port=5000)
